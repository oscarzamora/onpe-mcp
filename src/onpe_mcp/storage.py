from __future__ import annotations

import csv
import logging
import json
import sqlite3
import unicodedata
from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path
from typing import Any


_logger = logging.getLogger("onpe_mcp.storage")


def _norm_text(text: str) -> str:
    base = unicodedata.normalize("NFKD", text or "")
    return "".join(ch for ch in base if not unicodedata.combining(ch)).casefold().strip()


# Mapa de nombres populares de ciudades peruanas → nombres administrativos oficiales en ubigeo_reniec.
# Clave: nombre normalizado (sin tildes, minúsculas). Valor: nombre de provincia o departamento.
_CITY_ALIASES: dict[str, str] = {
    # Ucayali
    "pucallpa": "coronel portillo",
    # Loreto
    "iquitos": "maynas",
    "yurimaguas": "alto amazonas",
    "nauta": "loreto",
    # San Martín
    "tarapoto": "san martin",
    "moyobamba": "moyobamba",
    # Madre de Dios
    "puerto maldonado": "tambopata",
    # Amazonas
    "chachapoyas": "chachapoyas",
    # La Libertad
    "trujillo": "trujillo",
    # Lambayeque
    "chiclayo": "chiclayo",
    "lambayeque": "lambayeque",
    # Piura
    "piura": "piura",
    "sullana": "sullana",
    "tumbes": "tumbes",
    # Ancash
    "chimbote": "santa",
    "huaraz": "huaraz",
    # Lima
    "lima": "lima",
    "callao": "callao",
    # Ica
    "ica": "ica",
    "nazca": "nazca",
    # Arequipa
    "arequipa": "arequipa",
    # Tacna
    "tacna": "tacna",
    # Puno
    "puno": "puno",
    "juliaca": "san roman",
    # Cusco
    "cusco": "cusco",
    "cuzco": "cusco",
    # Apurimac
    "abancay": "abancay",
    "andahuaylalas": "andahuaylas",
    # Ayacucho — «ayacucho» es nombre de departamento; se resuelve por lookup RENIEC departamental.
    # No se alias: «ayacucho» → «huamanga» porque perdería el resto de provincias del dept.
    "huamanga": "huamanga",
    # Huancavelica
    "huancavelica": "huancavelica",
    # Junin
    "huancayo": "huancayo",
    # Pasco
    "cerro de pasco": "pasco",
    # Huanuco
    "huanuco": "huanuco",
    # Cajamarca
    "cajamarca": "cajamarca",
    "jaen": "jaen",
    # Lima — distritos urbanos frecuentes
    "la victoria": "la victoria",
    "buenos aires": "buenos aires",
    "miraflores": "miraflores",
    "san isidro": "san isidro",
    "lince": "lince",
    "barranco": "barranco",
    "surco": "santiago de surco",
    "surquillo": "surquillo",
    "la molina": "la molina",
    "san borja": "san borja",
    "pueblo libre": "pueblo libre",
    "jesus maria": "jesus maria",
    "magdalena": "magdalena del mar",
    "san miguel": "san miguel",
    "breña": "breña",
    "rimac": "rimac",
    "los olivos": "los olivos",
    "san martin de porres": "san martin de porres",
    "independencia": "independencia",
    "comas": "comas",
    "carabayllo": "carabayllo",
    "ate": "ate",
    "santa anita": "santa anita",
    "el agustino": "el agustino",
    "san juan de lurigancho": "san juan de lurigancho",
    "san juan de miraflores": "san juan de miraflores",
    "villa maria del triunfo": "villa maria del triunfo",
    "villa el salvador": "villa el salvador",
    "lurin": "lurin",
    "chorrillos": "chorrillos",
    "puente piedra": "puente piedra",
    "ventanilla": "ventanilla",
}

# Mapeo VOTOS_Pn → (partido_id, nombre_partido, candidato) para 1ra vuelta 2021.
# Verificado contra los sumatorios reales de las columnas en el CSV oficial PCM
# y los resultados publicados por ONPE/Wikipedia (junio 2021). Las columnas
# `VOTOS_Pn` del CSV NO siguen el orden alfabético por partido; sus posiciones
# fueron determinadas empíricamente por sum-matching con los totales históricos.
_PARTY_MAP_2021_1V: dict[str, tuple[str, str, str]] = {
    "VOTOS_P1":  ("PNP", "Partido Nacionalista Peruano", "Ollanta Humala Tasso"),
    "VOTOS_P2":  ("FA",  "Frente Amplio", "Marco Arana Zegarra"),
    "VOTOS_P3":  ("PM",  "Partido Morado", "Julio Guzmán"),
    "VOTOS_P4":  ("PPS", "Perú Patria Segura", "Rafael Santos Alvarado"),
    "VOTOS_P5":  ("VN",  "Victoria Nacional", "George Forsyth Sommer"),
    "VOTOS_P6":  ("AP",  "Acción Popular", "Yonhy Lescano Ancieta"),
    "VOTOS_P7":  ("AP2", "Avanza País", "Hernando de Soto"),
    "VOTOS_P8":  ("PP",  "Podemos Perú", "Daniel Urresti"),
    "VOTOS_P9":  ("JP",  "Juntos por el Perú", "Verónika Mendoza"),
    "VOTOS_P10": ("SP",  "Somos Perú", "Daniel Salaverry"),
    "VOTOS_P11": ("K",   "Fuerza Popular", "Keiko Fujimori Higuchi"),
    "VOTOS_P12": ("DD",  "Democracia Directa", "Andrés Alcántara"),
    "VOTOS_P13": ("RL",  "Renovación Popular", "Rafael López Aliaga"),
    "VOTOS_P14": ("PPC", "Partido Popular Cristiano", "Alberto Beingolea Delgado"),
    "VOTOS_P15": ("RUN", "Renacimiento Unido Nacional", "Ciro Gálvez Herrera"),
    "VOTOS_P16": ("PC",  "Perú Libre", "Pedro Castillo Terrones"),
    "VOTOS_P17": ("UPP", "Unión por el Perú", "José Vega Antonio"),
    "VOTOS_P18": ("APP", "Alianza para el Progreso", "César Acuña Peralta"),
}

_PARTY_MAP_2021_2V: dict[str, tuple[str, str, str]] = {
    "VOTOS_P1": ("PC", "Perú Libre", "Pedro Castillo Terrones"),
    "VOTOS_P2": ("K", "Fuerza Popular", "Keiko Fujimori Higuchi"),
}

# Sentinel key stored in `sv_sync_meta` after each successful 2021 bootstrap.
# Used to detect stale SQLite caches when the in-code party map changes between
# repo updates (avoids silent skip in `bootstrap_elecciones_2021`).
_PARTY_MAP_2021_FINGERPRINT_KEY = "party_map_2021_fingerprint"


def _compute_party_map_2021_fingerprint() -> str:
    """Stable short hash of both 2021 party maps used as a cache-invalidation key."""
    payload = json.dumps(
        {
            "v1": sorted(
                (col, pid, partido, candidato)
                for col, (pid, partido, candidato) in _PARTY_MAP_2021_1V.items()
            ),
            "v2": sorted(
                (col, pid, partido, candidato)
                for col, (pid, partido, candidato) in _PARTY_MAP_2021_2V.items()
            ),
        },
        ensure_ascii=False,
        sort_keys=True,
    ).encode("utf-8")
    return sha256(payload).hexdigest()[:16]


def _to_int_safe(value: Any) -> int:
    if value is None:
        return 0
    s = str(value).strip().replace('"', "")
    if s == "":
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0

class DataStore:
    def __init__(self, data_dir: Path) -> None:
        self.data_dir = data_dir
        # Runtime DB única del MCP: denorm (offline-first).
        # onpe.db queda sólo como artefacto de construcción para build_denorm.py.
        self.db_path = data_dir / "onpe_denorm.db"
        self.legacy_oltp_db_path = data_dir / "onpe.db"
        self.raw_dir = data_dir / "raw"
        self.reports_dir = data_dir / "reports"
        self.denorm_db_path = data_dir / "onpe_denorm.db"
        self._denorm_ready: bool | None = None

        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.raw_dir.mkdir(parents=True, exist_ok=True)
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        self._init_schema()
        self._migrate_legacy_oltp_if_needed()

    @property
    def denorm_available(self) -> bool:
        if self._denorm_ready is None:
            if not self.denorm_db_path.exists():
                self._denorm_ready = False
            else:
                try:
                    with sqlite3.connect(self.denorm_db_path) as conn:
                        row = conn.execute(
                            """
                            SELECT COUNT(*) AS c
                            FROM sqlite_master
                            WHERE type='table'
                              AND name IN ('fact_votos_nacional', 'fact_votos_mesa')
                            """
                        ).fetchone()
                    self._denorm_ready = bool(row and int(row[0] or 0) == 2)
                except Exception:
                    self._denorm_ready = False
        return self._denorm_ready

    def _connect_denorm(self) -> sqlite3.Connection:
        uri = f"file:{self.denorm_db_path}?mode=ro"
        conn = sqlite3.connect(uri, uri=True)
        conn.execute("PRAGMA cache_size=-131072")   # 128 MB page cache
        conn.execute("PRAGMA mmap_size=536870912")  # 512 MB mmap
        conn.row_factory = sqlite3.Row
        return conn

    def _query_sv_geo_denorm(self, nivel: str, ubigeo: str | None, nombre: str | None, top_n: int) -> list[dict]:
        """Denorm fast-path for query_sv_geo."""
        conn = self._connect_denorm()
        try:
            nivel_norm = nivel.lower().strip() if nivel else "nacional"

            if nivel_norm == "nacional":
                rows = conn.execute("""
                    SELECT partido_id, nombre_partido AS nombre_agrupacion,
                           candidato AS nombre_candidato,
                           votos, pct_votos_validos,
                           total_mesas, mesas_contabilizadas,
                           total_electores_habiles, total_votos_emitidos, total_votos_validos
                    FROM fact_votos_nacional
                    WHERE election_year=2026 AND vuelta=2 AND es_especial=0
                    ORDER BY votos DESC
                    LIMIT ?
                """, (top_n,)).fetchall()
                result = []
                for r in rows:
                    result.append({
                        "partido_id": r["partido_id"],
                        "nombre_agrupacion": r["nombre_agrupacion"],
                        "nombre_candidato": r["nombre_candidato"],
                        "votos_validos": r["votos"],
                        "pct_votos_validos": r["pct_votos_validos"],
                        "ubigeo": "000000",
                    })
                return result
            elif nivel_norm in ("pais_exterior", "pais"):
                where_country = ""
                params_country: list[Any] = []
                if nombre:
                    where_country = "AND UPPER(pais) = UPPER(?)"
                    params_country.append(nombre)
                rows = conn.execute(
                    f"""
                    SELECT pais AS nombre_geo, partido_id,
                           nombre_partido AS nombre_agrupacion,
                           candidato AS nombre_candidato,
                           votos AS votos_validos,
                           ROUND(CASE WHEN total_votos_validos > 0
                                      THEN (100.0 * votos / total_votos_validos)
                                      ELSE 0 END, 4) AS pct_votos_validos
                    FROM fact_votos_pais
                    WHERE election_year=2026 AND vuelta=2 AND es_especial=0 {where_country}
                    ORDER BY votos DESC
                    LIMIT ?
                    """,
                    params_country + [top_n],
                ).fetchall()
                return [dict(r) for r in rows]
            elif nivel_norm == "continente":
                rows = conn.execute(
                    """
                    SELECT continente AS nombre_geo, partido_id,
                           nombre_partido AS nombre_agrupacion,
                           candidato AS nombre_candidato,
                           SUM(votos) AS votos_validos
                    FROM fact_votos_pais
                    WHERE election_year=2026 AND vuelta=2 AND es_especial=0
                    GROUP BY continente, partido_id, nombre_partido, candidato
                    ORDER BY votos_validos DESC
                    LIMIT ?
                    """,
                    (top_n,),
                ).fetchall()
                return [dict(r) for r in rows]

            elif nivel_norm in ("departamento", "region", "dpto"):
                filter_clause = ""
                params: list = [top_n]
                if ubigeo:
                    filter_clause = "AND SUBSTR('000000'||cod_departamento,-2) = SUBSTR(?,1,2)"
                    params = [ubigeo[:2].zfill(2)] + [top_n]
                elif nombre:
                    filter_clause = "AND UPPER(departamento) = UPPER(?)"
                    params = [nombre] + [top_n]
                rows = conn.execute(f"""
                    SELECT cod_departamento, departamento AS nombre_departamento,
                           partido_id, nombre_partido AS nombre_agrupacion,
                           candidato AS nombre_candidato,
                           votos,
                           ROUND(CASE WHEN total_votos_validos > 0
                                      THEN (100.0 * votos / total_votos_validos)
                                      ELSE 0 END, 4) AS pct_votos_validos
                    FROM fact_votos_departamento
                    WHERE election_year=2026 AND vuelta=2 AND es_especial=0 {filter_clause}
                    ORDER BY votos DESC
                    LIMIT ?
                """, params).fetchall()
                result = []
                for r in rows:
                    ub = str(r["cod_departamento"]).zfill(2) + "0000"
                    result.append({
                        "partido_id": r["partido_id"],
                        "nombre_agrupacion": r["nombre_agrupacion"],
                        "nombre_candidato": r["nombre_candidato"],
                        "votos_validos": r["votos"],
                        "pct_votos_validos": r["pct_votos_validos"],
                        "ubigeo": ub,
                        "nombre_departamento": r["nombre_departamento"],
                    })
                return result

            elif nivel_norm in ("provincia",):
                filter_clause = ""
                params2: list = [top_n]
                if ubigeo:
                    filter_clause = "AND SUBSTR('000000'||cod_provincia,-4) = SUBSTR(?,1,4)"
                    params2 = [ubigeo[:4].zfill(4)] + [top_n]
                elif nombre:
                    filter_clause = "AND UPPER(provincia) = UPPER(?)"
                    params2 = [nombre] + [top_n]
                rows2 = conn.execute(f"""
                    SELECT cod_provincia, provincia AS nombre_provincia,
                           partido_id, nombre_partido AS nombre_agrupacion,
                           candidato AS nombre_candidato,
                           votos,
                           ROUND(CASE WHEN total_votos_validos > 0
                                      THEN (100.0 * votos / total_votos_validos)
                                      ELSE 0 END, 4) AS pct_votos_validos
                    FROM fact_votos_provincia
                    WHERE election_year=2026 AND vuelta=2 AND es_especial=0 {filter_clause}
                    ORDER BY votos DESC
                    LIMIT ?
                """, params2).fetchall()
                result2 = []
                for r in rows2:
                    ub = str(r["cod_provincia"]).zfill(4) + "00"
                    result2.append({
                        "partido_id": r["partido_id"],
                        "nombre_agrupacion": r["nombre_agrupacion"],
                        "nombre_candidato": r["nombre_candidato"],
                        "nombre_geo": r["nombre_provincia"],
                        "votos_validos": r["votos"],
                        "pct_votos_validos": r["pct_votos_validos"],
                        "ubigeo": ub,
                        "nombre_provincia": r["nombre_provincia"],
                    })
                return result2

            else:  # distrito / ubigeo — also handles continente/pais_exterior via raise
                if nivel_norm in ("continente", "pais_exterior", "ciudad"):
                    raise ValueError(f"nivel '{nivel_norm}' not supported in denorm fast-path")
                filter_clause2 = ""
                params3: list = [top_n]
                if ubigeo:
                    filter_clause2 = "AND SUBSTR('000000'||ubigeo,-6) = SUBSTR('000000'||?,-6)"
                    params3 = [ubigeo] + [top_n]
                rows3 = conn.execute(f"""
                    SELECT ubigeo, distrito AS nombre_distrito,
                           partido_id, nombre_partido AS nombre_agrupacion,
                           candidato AS nombre_candidato,
                           votos,
                           ROUND(CASE WHEN total_votos_validos > 0
                                      THEN (100.0 * votos / total_votos_validos)
                                      ELSE 0 END, 4) AS pct_votos_validos
                    FROM fact_votos_ubigeo
                    WHERE election_year=2026 AND vuelta=2 AND es_especial=0 {filter_clause2}
                    ORDER BY votos DESC
                    LIMIT ?
                """, params3).fetchall()
                return [
                    {
                        "partido_id": r["partido_id"],
                        "nombre_agrupacion": r["nombre_agrupacion"],
                        "nombre_candidato": r["nombre_candidato"],
                        "votos_validos": r["votos"],
                        "pct_votos_validos": r["pct_votos_validos"],
                        "ubigeo": r["ubigeo"],
                        "nombre_distrito": r["nombre_distrito"],
                    }
                    for r in rows3
                ]
        finally:
            conn.close()

    def _resultados_geo_1v_denorm(self, nivel: str, filtro: str | None, top_n: int) -> list[dict] | None:
        """Returns resultados_geo_2026_1v result from denorm. None = use OLTP."""
        try:
            conn = self._connect_denorm()
            try:
                nivel_norm = nivel.lower().strip() if nivel else "nacional"
                if nivel_norm in ("nacional",):
                    rows = conn.execute("""
                        SELECT partido_id, nombre_partido, candidato,
                               votos, pct_votos_validos,
                               total_mesas, mesas_contabilizadas
                        FROM fact_votos_nacional
                        WHERE election_year=2026 AND vuelta=1 AND es_especial=0
                        ORDER BY votos DESC LIMIT ?
                    """, (top_n,)).fetchall()
                elif nivel_norm in ("departamento", "region", "dpto"):
                    if filtro:
                        import unicodedata as _ud

                        def _norm_dn(t: str) -> str:
                            return "".join(c for c in _ud.normalize("NFKD", t) if not _ud.combining(c)).casefold()

                        rows_all = conn.execute("""
                            SELECT departamento AS nombre_departamento, partido_id, nombre_partido, candidato,
                                   votos,
                                   ROUND(CASE WHEN total_votos_validos > 0
                                              THEN (100.0 * votos / total_votos_validos)
                                              ELSE 0 END, 4) AS pct_votos_validos
                            FROM fact_votos_departamento
                            WHERE election_year=2026 AND vuelta=1 AND es_especial=0
                        """).fetchall()
                        rows = [r for r in rows_all if _norm_dn(r["nombre_departamento"]) == _norm_dn(filtro)][:top_n]
                    else:
                        rows = conn.execute("""
                            SELECT partido_id, nombre_partido, candidato,
                                   votos,
                                   ROUND(CASE WHEN total_votos_validos > 0
                                              THEN (100.0 * votos / total_votos_validos)
                                              ELSE 0 END, 4) AS pct_votos_validos
                            FROM fact_votos_departamento
                            WHERE election_year=2026 AND vuelta=1 AND es_especial=0
                            ORDER BY votos DESC LIMIT ?
                        """, (top_n,)).fetchall()
                else:
                    return None
                return [dict(r) for r in rows]
            finally:
                conn.close()
        except Exception:
            return None

    @staticmethod
    def now_iso() -> str:
        return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    @staticmethod
    def _payload_hash(payload: dict[str, Any]) -> str:
        canonical = json.dumps(payload, sort_keys=True, ensure_ascii=False)
        return sha256(canonical.encode("utf-8")).hexdigest()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path, timeout=30)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA cache_size=-65536")   # 64 MB page cache
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA temp_store=MEMORY")
        return conn

    def _migrate_legacy_oltp_if_needed(self) -> None:
        """One-shot copy of legacy onpe.db runtime tables into onpe_denorm.db.

        This keeps MCP runtime denorm-only while preserving existing tool behavior
        for tables that are hydrated from scrapers.
        """
        if self.legacy_oltp_db_path == self.db_path:
            return
        if not self.legacy_oltp_db_path.exists():
            return

        with self._connect() as conn:
            # If already hydrated in denorm runtime tables, skip migration.
            try:
                row = conn.execute(
                    """
                    SELECT
                      (SELECT COUNT(*) FROM mesas_data) AS c1,
                      (SELECT COUNT(*) FROM votos) AS c2,
                      (SELECT COUNT(*) FROM mesas_sv) AS c3,
                      (SELECT COUNT(*) FROM votos_sv) AS c4,
                      (SELECT COUNT(*) FROM mesas_2021) AS c5,
                      (SELECT COUNT(*) FROM votos_2021) AS c6
                    """
                ).fetchone()
                if row and all(int(row[k] or 0) > 0 for k in ("c1", "c2", "c3", "c4", "c5", "c6")):
                    return
            except Exception:
                return

            conn.execute("ATTACH DATABASE ? AS legacy", (str(self.legacy_oltp_db_path),))
            try:
                tables = conn.execute(
                    """
                    SELECT name
                    FROM legacy.sqlite_master
                    WHERE type='table'
                      AND name NOT LIKE 'sqlite_%'
                    """
                ).fetchall()

                for trow in tables:
                    table = str(trow["name"])
                    if table in {
                        "dim_eleccion",
                        "dim_partido",
                        "dim_geo",
                        "fact_votos_mesa",
                        "fact_votos_ubigeo",
                        "fact_votos_provincia",
                        "fact_votos_departamento",
                        "fact_votos_nacional",
                        "fact_votos_pais",
                    }:
                        continue

                    src_cols = [r["name"] for r in conn.execute(f'PRAGMA legacy.table_info("{table}")').fetchall()]
                    if not src_cols:
                        continue

                    conn.execute(f'CREATE TABLE IF NOT EXISTS "{table}" AS SELECT * FROM legacy."{table}" WHERE 0')
                    dst_cols = [r["name"] for r in conn.execute(f'PRAGMA table_info("{table}")').fetchall()]
                    common = [c for c in src_cols if c in dst_cols]
                    if not common:
                        continue

                    quoted = ", ".join(f'"{c}"' for c in common)
                    conn.execute(f'DELETE FROM "{table}"')
                    conn.execute(
                        f'INSERT INTO "{table}" ({quoted}) SELECT {quoted} FROM legacy."{table}"'
                    )
                conn.commit()
            finally:
                try:
                    conn.execute("DETACH DATABASE legacy")
                except sqlite3.OperationalError:
                    pass

    def _init_schema(self) -> None:
        with self._connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS mesa_cache (
                    codigo_mesa TEXT PRIMARY KEY,
                    payload_json TEXT NOT NULL,
                    fetched_at TEXT NOT NULL,
                    source TEXT NOT NULL,
                    id_eleccion INTEGER NOT NULL,
                    payload_hash TEXT NOT NULL,
                    schema_version INTEGER NOT NULL DEFAULT 1
                );

                CREATE TABLE IF NOT EXISTS mesas_data (
                    codigo_mesa TEXT PRIMARY KEY,
                    ubigeo TEXT,
                    local_votacion TEXT,
                    electores_habiles INTEGER,
                    votos_emitidos INTEGER,
                    votos_validos INTEGER,
                    blancos INTEGER,
                    nulos INTEGER,
                    impugnados INTEGER,
                    estado_acta TEXT,
                    fetched_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS votos (
                    codigo_mesa TEXT NOT NULL,
                    partido_id TEXT NOT NULL,
                    votos INTEGER,
                    fetched_at TEXT NOT NULL,
                    PRIMARY KEY (codigo_mesa, partido_id)
                );

                CREATE TABLE IF NOT EXISTS agrupaciones (
                    partido_id TEXT PRIMARY KEY,
                    nombre TEXT,
                    candidato TEXT NOT NULL DEFAULT '',
                    fetched_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS foreign_catalog (
                    ubigeo TEXT PRIMARY KEY,
                    continente TEXT,
                    pais TEXT,
                    ciudad TEXT,
                    fetched_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS ubigeo_location_cache (
                    ubigeo TEXT PRIMARY KEY,
                    ambito TEXT,
                    departamento TEXT,
                    ciudad TEXT,
                    pais TEXT,
                    fetched_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS votos_by_ubigeo_partido (
                    ubigeo TEXT NOT NULL,
                    partido_id TEXT NOT NULL,
                    total_votos INTEGER NOT NULL,
                    fetched_at TEXT NOT NULL,
                    PRIMARY KEY (ubigeo, partido_id)
                );

                CREATE TABLE IF NOT EXISTS geo_query_cache (
                    query_key TEXT PRIMARY KEY,
                    payload_json TEXT NOT NULL,
                    fetched_at TEXT NOT NULL
                );

                CREATE INDEX IF NOT EXISTS idx_mesas_ubigeo ON mesas_data (ubigeo);
                CREATE INDEX IF NOT EXISTS idx_mesas_estado ON mesas_data (estado_acta);
                CREATE INDEX IF NOT EXISTS idx_votos_partido ON votos (partido_id);
                CREATE INDEX IF NOT EXISTS idx_foreign_pais ON foreign_catalog (pais);
                CREATE INDEX IF NOT EXISTS idx_foreign_ciudad ON foreign_catalog (ciudad);
                CREATE INDEX IF NOT EXISTS idx_ubigeo_location_depto ON ubigeo_location_cache (departamento);
                CREATE INDEX IF NOT EXISTS idx_ubigeo_location_ciudad ON ubigeo_location_cache (ciudad);
                CREATE INDEX IF NOT EXISTS idx_votos_ubigeo_partido ON votos_by_ubigeo_partido (ubigeo, partido_id);

                CREATE TABLE IF NOT EXISTS mesa_prefix_totals (
                    prefix TEXT PRIMARY KEY,
                    n_mesas INTEGER NOT NULL DEFAULT 0,
                    mesas_con_votos INTEGER NOT NULL DEFAULT 0,
                    votos_emitidos INTEGER NOT NULL DEFAULT 0,
                    votos_validos INTEGER NOT NULL DEFAULT 0,
                    rebuilt_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS mesa_prefix_party_summary (
                    prefix TEXT NOT NULL,
                    partido_id TEXT NOT NULL,
                    total_votos INTEGER NOT NULL DEFAULT 0,
                    n_mesas INTEGER NOT NULL DEFAULT 0,
                    PRIMARY KEY (prefix, partido_id)
                );
                CREATE INDEX IF NOT EXISTS idx_prefix_party ON mesa_prefix_party_summary (prefix, total_votos DESC);

                CREATE TABLE IF NOT EXISTS mesa_winner (
                    codigo_mesa TEXT PRIMARY KEY,
                    partido_id TEXT NOT NULL,
                    max_votos INTEGER NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_mesa_winner_partido ON mesa_winner (partido_id);
                CREATE INDEX IF NOT EXISTS idx_mesa_winner_mesa_partido ON mesa_winner (codigo_mesa, partido_id);

                CREATE TABLE IF NOT EXISTS ubigeo_reniec (
                    ubigeo TEXT PRIMARY KEY,
                    distrito TEXT,
                    provincia TEXT,
                    departamento TEXT,
                    distrito_norm TEXT,
                    provincia_norm TEXT,
                    departamento_norm TEXT,
                    fetched_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_reniec_dept_norm ON ubigeo_reniec (departamento_norm);
                CREATE INDEX IF NOT EXISTS idx_reniec_prov_norm ON ubigeo_reniec (provincia_norm);
                CREATE INDEX IF NOT EXISTS idx_reniec_dist_norm ON ubigeo_reniec (distrito_norm);

                CREATE TABLE IF NOT EXISTS ubigeo_onpe_api (
                    ubigeo TEXT PRIMARY KEY,
                    distrito TEXT,
                    provincia TEXT,
                    departamento TEXT,
                    distrito_norm TEXT,
                    provincia_norm TEXT,
                    departamento_norm TEXT,
                    fetched_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_onpe_api_dept_norm ON ubigeo_onpe_api (departamento_norm);
                CREATE INDEX IF NOT EXISTS idx_onpe_api_prov_norm ON ubigeo_onpe_api (provincia_norm);
                CREATE INDEX IF NOT EXISTS idx_onpe_api_dist_norm ON ubigeo_onpe_api (distrito_norm);

                -- ═══ ELECCIONES 2021 (1V y 2V) ════════════════════════════════════════════

                CREATE TABLE IF NOT EXISTS mesas_2021 (
                    vuelta INTEGER NOT NULL,
                    codigo_mesa TEXT NOT NULL,
                    ubigeo TEXT,
                    departamento TEXT,
                    provincia TEXT,
                    distrito TEXT,
                    tipo_eleccion TEXT,
                    descrip_estado_acta TEXT,
                    tipo_observacion TEXT,
                    n_cvas INTEGER,
                    n_elec_habil INTEGER,
                    votos_vb INTEGER,
                    votos_vn INTEGER,
                    votos_vi INTEGER,
                    votos_emitidos INTEGER,
                    votos_validos INTEGER,
                    fetched_at TEXT NOT NULL,
                    PRIMARY KEY (vuelta, codigo_mesa)
                );
                CREATE INDEX IF NOT EXISTS idx_mesas2021_geo ON mesas_2021 (vuelta, departamento, provincia, distrito);
                CREATE INDEX IF NOT EXISTS idx_mesas2021_ubigeo ON mesas_2021 (vuelta, ubigeo);

                CREATE TABLE IF NOT EXISTS votos_2021 (
                    vuelta INTEGER NOT NULL,
                    codigo_mesa TEXT NOT NULL,
                    partido_id TEXT NOT NULL,
                    votos INTEGER NOT NULL,
                    fetched_at TEXT NOT NULL,
                    PRIMARY KEY (vuelta, codigo_mesa, partido_id)
                );
                CREATE INDEX IF NOT EXISTS idx_votos2021_partido ON votos_2021 (vuelta, partido_id);
                CREATE INDEX IF NOT EXISTS idx_votos2021_mesa ON votos_2021 (vuelta, codigo_mesa);

                CREATE TABLE IF NOT EXISTS partidos_2021 (
                    vuelta INTEGER NOT NULL,
                    partido_id TEXT NOT NULL,
                    nombre_partido TEXT NOT NULL,
                    candidato TEXT NOT NULL,
                    fetched_at TEXT NOT NULL,
                    PRIMARY KEY (vuelta, partido_id)
                );
                CREATE INDEX IF NOT EXISTS idx_partidos2021_nombre ON partidos_2021 (vuelta, nombre_partido);
                CREATE INDEX IF NOT EXISTS idx_partidos2021_candidato ON partidos_2021 (vuelta, candidato);

                -- ═══ SEGUNDA VUELTA (SV) ═══════════════════════════════════════════════════

                CREATE TABLE IF NOT EXISTS mesas_sv (
                    codigo_mesa TEXT PRIMARY KEY,
                    id_ubigeo TEXT,
                    nombre_local TEXT,
                    id_ambito INTEGER,
                    electores_habiles INTEGER,
                    votos_emitidos INTEGER,
                    votos_validos INTEGER,
                    total_asistentes INTEGER,
                    codigo_estado_acta TEXT,
                    fetched_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_mesas_sv_ubigeo ON mesas_sv (id_ubigeo);
                CREATE INDEX IF NOT EXISTS idx_mesas_sv_estado ON mesas_sv (codigo_estado_acta);

                CREATE TABLE IF NOT EXISTS votos_sv (
                    codigo_mesa TEXT NOT NULL,
                    partido_id TEXT NOT NULL,
                    votos INTEGER,
                    fetched_at TEXT NOT NULL,
                    PRIMARY KEY (codigo_mesa, partido_id)
                );
                CREATE INDEX IF NOT EXISTS idx_votos_sv_partido ON votos_sv (partido_id);
                CREATE INDEX IF NOT EXISTS idx_votos_sv_mesa ON votos_sv (codigo_mesa);

                CREATE TABLE IF NOT EXISTS agrupaciones_sv (
                    partido_id TEXT PRIMARY KEY,
                    nombre TEXT,
                    fetched_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS ubicaciones_sv (
                    ubigeo TEXT PRIMARY KEY,
                    ambito TEXT,
                    departamento TEXT,
                    provincia TEXT,
                    distrito TEXT,
                    continente TEXT,
                    pais TEXT,
                    ciudad TEXT,
                    fetched_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_ubicaciones_sv_dep ON ubicaciones_sv (departamento);
                CREATE INDEX IF NOT EXISTS idx_ubicaciones_sv_prov ON ubicaciones_sv (provincia);

                CREATE TABLE IF NOT EXISTS locales_reasignados_sv (
                    nro INTEGER PRIMARY KEY,
                    odpe TEXT,
                    dpto TEXT,
                    provincia TEXT,
                    distrito TEXT,
                    ccpp TEXT,
                    nombre_local_original TEXT,
                    nombre_local_nuevo TEXT,
                    motivo TEXT,
                    mesas_afectadas INTEGER,
                    estado_parseo TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_reasignados_dpto ON locales_reasignados_sv (dpto);
                CREATE INDEX IF NOT EXISTS idx_reasignados_motivo ON locales_reasignados_sv (motivo);

                -- sv_resumen_* tables loaded from scraper's resumen/ files

                CREATE TABLE IF NOT EXISTS sv_resumen_nacional (
                    partido_id TEXT PRIMARY KEY,
                    nombre_candidato TEXT,
                    nombre_agrupacion TEXT,
                    votos_validos INTEGER,
                    pct_votos_validos REAL,
                    pct_votos_emitidos REAL,
                    actas_contabilizadas_pct REAL,
                    contabilizadas INTEGER,
                    total_actas INTEGER,
                    participacion_ciudadana REAL,
                    fecha_actualizacion TEXT,
                    fuente TEXT,
                    loaded_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS sv_resumen_departamentos (
                    ubigeo TEXT NOT NULL,
                    partido_id TEXT NOT NULL,
                    nombre_candidato TEXT,
                    nombre_agrupacion TEXT,
                    votos_validos INTEGER,
                    pct_votos_validos REAL,
                    pct_votos_emitidos REAL,
                    total_votos_validos_geo INTEGER,
                    total_votos_emitidos_geo INTEGER,
                    fuente TEXT,
                    loaded_at TEXT NOT NULL,
                    PRIMARY KEY (ubigeo, partido_id)
                );
                CREATE INDEX IF NOT EXISTS idx_sv_rdept_ubigeo ON sv_resumen_departamentos (ubigeo);

                CREATE TABLE IF NOT EXISTS sv_resumen_provincias (
                    ubigeo TEXT NOT NULL,
                    partido_id TEXT NOT NULL,
                    nombre_candidato TEXT,
                    nombre_agrupacion TEXT,
                    nombre_geo TEXT,
                    votos_validos INTEGER,
                    pct_votos_validos REAL,
                    pct_votos_emitidos REAL,
                    total_votos_validos_geo INTEGER,
                    total_votos_emitidos_geo INTEGER,
                    fuente TEXT,
                    loaded_at TEXT NOT NULL,
                    PRIMARY KEY (ubigeo, partido_id)
                );
                CREATE INDEX IF NOT EXISTS idx_sv_rprov_ubigeo ON sv_resumen_provincias (ubigeo);
                CREATE INDEX IF NOT EXISTS idx_sv_rprov_nombre_geo ON sv_resumen_provincias (nombre_geo);

                CREATE TABLE IF NOT EXISTS sv_resumen_cobertura (
                    ubigeo TEXT PRIMARY KEY,
                    nombre_departamento TEXT,
                    actas_contabilizadas INTEGER,
                    pct_actas_contabilizadas REAL,
                    fuente TEXT,
                    loaded_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_sv_rcob_ubigeo ON sv_resumen_cobertura (ubigeo);

                -- CTAS aggregation tables (distrito and ciudad level — scraper doesn't pre-compute these)

                CREATE TABLE IF NOT EXISTS sv_agg_distrito (
                    ubigeo TEXT NOT NULL,
                    partido_id TEXT NOT NULL,
                    nombre_candidato TEXT,
                    votos INTEGER,
                    total_mesas INTEGER,
                    mesas_contabilizadas INTEGER,
                    rebuilt_at TEXT NOT NULL,
                    PRIMARY KEY (ubigeo, partido_id)
                );
                CREATE INDEX IF NOT EXISTS idx_sv_agg_distrito_ubigeo ON sv_agg_distrito (ubigeo);

                CREATE TABLE IF NOT EXISTS sv_agg_ciudad (
                    ubigeo TEXT NOT NULL,
                    ciudad TEXT NOT NULL,
                    partido_id TEXT NOT NULL,
                    nombre_candidato TEXT,
                    votos INTEGER,
                    total_mesas INTEGER,
                    mesas_contabilizadas INTEGER,
                    rebuilt_at TEXT NOT NULL,
                    PRIMARY KEY (ubigeo, ciudad, partido_id)
                );
                CREATE INDEX IF NOT EXISTS idx_sv_agg_ciudad_ubigeo ON sv_agg_ciudad (ubigeo);
                CREATE INDEX IF NOT EXISTS idx_sv_agg_ciudad_nombre ON sv_agg_ciudad (ciudad);

                -- Vote transfer projection table
                CREATE TABLE IF NOT EXISTS proyeccion_sv_by_ubigeo (
                    ubigeo TEXT PRIMARY KEY,
                    votos_1v_total INTEGER,
                    votos_proyectados_keiko INTEGER,
                    votos_proyectados_sanchez INTEGER,
                    votos_proyectados_bn INTEGER,
                    votos_abstencion_estimada INTEGER,
                    rebuilt_at TEXT NOT NULL
                );

                -- Transfer map seeds (static knowledge)
                CREATE TABLE IF NOT EXISTS voto_transfer_map (
                    partido_nombre_norm TEXT PRIMARY KEY,
                    peso_keiko REAL NOT NULL,
                    peso_sanchez REAL NOT NULL,
                    peso_bn REAL NOT NULL,
                    fuente TEXT NOT NULL,
                    loaded_at TEXT NOT NULL
                );

                -- Scraper commit guard (avoid redundant full rebuilds)
                CREATE TABLE IF NOT EXISTS sv_sync_meta (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                """
            )
            cols = {
                str(r["name"])
                for r in conn.execute("PRAGMA table_info(agrupaciones)").fetchall()
            }
            if "candidato" not in cols:
                conn.execute("ALTER TABLE agrupaciones ADD COLUMN candidato TEXT NOT NULL DEFAULT ''")

    @staticmethod
    def _mesa_prefix_like(prefix: str) -> str:
        normalized = str(prefix or "").strip()
        # Convención conversacional: 900K => 900000, que representa el bloque 900xxx.
        if len(normalized) == 6 and normalized.endswith("000"):
            return normalized[:3]
        return normalized

    def append_raw_event(self, event_type: str, payload: dict[str, Any]) -> None:
        line = {
            "timestamp": self.now_iso(),
            "event_type": event_type,
            "payload": payload,
        }
        path = self.raw_dir / "events.jsonl"
        with path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(line, ensure_ascii=False) + "\n")

    def get_cached_mesa(self, codigo_mesa: str, max_age_seconds: int) -> dict[str, Any] | None:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT payload_json, fetched_at FROM mesa_cache WHERE codigo_mesa = ?",
                (codigo_mesa,),
            ).fetchone()

        if row is None:
            return None

        fetched_at = datetime.fromisoformat(str(row["fetched_at"]).replace("Z", "+00:00"))
        age_seconds = (datetime.now(timezone.utc) - fetched_at).total_seconds()
        if age_seconds > max_age_seconds:
            return None

        return json.loads(str(row["payload_json"]))

    def get_mesa_from_local(self, codigo_mesa: str) -> dict[str, Any] | None:
        """Construye bundle de mesa desde tablas locales (mesas_data + votos + agrupaciones).

        Evita llamada al API live cuando los datos ya están hidratados.
        Retorna None si la mesa no existe en mesas_data.
        """
        with self._connect() as conn:
            mesa_row = conn.execute(
                "SELECT * FROM mesas_data WHERE codigo_mesa = ?",
                (codigo_mesa,),
            ).fetchone()
            if mesa_row is None:
                return None

            votos_rows = conn.execute(
                """SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre_partido, v.votos
                   FROM votos v
                   LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                   WHERE v.codigo_mesa = ?
                   ORDER BY v.votos DESC""",
                (codigo_mesa,),
            ).fetchall()

            location_row = conn.execute(
                "SELECT departamento, ciudad, pais FROM ubigeo_location_cache WHERE ubigeo = ?",
                (str(mesa_row["ubigeo"] or ""),),
            ).fetchone()

        mesa_data: dict[str, Any] = {
            "codigo_mesa": codigo_mesa,
            "ubigeo": str(mesa_row["ubigeo"] or ""),
            "local_votacion": str(mesa_row["local_votacion"] or ""),
            "electores_habiles": int(mesa_row["electores_habiles"] or 0),
            "votos_emitidos": int(mesa_row["votos_emitidos"] or 0),
            "votos_validos": int(mesa_row["votos_validos"] or 0),
            "blancos": int(mesa_row["blancos"] or 0),
            "nulos": int(mesa_row["nulos"] or 0),
            "impugnados": int(mesa_row["impugnados"] or 0),
            "estado_acta": str(mesa_row["estado_acta"] or ""),
        }
        if location_row:
            mesa_data.update({
                "departamento": str(location_row["departamento"] or ""),
                "ciudad": str(location_row["ciudad"] or ""),
            })

        votos = [
            {
                "partido_id": str(r["partido_id"]),
                "nombre_partido": str(r["nombre_partido"]),
                "votos": int(r["votos"] or 0),
            }
            for r in votos_rows
        ]

        return {
            "codigo_mesa": codigo_mesa,
            "found": True,
            "mesa_data": mesa_data,
            "agrupaciones": [{"partido_id": v["partido_id"], "nombre": v["nombre_partido"]} for v in votos],
            "votos": votos,
            "source": "local_db",
        }

    def upsert_mesa_bundle(
        self,
        codigo_mesa: str,
        mesa_payload: dict[str, Any],
        *,
        source: str,
        id_eleccion: int,
    ) -> None:
        now = self.now_iso()
        payload_hash = self._payload_hash(mesa_payload)

        mesa_data = mesa_payload.get("mesa_data") or {}
        votos = mesa_payload.get("votos") or []
        agrupaciones = mesa_payload.get("agrupaciones") or []

        with self._connect() as conn:
            prev_mesa_row = conn.execute(
                "SELECT ubigeo FROM mesas_data WHERE codigo_mesa = ?",
                (codigo_mesa,),
            ).fetchone()
            prev_ubigeo = str(prev_mesa_row["ubigeo"] or "") if prev_mesa_row else ""

            prev_votes_rows = conn.execute(
                "SELECT partido_id, votos FROM votos WHERE codigo_mesa = ?",
                (codigo_mesa,),
            ).fetchall()
            prev_votes: dict[str, int] = {
                str(row["partido_id"]): int(row["votos"] or 0)
                for row in prev_votes_rows
            }

            conn.execute(
                """
                INSERT INTO mesa_cache (codigo_mesa, payload_json, fetched_at, source, id_eleccion, payload_hash, schema_version)
                VALUES (?, ?, ?, ?, ?, ?, 1)
                ON CONFLICT(codigo_mesa) DO UPDATE SET
                    payload_json=excluded.payload_json,
                    fetched_at=excluded.fetched_at,
                    source=excluded.source,
                    id_eleccion=excluded.id_eleccion,
                    payload_hash=excluded.payload_hash
                """,
                (
                    codigo_mesa,
                    json.dumps(mesa_payload, ensure_ascii=False),
                    now,
                    source,
                    id_eleccion,
                    payload_hash,
                ),
            )

            if mesa_data:
                conn.execute(
                    """
                    INSERT INTO mesas_data (
                        codigo_mesa, ubigeo, local_votacion, electores_habiles, votos_emitidos, votos_validos,
                        blancos, nulos, impugnados, estado_acta, fetched_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(codigo_mesa) DO UPDATE SET
                        ubigeo=excluded.ubigeo,
                        local_votacion=excluded.local_votacion,
                        electores_habiles=excluded.electores_habiles,
                        votos_emitidos=excluded.votos_emitidos,
                        votos_validos=excluded.votos_validos,
                        blancos=excluded.blancos,
                        nulos=excluded.nulos,
                        impugnados=excluded.impugnados,
                        estado_acta=excluded.estado_acta,
                        fetched_at=excluded.fetched_at
                    """,
                    (
                        mesa_data.get("codigo_mesa"),
                        mesa_data.get("ubigeo"),
                        mesa_data.get("local_votacion"),
                        mesa_data.get("electores_habiles") or 0,
                        mesa_data.get("votos_emitidos") or 0,
                        mesa_data.get("votos_validos") or 0,
                        mesa_data.get("blancos") or 0,
                        mesa_data.get("nulos") or 0,
                        mesa_data.get("impugnados") or 0,
                        mesa_data.get("estado_acta"),
                        now,
                    ),
                )

            for item in agrupaciones:
                conn.execute(
                    """
                    INSERT INTO agrupaciones (partido_id, nombre, fetched_at)
                    VALUES (?, ?, ?)
                    ON CONFLICT(partido_id) DO UPDATE SET
                        nombre=excluded.nombre,
                        fetched_at=excluded.fetched_at
                    """,
                    (str(item.get("partido_id", "")), str(item.get("nombre", "")), now),
                )

            for item in votos:
                conn.execute(
                    """
                    INSERT INTO votos (codigo_mesa, partido_id, votos, fetched_at)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(codigo_mesa, partido_id) DO UPDATE SET
                        votos=excluded.votos,
                        fetched_at=excluded.fetched_at
                    """,
                    (
                        str(item.get("codigo_mesa", codigo_mesa)),
                        str(item.get("partido_id", "")),
                        int(item.get("votos") or 0),
                        now,
                    ),
                )

            new_ubigeo = str(mesa_data.get("ubigeo") or prev_ubigeo or "")
            new_votes: dict[str, int] = {}
            for item in votos:
                partido_id = str(item.get("partido_id", ""))
                if not partido_id:
                    continue
                new_votes[partido_id] = int(item.get("votos") or 0)

            self._apply_votes_aggregate_delta(
                conn,
                prev_ubigeo=prev_ubigeo,
                prev_votes=prev_votes,
                new_ubigeo=new_ubigeo,
                new_votes=new_votes,
                now=now,
            )

    @staticmethod
    def _apply_votes_aggregate_delta(
        conn: sqlite3.Connection,
        *,
        prev_ubigeo: str,
        prev_votes: dict[str, int],
        new_ubigeo: str,
        new_votes: dict[str, int],
        now: str,
    ) -> None:
        def _accumulate(ubigeo: str, delta_map: dict[str, int]) -> None:
            if not ubigeo or not delta_map:
                return
            for partido_id, delta in delta_map.items():
                if not partido_id or delta == 0:
                    continue
                conn.execute(
                    """
                    INSERT INTO votos_by_ubigeo_partido (ubigeo, partido_id, total_votos, fetched_at)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(ubigeo, partido_id) DO UPDATE SET
                        total_votos=votos_by_ubigeo_partido.total_votos + excluded.total_votos,
                        fetched_at=excluded.fetched_at
                    """,
                    (ubigeo, partido_id, int(delta), now),
                )

        if prev_ubigeo:
            prev_delta = {pid: -int(v) for pid, v in prev_votes.items() if int(v) != 0}
            _accumulate(prev_ubigeo, prev_delta)

        if new_ubigeo:
            new_delta = {pid: int(v) for pid, v in new_votes.items() if int(v) != 0}
            _accumulate(new_ubigeo, new_delta)

        conn.execute("DELETE FROM votos_by_ubigeo_partido WHERE total_votos <= 0")

    def _ensure_votes_by_ubigeo_partido_backfilled(self, conn: sqlite3.Connection) -> None:
        row = conn.execute("SELECT COUNT(*) AS c FROM votos_by_ubigeo_partido").fetchone()
        if row and int(row["c"] or 0) > 0:
            return

        votos_row = conn.execute("SELECT COUNT(*) AS c FROM votos").fetchone()
        if not votos_row or int(votos_row["c"] or 0) == 0:
            return

        now = self.now_iso()
        conn.execute(
            """
            INSERT INTO votos_by_ubigeo_partido (ubigeo, partido_id, total_votos, fetched_at)
            SELECT m.ubigeo, v.partido_id, SUM(v.votos) AS total_votos, ?
            FROM votos v
            INNER JOIN mesas_data m ON m.codigo_mesa = v.codigo_mesa
            WHERE COALESCE(m.ubigeo, '') <> ''
            GROUP BY m.ubigeo, v.partido_id
            """,
            (now,),
        )


    @staticmethod
    def _norm(text: str) -> str:
        base = unicodedata.normalize("NFKD", text or "")
        stripped = "".join(ch for ch in base if not unicodedata.combining(ch))
        return stripped.casefold().strip()

    @staticmethod
    def load_candidate_map(path: Path) -> dict[str, str]:
        if not path.exists():
            return {}

        result: dict[str, str] = {}
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            for row in reader:
                partido_id = str(row.get("partido_id", "")).strip()
                candidato = str(row.get("Candidato", "")).strip()
                if not partido_id:
                    continue
                result[partido_id] = candidato
        return result

    @staticmethod
    def load_foreign_catalog(path: Path) -> list[dict[str, str]]:
        if not path.exists():
            return []

        rows: list[dict[str, str]] = []
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            for row in reader:
                ubigeo = str(row.get("ubigeo", "")).strip()
                if not ubigeo:
                    continue
                rows.append(
                    {
                        "ubigeo": ubigeo,
                        "Continente": str(row.get("Continente", "")).strip(),
                        "pais": str(row.get("pais", "")).strip(),
                        "ciudad": str(row.get("ciudad", "")).strip(),
                    }
                )
        return rows

    def upsert_foreign_catalog(self, rows: list[dict[str, str]]) -> int:
        if not rows:
            return 0

        now = self.now_iso()
        count = 0
        with self._connect() as conn:
            for row in rows:
                ubigeo = str(row.get("ubigeo", "")).strip()
                if not ubigeo:
                    continue
                conn.execute(
                    """
                    INSERT INTO foreign_catalog (ubigeo, continente, pais, ciudad, fetched_at)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(ubigeo) DO UPDATE SET
                        continente=excluded.continente,
                        pais=excluded.pais,
                        ciudad=excluded.ciudad,
                        fetched_at=excluded.fetched_at
                    """,
                    (
                        ubigeo,
                        str(row.get("Continente", "")).strip(),
                        str(row.get("pais", "")).strip(),
                        str(row.get("ciudad", "")).strip(),
                        now,
                    ),
                )
                count += 1
        return count

    def upsert_ubigeo_location(self, row: dict[str, str]) -> bool:
        ubigeo = str(row.get("ubigeo", "")).strip()
        if not ubigeo:
            return False

        ambito = str(row.get("ambito", "")).strip()
        departamento = str(row.get("departamento", "")).strip()
        ciudad = str(row.get("ciudad", "")).strip()
        pais = str(row.get("pais", "")).strip()
        if not departamento and not ciudad and not pais:
            return False

        now = self.now_iso()
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO ubigeo_location_cache (ubigeo, ambito, departamento, ciudad, pais, fetched_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(ubigeo) DO UPDATE SET
                    ambito=excluded.ambito,
                    departamento=excluded.departamento,
                    ciudad=excluded.ciudad,
                    pais=excluded.pais,
                    fetched_at=excluded.fetched_at
                """,
                (ubigeo, ambito, departamento, ciudad, pais, now),
            )
        return True

    def find_ubigeos_missing_city_or_department_by_mesa_prefix(
        self,
        mesa_prefix: str,
        *,
        limit: int = 50,
    ) -> list[str]:
        prefix = str(mesa_prefix or "").strip()
        if not prefix:
            return []

        like_prefix = self._mesa_prefix_like(prefix)
        limit = max(1, min(int(limit), 500))
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT DISTINCT m.ubigeo AS ubigeo
                FROM mesas_data m
                LEFT JOIN foreign_catalog f ON f.ubigeo = m.ubigeo
                LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                WHERE m.codigo_mesa LIKE ?
                  AND COALESCE(m.ubigeo, '') <> ''
                  AND (
                        COALESCE(NULLIF(TRIM(COALESCE(f.ciudad, ul.ciudad, '')), ''), '') = ''
                     OR COALESCE(NULLIF(TRIM(COALESCE(ul.departamento, '')), ''), '') = ''
                  )
                LIMIT ?
                """,
                (f"{like_prefix}%", limit),
            ).fetchall()
        return [str(row["ubigeo"] or "") for row in rows if str(row["ubigeo"] or "").strip()]

    def find_ubigeos_missing_city_or_department_by_ubigeo_prefix(
        self,
        ubigeo_prefix: str,
        *,
        limit: int = 50,
    ) -> list[str]:
        prefix = str(ubigeo_prefix or "").strip()
        if not prefix:
            return []

        limit = max(1, min(int(limit), 500))
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT DISTINCT m.ubigeo AS ubigeo
                FROM mesas_data m
                LEFT JOIN foreign_catalog f ON f.ubigeo = m.ubigeo
                LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                WHERE COALESCE(m.ubigeo, '') LIKE ?
                  AND (
                        COALESCE(NULLIF(TRIM(COALESCE(f.ciudad, ul.ciudad, '')), ''), '') = ''
                     OR COALESCE(NULLIF(TRIM(COALESCE(ul.departamento, '')), ''), '') = ''
                  )
                LIMIT ?
                """,
                (f"{prefix}%", limit),
            ).fetchall()
        return [str(row["ubigeo"] or "") for row in rows if str(row["ubigeo"] or "").strip()]

    def summarize_ubigeo_prefix(self, ubigeo_prefix: str, sample_size: int = 5) -> dict[str, Any]:
        prefix = str(ubigeo_prefix or "").strip()
        if not prefix:
            raise ValueError("ubigeo_prefix no puede estar vacío")

        sample_size = max(1, min(int(sample_size), 20))
        like_value = f"{prefix}%"

        with self._connect() as conn:
            sample_rows = conn.execute(
                """
                SELECT
                    m.ubigeo AS ubigeo,
                    COALESCE(ul.departamento, '') AS departamento,
                    COALESCE(f.continente, '') AS continente,
                    COALESCE(f.pais, ul.pais, '') AS pais,
                    COALESCE(f.ciudad, ul.ciudad, '') AS ciudad,
                    COUNT(*) AS mesas,
                    COALESCE(SUM(COALESCE(m.votos_emitidos, 0)), 0) AS votos_emitidos,
                    COALESCE(SUM(COALESCE(m.votos_validos, 0)), 0) AS votos_validos
                FROM mesas_data m
                LEFT JOIN foreign_catalog f ON f.ubigeo = m.ubigeo
                LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                WHERE COALESCE(m.ubigeo, '') LIKE ?
                GROUP BY m.ubigeo, ul.departamento, f.continente, f.pais, ul.pais, f.ciudad, ul.ciudad
                ORDER BY mesas DESC, votos_validos DESC, m.ubigeo ASC
                LIMIT ?
                """,
                (like_value, sample_size),
            ).fetchall()

        sample: list[dict[str, Any]] = []
        for row in sample_rows:
            sample.append(
                {
                    "ubigeo": str(row["ubigeo"] or ""),
                    "departamento": str(row["departamento"] or ""),
                    "continente": str(row["continente"] or ""),
                    "pais": str(row["pais"] or ""),
                    "ciudad": str(row["ciudad"] or ""),
                    "mesas": int(row["mesas"] or 0),
                    "votos_emitidos": int(row["votos_emitidos"] or 0),
                    "votos_validos": int(row["votos_validos"] or 0),
                }
            )

        return {
            "ubigeo_prefix": prefix,
            "sample": sample,
        }

    def find_foreign_ubigeos(self, query: str, field: str | None = None) -> list[dict[str, str]]:
        q_norm = self._norm(query)
        if not q_norm:
            return []

        field_norm = self._norm(field or "")
        allowed_fields = {"pais", "ciudad", "any"}
        if field_norm and field_norm not in allowed_fields:
            raise ValueError("field debe ser 'pais', 'ciudad' o None")

        with self._connect() as conn:
            rows = conn.execute(
                "SELECT ubigeo, continente, pais, ciudad FROM foreign_catalog"
            ).fetchall()

        def _project(row: sqlite3.Row) -> dict[str, str]:
            return {
                "ubigeo": str(row["ubigeo"]),
                "Continente": str(row["continente"] or ""),
                "pais": str(row["pais"] or ""),
                "ciudad": str(row["ciudad"] or ""),
            }

        exact: list[dict[str, str]] = []
        partial: list[dict[str, str]] = []

        for row in rows:
            pais = str(row["pais"] or "")
            ciudad = str(row["ciudad"] or "")
            pais_norm = self._norm(pais)
            ciudad_norm = self._norm(ciudad)

            if field_norm == "pais":
                is_exact = pais_norm == q_norm
                is_partial = q_norm in pais_norm
            elif field_norm == "ciudad":
                is_exact = ciudad_norm == q_norm
                is_partial = q_norm in ciudad_norm
            else:
                is_exact = pais_norm == q_norm or ciudad_norm == q_norm
                is_partial = q_norm in pais_norm or q_norm in ciudad_norm

            if is_exact:
                exact.append(_project(row))
            elif is_partial:
                partial.append(_project(row))

        # Prioriza coincidencia exacta para evitar mezclar entidades parecidas.
        return exact if exact else partial

    def total_mesas_local(self) -> int:
        with self._connect() as conn:
            row = conn.execute("SELECT COUNT(*) AS c FROM mesas_data").fetchone()
        return int(row["c"] if row else 0)

    def get_geo_query_cache(self, query_key: str, max_age_seconds: int) -> dict[str, Any] | None:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT payload_json, fetched_at FROM geo_query_cache WHERE query_key = ?",
                (query_key,),
            ).fetchone()

        if row is None:
            return None

        fetched_at = datetime.fromisoformat(str(row["fetched_at"]).replace("Z", "+00:00"))
        age_seconds = (datetime.now(timezone.utc) - fetched_at).total_seconds()
        if age_seconds > max_age_seconds:
            return None

        return json.loads(str(row["payload_json"]))

    def upsert_geo_query_cache(self, query_key: str, payload: dict[str, Any]) -> None:
        now = self.now_iso()
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO geo_query_cache (query_key, payload_json, fetched_at)
                VALUES (?, ?, ?)
                ON CONFLICT(query_key) DO UPDATE SET
                    payload_json=excluded.payload_json,
                    fetched_at=excluded.fetched_at
                """,
                (query_key, json.dumps(payload, ensure_ascii=False), now),
            )

    def fill_ubigeo_location_cache_from_reniec(self, source_dir: Path) -> dict[str, int]:
        """Populates ubigeo_location_cache from geodir-ubigeo-reniec.xlsx.

        Strategy:
        1. Load all 6-digit ubigeos from Excel → ubigeo_location_cache.
        2. For 5-digit ubigeos in mesas_data, try zero-padded match.
        3. For still-missing ones, fill by province prefix (first 4 digits).
        """
        try:
            import openpyxl  # type: ignore[import]
        except ImportError:
            return {"excel": 0, "padded": 0, "prefix": 0}

        xlsx_path = source_dir / "geodir-ubigeo-reniec.xlsx"
        if not xlsx_path.exists():
            return {"excel": 0, "padded": 0, "prefix": 0}

        now = self.now_iso()
        excel_inserted = 0
        padded_fixed = 0
        prefix_fixed = 0

        try:
            wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]  # type: ignore

            def _col(name: str) -> int:
                return headers.index(name)

            ubigeo_col = _col("ubigeo")
            distrito_col = _col("distrito")
            provincia_col = _col("provincia")
            departamento_col = _col("departamento")

            with self._connect() as conn:
                # Step 1: load from Excel
                for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore
                    ubigeo = str(row[ubigeo_col] or "").strip()
                    if not ubigeo:
                        continue
                    distrito = str(row[distrito_col] or "").strip()
                    provincia = str(row[provincia_col] or "").strip()
                    departamento = str(row[departamento_col] or "").strip()
                    ciudad = f"{distrito}, {provincia}" if provincia else distrito
                    conn.execute(
                        """INSERT OR IGNORE INTO ubigeo_location_cache (ubigeo, ciudad, departamento, fetched_at)
                           VALUES (?, ?, ?, ?)""",
                        (ubigeo, ciudad, departamento, now),
                    )
                    excel_inserted += 1

                # Step 2: fill 5-digit ubigeos using zero-padded lookup
                missing = conn.execute("""
                    SELECT DISTINCT m.ubigeo FROM mesas_data m
                    LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                    LEFT JOIN foreign_catalog fc ON fc.ubigeo = m.ubigeo
                    WHERE ul.ubigeo IS NULL AND fc.ubigeo IS NULL AND m.ubigeo != ''
                """).fetchall()

                for (ubigeo,) in missing:
                    padded = ubigeo.zfill(6)
                    if padded == ubigeo:
                        continue
                    row_data = conn.execute(
                        "SELECT ciudad, departamento FROM ubigeo_location_cache WHERE ubigeo=?",
                        (padded,),
                    ).fetchone()
                    if row_data:
                        conn.execute(
                            """INSERT OR IGNORE INTO ubigeo_location_cache (ubigeo, ciudad, departamento, fetched_at)
                               VALUES (?, ?, ?, ?)""",
                            (ubigeo, row_data[0], row_data[1], now),
                        )
                        padded_fixed += 1

                # Step 3: fill remaining by province prefix (first 4 digits)
                still_missing = conn.execute("""
                    SELECT DISTINCT m.ubigeo FROM mesas_data m
                    LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                    LEFT JOIN foreign_catalog fc ON fc.ubigeo = m.ubigeo
                    WHERE ul.ubigeo IS NULL AND fc.ubigeo IS NULL AND m.ubigeo != ''
                """).fetchall()

                for (ubigeo,) in still_missing:
                    padded = ubigeo.zfill(6)
                    resolved = None
                    for prefix_len in [4, 2]:
                        prefix = padded[:prefix_len]
                        row_data = conn.execute(
                            """SELECT ciudad, departamento FROM ubigeo_location_cache
                               WHERE ubigeo LIKE ? AND departamento != '' LIMIT 1""",
                            (prefix + "%",),
                        ).fetchone()
                        if row_data:
                            resolved = row_data
                            break
                    if resolved:
                        conn.execute(
                            """INSERT OR IGNORE INTO ubigeo_location_cache (ubigeo, ciudad, departamento, fetched_at)
                               VALUES (?, ?, ?, ?)""",
                            (ubigeo, resolved[0], resolved[1], now),
                        )
                        prefix_fixed += 1

            wb.close()
            _logger.info(
                "ubigeo_location_cache: excel=%d padded=%d prefix=%d",
                excel_inserted, padded_fixed, prefix_fixed,
            )
        except Exception:
            _logger.exception("Error en fill_ubigeo_location_cache_from_reniec")

        return {"excel": excel_inserted, "padded": padded_fixed, "prefix": prefix_fixed}

    def try_bootstrap_reniec(self, source_dir: Path) -> int:
        """Pobla ubigeo_reniec desde geodir-ubigeo-reniec.xlsx (requiere openpyxl).
        Retorna filas insertadas/actualizadas. Silencia si openpyxl no está disponible.
        No re-lee el xlsx si la tabla ya tiene datos (evita costosa carga en arranque)."""
        # Skip rápido si ya hay filas en la tabla (opción de refresco: borrar tabla manualmente)
        try:
            with self._connect() as conn:
                cnt = conn.execute("SELECT COUNT(*) AS c FROM ubigeo_reniec").fetchone()["c"]
            if cnt > 0:
                return cnt
        except Exception:
            pass

        try:
            import openpyxl  # type: ignore[import]
        except ImportError:
            _logger.debug("openpyxl no disponible; salteando bootstrap de ubigeo_reniec")
            return 0

        xlsx_path = source_dir / "geodir-ubigeo-reniec.xlsx"
        if not xlsx_path.exists():
            _logger.debug("geodir-ubigeo-reniec.xlsx no encontrado en %s", source_dir)
            return 0

        try:
            wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(max_row=1))]  # type: ignore[arg-type]

            def _col(name: str) -> int:
                return headers.index(name)

            ubigeo_col = _col("ubigeo")
            distrito_col = _col("distrito")
            provincia_col = _col("provincia")
            departamento_col = _col("departamento")

            now = self.now_iso()
            inserted = 0
            with self._connect() as conn:
                for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore[union-attr]
                    ubigeo = str(row[ubigeo_col] or "").strip()
                    if not ubigeo:
                        continue
                    distrito = str(row[distrito_col] or "").strip()
                    provincia = str(row[provincia_col] or "").strip()
                    departamento = str(row[departamento_col] or "").strip()
                    conn.execute(
                        """
                        INSERT INTO ubigeo_reniec (
                            ubigeo, distrito, provincia, departamento,
                            distrito_norm, provincia_norm, departamento_norm, fetched_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(ubigeo) DO UPDATE SET
                            distrito=excluded.distrito,
                            provincia=excluded.provincia,
                            departamento=excluded.departamento,
                            distrito_norm=excluded.distrito_norm,
                            provincia_norm=excluded.provincia_norm,
                            departamento_norm=excluded.departamento_norm,
                            fetched_at=excluded.fetched_at
                        """,
                        (
                            ubigeo, distrito, provincia, departamento,
                            _norm_text(distrito), _norm_text(provincia), _norm_text(departamento),
                            now,
                        ),
                    )
                    inserted += 1

            wb.close()
            _logger.info("ubigeo_reniec: %d filas desde %s", inserted, xlsx_path)
            return inserted
        except Exception:
            _logger.exception("Error en try_bootstrap_reniec")
            return 0

    def rebuild_prefix_summaries(self, prefix_lengths: tuple[int, ...] = (1, 2, 3, 4)) -> dict[str, int]:
        """Pre-computes mesa_prefix_totals and mesa_prefix_party_summary for fast prefix queries.

        Covering prefixes of 1–4 digits lets all '9%', '90%', '900%', '9000%' queries
        skip the 3.8M-row votos scan and return in <1ms.
        """
        now = self.now_iso()
        totals_inserted = 0
        party_inserted = 0

        with self._connect() as conn:
            conn.execute("DELETE FROM mesa_prefix_totals")
            conn.execute("DELETE FROM mesa_prefix_party_summary")

            # Pre-compute the winner of each mesa (partido with max votes)
            conn.execute("DELETE FROM mesa_winner")
            conn.execute(
                """
                INSERT INTO mesa_winner (codigo_mesa, partido_id, max_votos)
                SELECT v.codigo_mesa, v.partido_id, v.votos
                FROM votos v
                INNER JOIN (
                    SELECT codigo_mesa, MAX(votos) AS max_votos
                    FROM votos GROUP BY codigo_mesa
                ) mx ON mx.codigo_mesa = v.codigo_mesa AND mx.max_votos = v.votos
                ON CONFLICT(codigo_mesa) DO UPDATE SET
                    partido_id=excluded.partido_id,
                    max_votos=excluded.max_votos
                """
            )
            _logger.info("mesa_winner: populated")

            for length in prefix_lengths:
                # Totals per prefix
                conn.execute(
                    f"""
                    INSERT INTO mesa_prefix_totals (prefix, n_mesas, mesas_con_votos, votos_emitidos, votos_validos, rebuilt_at)
                    SELECT
                        SUBSTR(codigo_mesa, 1, {length}) AS prefix,
                        COUNT(*) AS n_mesas,
                        SUM(CASE WHEN votos_emitidos > 0 THEN 1 ELSE 0 END) AS mesas_con_votos,
                        SUM(COALESCE(votos_emitidos, 0)) AS votos_emitidos,
                        SUM(COALESCE(votos_validos, 0)) AS votos_validos,
                        ? AS rebuilt_at
                    FROM mesas_data
                    GROUP BY 1
                    ON CONFLICT(prefix) DO UPDATE SET
                        n_mesas=excluded.n_mesas,
                        mesas_con_votos=excluded.mesas_con_votos,
                        votos_emitidos=excluded.votos_emitidos,
                        votos_validos=excluded.votos_validos,
                        rebuilt_at=excluded.rebuilt_at
                    """,
                    (now,),
                )
                n = conn.execute(f"SELECT CHANGES()").fetchone()[0]
                totals_inserted += n

                # Party aggregation per prefix
                conn.execute(
                    f"""
                    INSERT INTO mesa_prefix_party_summary (prefix, partido_id, total_votos, n_mesas)
                    SELECT
                        SUBSTR(v.codigo_mesa, 1, {length}) AS prefix,
                        v.partido_id,
                        SUM(v.votos) AS total_votos,
                        COUNT(DISTINCT v.codigo_mesa) AS n_mesas
                    FROM votos v
                    GROUP BY 1, 2
                    ON CONFLICT(prefix, partido_id) DO UPDATE SET
                        total_votos=excluded.total_votos,
                        n_mesas=excluded.n_mesas
                    """,
                )
                n = conn.execute("SELECT CHANGES()").fetchone()[0]
                party_inserted += n

        _logger.info("rebuild_prefix_summaries: totals=%d party=%d", totals_inserted, party_inserted)
        return {"totals": totals_inserted, "party": party_inserted}

    def bootstrap_from_onpescraper(
        self,
        *,
        output_dir: Path,
        source_dir: Path | None = None,
        include_votes: bool = True,
        source: str = "onpescraper_snapshot",
        id_eleccion: int = 10,
        force: bool = False,
    ) -> dict[str, int | bool]:
        if source_dir is None:
            source_dir = output_dir.parent / "source_data"
        candidate_map = self.load_candidate_map(source_dir / "candidato.txt")

        with self._connect() as conn:
            mesas_existing = int(
                (conn.execute("SELECT COUNT(*) AS c FROM mesas_data").fetchone() or {"c": 0})["c"]
            )
            votos_existing = int(
                (conn.execute("SELECT COUNT(*) AS c FROM votos").fetchone() or {"c": 0})["c"]
            )

            if not force and (mesas_existing > 0 or (include_votes and votos_existing > 0)):
                if candidate_map:
                    conn.executemany(
                        "UPDATE agrupaciones SET candidato = ? WHERE partido_id = ?",
                        [(str(cand or ""), str(pid)) for pid, cand in candidate_map.items()],
                    )
                reniec_inserted = self.try_bootstrap_reniec(source_dir)
                return {
                    "skipped": True,
                    "mesas": 0,
                    "votos": 0,
                    "agrupaciones": 0,
                    "foreign_catalog": 0,
                    "reniec_ubigeos": reniec_inserted,
                }

            now = self.now_iso()
            mesas_inserted = 0
            votos_inserted = 0
            agrupaciones_inserted = 0
            foreign_inserted = 0

            mesas_file = output_dir / "mesas_data.txt"
            if mesas_file.exists():
                with mesas_file.open("r", encoding="utf-8-sig", newline="") as handle:
                    reader = csv.DictReader(handle, delimiter="\t")
                    for row in reader:
                        codigo_mesa = str(row.get("codigo_mesa", "")).strip()
                        if not codigo_mesa:
                            continue
                        conn.execute(
                            """
                            INSERT INTO mesas_data (
                                codigo_mesa, ubigeo, local_votacion, electores_habiles, votos_emitidos, votos_validos,
                                blancos, nulos, impugnados, estado_acta, fetched_at
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ON CONFLICT(codigo_mesa) DO UPDATE SET
                                ubigeo=excluded.ubigeo,
                                local_votacion=excluded.local_votacion,
                                electores_habiles=excluded.electores_habiles,
                                votos_emitidos=excluded.votos_emitidos,
                                votos_validos=excluded.votos_validos,
                                blancos=excluded.blancos,
                                nulos=excluded.nulos,
                                impugnados=excluded.impugnados,
                                estado_acta=excluded.estado_acta,
                                fetched_at=excluded.fetched_at
                            """,
                            (
                                codigo_mesa,
                                str(row.get("ubigeo", "")).strip(),
                                str(row.get("local_votacion", "")).strip(),
                                int(str(row.get("electores_habiles", "0")).strip() or 0),
                                int(str(row.get("votos_emitidos", "0")).strip() or 0),
                                int(str(row.get("votos_validos", "0")).strip() or 0),
                                int(str(row.get("blancos", "0")).strip() or 0),
                                int(str(row.get("nulos", "0")).strip() or 0),
                                int(str(row.get("impugnados", "0")).strip() or 0),
                                str(row.get("estado_acta", "")).strip(),
                                now,
                            ),
                        )
                        mesas_inserted += 1

            agrupaciones_file = output_dir / "agrupaciones.txt"
            if agrupaciones_file.exists():
                with agrupaciones_file.open("r", encoding="utf-8-sig", newline="") as handle:
                    reader = csv.DictReader(handle, delimiter="\t")
                    for row in reader:
                        partido_id = str(row.get("partido_id", "")).strip()
                        if not partido_id:
                            continue
                        conn.execute(
                            """
                            INSERT INTO agrupaciones (partido_id, nombre, candidato, fetched_at)
                            VALUES (?, ?, ?, ?)
                            ON CONFLICT(partido_id) DO UPDATE SET
                                nombre=excluded.nombre,
                                candidato=excluded.candidato,
                                fetched_at=excluded.fetched_at
                            """,
                            (
                                partido_id,
                                str(row.get("nombre", "")).strip(),
                                str(candidate_map.get(partido_id, "") or ""),
                                now,
                            ),
                        )
                        agrupaciones_inserted += 1

            if include_votes:
                votos_file = output_dir / "votos.txt"
                if votos_file.exists():
                    with votos_file.open("r", encoding="utf-8-sig", newline="") as handle:
                        reader = csv.DictReader(handle, delimiter="\t")
                        for row in reader:
                            codigo_mesa = str(row.get("codigo_mesa", "")).strip()
                            partido_id = str(row.get("partido_id", "")).strip()
                            votos_raw = str(row.get("votos", "0")).strip()
                            # Saltar filas de cabecera repetidas o malformadas
                            if not codigo_mesa or not partido_id or not votos_raw.lstrip("-").isdigit():
                                continue
                            conn.execute(
                                """
                                INSERT INTO votos (codigo_mesa, partido_id, votos, fetched_at)
                                VALUES (?, ?, ?, ?)
                                ON CONFLICT(codigo_mesa, partido_id) DO UPDATE SET
                                    votos=excluded.votos,
                                    fetched_at=excluded.fetched_at
                                """,
                                (
                                    codigo_mesa,
                                    partido_id,
                                    int(votos_raw),
                                    now,
                                ),
                            )
                            votos_inserted += 1

            foreign_file = output_dir / "ubigeo_extranjero.txt"
            if foreign_file.exists():
                with foreign_file.open("r", encoding="utf-8-sig", newline="") as handle:
                    reader = csv.DictReader(handle, delimiter="\t")
                    for row in reader:
                        ubigeo = str(row.get("ubigeo", "")).strip()
                        if not ubigeo:
                            continue
                        conn.execute(
                            """
                            INSERT INTO foreign_catalog (ubigeo, continente, pais, ciudad, fetched_at)
                            VALUES (?, ?, ?, ?, ?)
                            ON CONFLICT(ubigeo) DO UPDATE SET
                                continente=excluded.continente,
                                pais=excluded.pais,
                                ciudad=excluded.ciudad,
                                fetched_at=excluded.fetched_at
                            """,
                            (
                                ubigeo,
                                str(row.get("Continente", "")).strip(),
                                str(row.get("pais", "")).strip(),
                                str(row.get("ciudad", "")).strip(),
                                now,
                            ),
                        )
                        foreign_inserted += 1

            conn.execute("DELETE FROM votos_by_ubigeo_partido")
            self._ensure_votes_by_ubigeo_partido_backfilled(conn)

            self.append_raw_event(
                "bootstrap_onpescraper_snapshot",
                {
                    "source": source,
                    "id_eleccion": id_eleccion,
                    "include_votes": include_votes,
                    "mesas": mesas_inserted,
                    "votos": votos_inserted,
                    "agrupaciones": agrupaciones_inserted,
                    "foreign_catalog": foreign_inserted,
                },
            )

        reniec_inserted = self.try_bootstrap_reniec(source_dir)
        self.fill_ubigeo_location_cache_from_reniec(source_dir)
        self.rebuild_prefix_summaries()

        return {
            "skipped": False,
            "mesas": mesas_inserted,
            "votos": votos_inserted,
            "agrupaciones": agrupaciones_inserted,
            "foreign_catalog": foreign_inserted,
            "reniec_ubigeos": reniec_inserted,
        }

    def bootstrap_from_atu_manera_csv(
        self,
        csv_path: Path | None = None,
        *,
        url: str = (
            "https://media.githubusercontent.com/media/ATuManera/"
            "Peru_elecciones2026/main/data/output/por_votacion/mesas_presidencial.csv"
        ),
        id_eleccion: int = 12,
        force: bool = False,
        timeout: int = 300,
    ) -> dict[str, int | bool]:
        """Carga 92,766 mesas presidenciales desde el CSV público de ATuManera/Peru_elecciones2026.

        Si csv_path es None, intenta descargarlo desde `url` (archivo Git-LFS).
        Requiere que curl_cffi o urllib esté disponible para la descarga.

        Columns esperadas: codigoMesa, ubigeoNivel03, nombreLocalVotacion,
        totalElectoresHabiles, totalVotosEmitidos, totalVotosValidos, estadoActa,
        detalle_N_nposicion, detalle_N_nvotos, detalle_N_descripcion (N=1..82).
        """
        with self._connect() as conn:
            mesas_existing = int(
                (conn.execute("SELECT COUNT(*) AS c FROM mesas_data").fetchone() or {"c": 0})["c"]
            )
            if not force and mesas_existing > 0:
                return {"skipped": True, "mesas": 0, "votos": 0, "agrupaciones": 0}

        # Descarga si no se proveyó ruta local
        _tmp_path: Path | None = None
        if csv_path is None or not csv_path.exists():
            import tempfile
            try:
                try:
                    from curl_cffi import requests as cffi_req
                    resp = cffi_req.get(url, impersonate="chrome124", timeout=timeout, stream=True)
                    resp.raise_for_status()
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
                    for chunk in resp.iter_content(chunk_size=1 << 20):
                        tmp.write(chunk)
                    tmp.close()
                    _tmp_path = Path(tmp.name)
                    csv_path = _tmp_path
                except ImportError:
                    from urllib.request import urlretrieve
                    tmp_file, _ = urlretrieve(url)  # noqa: S310
                    _tmp_path = Path(tmp_file)
                    csv_path = _tmp_path
            except Exception as exc:
                _logger.warning("bootstrap_from_atu_manera_csv: descarga falló: %s", exc)
                return {"skipped": True, "mesas": 0, "votos": 0, "agrupaciones": 0, "error": str(exc)}

        now = self.now_iso()
        mesas_inserted = 0
        votos_inserted = 0
        agrupaciones_seen: dict[str, str] = {}  # partido_id → nombre

        try:
            with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
                reader = csv.DictReader(fh)
                fieldnames = reader.fieldnames or []
                # Detectar columnas detalle_N_* dinámicamente
                detalle_indices: list[int] = sorted(
                    {
                        int(m.group(1))
                        for name in fieldnames
                        if (m := __import__("re").match(r"detalle_(\d+)_nvotos", name))
                    }
                )
                with self._connect() as conn:
                    for row in reader:
                        codigo_mesa = str(row.get("codigoMesa", "")).strip().zfill(6)
                        if not codigo_mesa or codigo_mesa == "000000":
                            continue
                        ubigeo = str(row.get("ubigeoNivel03", "")).strip()
                        local_votacion = str(row.get("nombreLocalVotacion", "")).strip()
                        electores = int(str(row.get("totalElectoresHabiles", "0")).strip() or 0)
                        votos_emitidos = int(str(row.get("totalVotosEmitidos", "0")).strip() or 0)
                        votos_validos = int(str(row.get("totalVotosValidos", "0")).strip() or 0)
                        estado_acta = str(row.get("estadoActa", "")).strip()

                        # Calcular blancos + nulos + impugnados desde detalle 80/81/82
                        blancos = int(str(row.get("detalle_80_nvotos", "0")).strip() or 0)
                        nulos = int(str(row.get("detalle_81_nvotos", "0")).strip() or 0)
                        impugnados = int(str(row.get("detalle_82_nvotos", "0")).strip() or 0)

                        conn.execute(
                            """INSERT INTO mesas_data (
                                codigo_mesa, ubigeo, local_votacion, electores_habiles,
                                votos_emitidos, votos_validos, blancos, nulos, impugnados,
                                estado_acta, fetched_at
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ON CONFLICT(codigo_mesa) DO UPDATE SET
                                ubigeo=excluded.ubigeo,
                                local_votacion=excluded.local_votacion,
                                electores_habiles=excluded.electores_habiles,
                                votos_emitidos=excluded.votos_emitidos,
                                votos_validos=excluded.votos_validos,
                                blancos=excluded.blancos,
                                nulos=excluded.nulos,
                                impugnados=excluded.impugnados,
                                estado_acta=excluded.estado_acta,
                                fetched_at=excluded.fetched_at""",
                            (
                                codigo_mesa, ubigeo, local_votacion, electores,
                                votos_emitidos, votos_validos, blancos, nulos, impugnados,
                                estado_acta, now,
                            ),
                        )
                        mesas_inserted += 1

                        # Insertar votos por partido (columnas detalle_N_*)
                        for n in detalle_indices:
                            pid_raw = str(row.get(f"detalle_{n}_nposicion", "")).strip()
                            votos_raw = str(row.get(f"detalle_{n}_nvotos", "")).strip()
                            nombre = str(row.get(f"detalle_{n}_descripcion", "")).strip()
                            if not pid_raw or not votos_raw:
                                continue
                            try:
                                pid = str(int(pid_raw))
                                nvotos = int(votos_raw)
                            except ValueError:
                                continue
                            if pid not in agrupaciones_seen and nombre:
                                agrupaciones_seen[pid] = nombre
                            conn.execute(
                                """INSERT INTO votos (codigo_mesa, partido_id, votos, fetched_at)
                                   VALUES (?, ?, ?, ?)
                                   ON CONFLICT(codigo_mesa, partido_id) DO UPDATE SET
                                       votos=excluded.votos, fetched_at=excluded.fetched_at""",
                                (codigo_mesa, pid, nvotos, now),
                            )
                            votos_inserted += 1

                    # Insertar agrupaciones descubiertas
                    for pid, nombre in agrupaciones_seen.items():
                        conn.execute(
                            """INSERT INTO agrupaciones (partido_id, nombre, fetched_at)
                               VALUES (?, ?, ?)
                               ON CONFLICT(partido_id) DO UPDATE SET
                                   nombre=excluded.nombre, fetched_at=excluded.fetched_at""",
                            (pid, nombre, now),
                        )
                    conn.execute("DELETE FROM votos_by_ubigeo_partido")
                    self._ensure_votes_by_ubigeo_partido_backfilled(conn)
        finally:
            if _tmp_path and _tmp_path.exists():
                try:
                    _tmp_path.unlink()
                except OSError:
                    pass

        self.append_raw_event(
            "bootstrap_atu_manera_csv",
            {"id_eleccion": id_eleccion, "mesas": mesas_inserted, "votos": votos_inserted,
             "agrupaciones": len(agrupaciones_seen)},
        )
        return {
            "skipped": False,
            "mesas": mesas_inserted,
            "votos": votos_inserted,
            "agrupaciones": len(agrupaciones_seen),
        }

    def bootstrap_elecciones_2021(
        self,
        repo_root: Path,
        *,
        force: bool = False,
    ) -> dict[str, Any]:
        """Hidrata SQLite con la data oficial 2021 (1V y 2V) desde CSV peruvoto2021.

        Auto-detects stale caches: if the in-code `_PARTY_MAP_2021_*` fingerprint
        no longer matches the value stored in `sv_sync_meta` (e.g. after pulling
        a fix), the cache is treated as invalid and a full re-hydration is
        forced even when `force=False`. This prevents silent data drift when
        the party/candidate mapping changes between repo versions.
        """
        data_dir = repo_root / "data"
        file_1v = data_dir / "Resultados_1ra_vuelta_Version_PCM.csv"
        file_2v = data_dir / "Resultados_2da_vuelta_Version_PCM.csv"
        if not file_1v.exists() or not file_2v.exists():
            raise FileNotFoundError(
                "No se encontraron los CSV 2021 en peruvoto2021/data "
                "(Resultados_1ra_vuelta_Version_PCM.csv y Resultados_2da_vuelta_Version_PCM.csv)."
            )

        current_fingerprint = _compute_party_map_2021_fingerprint()
        stored_fingerprint = self.get_sv_sync_meta(_PARTY_MAP_2021_FINGERPRINT_KEY)
        fingerprint_mismatch = (
            stored_fingerprint is not None and stored_fingerprint != current_fingerprint
        )

        with self._connect() as conn:
            c1 = int((conn.execute("SELECT COUNT(*) AS c FROM mesas_2021 WHERE vuelta = 1").fetchone() or {"c": 0})["c"])
            c2 = int((conn.execute("SELECT COUNT(*) AS c FROM mesas_2021 WHERE vuelta = 2").fetchone() or {"c": 0})["c"])
            if not force and c1 > 0 and c2 > 0:
                if not fingerprint_mismatch:
                    return {
                        "skipped": True,
                        "vuelta1_mesas": c1,
                        "vuelta2_mesas": c2,
                        "votos": 0,
                        "partidos": 0,
                        "fingerprint": current_fingerprint,
                    }
                _logger.warning(
                    "party_map_2021 fingerprint mismatch (stored=%s current=%s) — "
                    "forcing re-hydration to refresh stale cache.",
                    stored_fingerprint,
                    current_fingerprint,
                )
                force = True

        now = self.now_iso()
        votos_inserted = 0
        mesas_1v = 0
        mesas_2v = 0

        with self._connect() as conn:
            if force:
                conn.execute("DELETE FROM votos_2021")
                conn.execute("DELETE FROM mesas_2021")
                conn.execute("DELETE FROM partidos_2021")

            for col, (pid, partido, candidato) in _PARTY_MAP_2021_1V.items():
                conn.execute(
                    """
                    INSERT INTO partidos_2021 (vuelta, partido_id, nombre_partido, candidato, fetched_at)
                    VALUES (1, ?, ?, ?, ?)
                    ON CONFLICT(vuelta, partido_id) DO UPDATE SET
                        nombre_partido=excluded.nombre_partido,
                        candidato=excluded.candidato,
                        fetched_at=excluded.fetched_at
                    """,
                    (pid, partido, candidato, now),
                )
            for col, (pid, partido, candidato) in _PARTY_MAP_2021_2V.items():
                conn.execute(
                    """
                    INSERT INTO partidos_2021 (vuelta, partido_id, nombre_partido, candidato, fetched_at)
                    VALUES (2, ?, ?, ?, ?)
                    ON CONFLICT(vuelta, partido_id) DO UPDATE SET
                        nombre_partido=excluded.nombre_partido,
                        candidato=excluded.candidato,
                        fetched_at=excluded.fetched_at
                    """,
                    (pid, partido, candidato, now),
                )

            def _load_round(csv_file: Path, vuelta: int, party_map: dict[str, tuple[str, str, str]]) -> tuple[int, int]:
                mesas_local = 0
                votos_local = 0
                with csv_file.open("r", encoding="cp1252", newline="") as fh:
                    reader = csv.DictReader(fh, delimiter=";")
                    for row in reader:
                        codigo_mesa = str(row.get("MESA_DE_VOTACION") or "").strip().replace('"', "")
                        if not codigo_mesa:
                            continue
                        codigo_mesa = codigo_mesa.zfill(6)
                        ubigeo = str(row.get("UBIGEO") or "").strip().replace('"', "")
                        departamento = str(row.get("DEPARTAMENTO") or "").strip().replace('"', "")
                        provincia = str(row.get("PROVINCIA") or "").strip().replace('"', "")
                        distrito = str(row.get("DISTRITO") or "").strip().replace('"', "")
                        tipo_eleccion = str(row.get("TIPO_ELECCION") or "").strip().replace('"', "")
                        descrip_estado = str(row.get("DESCRIP_ESTADO_ACTA") or "").strip().replace('"', "")
                        tipo_obs = str(row.get("TIPO_OBSERVACION") or "").strip().replace('"', "")
                        n_cvas = _to_int_safe(row.get("N_CVAS"))
                        n_elec_habil = _to_int_safe(row.get("N_ELEC_HABIL"))
                        votos_vb = _to_int_safe(row.get("VOTOS_VB"))
                        votos_vn = _to_int_safe(row.get("VOTOS_VN"))
                        votos_vi = _to_int_safe(row.get("VOTOS_VI"))

                        votos_validos = 0
                        party_votes: list[tuple[str, int]] = []
                        for col, (pid, _, _) in party_map.items():
                            vv = _to_int_safe(row.get(col))
                            votos_validos += vv
                            party_votes.append((pid, vv))
                        votos_emitidos = votos_validos + votos_vb + votos_vn + votos_vi

                        conn.execute(
                            """
                            INSERT INTO mesas_2021 (
                                vuelta, codigo_mesa, ubigeo, departamento, provincia, distrito, tipo_eleccion,
                                descrip_estado_acta, tipo_observacion, n_cvas, n_elec_habil,
                                votos_vb, votos_vn, votos_vi, votos_emitidos, votos_validos, fetched_at
                            )
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ON CONFLICT(vuelta, codigo_mesa) DO UPDATE SET
                                ubigeo=excluded.ubigeo,
                                departamento=excluded.departamento,
                                provincia=excluded.provincia,
                                distrito=excluded.distrito,
                                tipo_eleccion=excluded.tipo_eleccion,
                                descrip_estado_acta=excluded.descrip_estado_acta,
                                tipo_observacion=excluded.tipo_observacion,
                                n_cvas=excluded.n_cvas,
                                n_elec_habil=excluded.n_elec_habil,
                                votos_vb=excluded.votos_vb,
                                votos_vn=excluded.votos_vn,
                                votos_vi=excluded.votos_vi,
                                votos_emitidos=excluded.votos_emitidos,
                                votos_validos=excluded.votos_validos,
                                fetched_at=excluded.fetched_at
                            """,
                            (
                                vuelta, codigo_mesa, ubigeo, departamento, provincia, distrito, tipo_eleccion,
                                descrip_estado, tipo_obs, n_cvas, n_elec_habil,
                                votos_vb, votos_vn, votos_vi, votos_emitidos, votos_validos, now,
                            ),
                        )
                        mesas_local += 1

                        for pid, vv in party_votes:
                            conn.execute(
                                """
                                INSERT INTO votos_2021 (vuelta, codigo_mesa, partido_id, votos, fetched_at)
                                VALUES (?, ?, ?, ?, ?)
                                ON CONFLICT(vuelta, codigo_mesa, partido_id) DO UPDATE SET
                                    votos=excluded.votos,
                                    fetched_at=excluded.fetched_at
                                """,
                                (vuelta, codigo_mesa, pid, vv, now),
                            )
                            votos_local += 1
                return mesas_local, votos_local

            mesas_1v, votos1 = _load_round(file_1v, 1, _PARTY_MAP_2021_1V)
            mesas_2v, votos2 = _load_round(file_2v, 2, _PARTY_MAP_2021_2V)
            votos_inserted = votos1 + votos2

        # Persist the fingerprint AFTER a successful bootstrap so subsequent
        # calls can detect staleness when the in-code mapping evolves.
        self.set_sv_sync_meta(_PARTY_MAP_2021_FINGERPRINT_KEY, current_fingerprint)

        self.append_raw_event(
            "bootstrap_2021_csv",
            {
                "repo_root": str(repo_root),
                "vuelta1_mesas": mesas_1v,
                "vuelta2_mesas": mesas_2v,
                "votos": votos_inserted,
                "partidos": len(_PARTY_MAP_2021_1V) + len(_PARTY_MAP_2021_2V),
                "fingerprint": current_fingerprint,
                "previous_fingerprint": stored_fingerprint,
            },
        )
        return {
            "skipped": False,
            "vuelta1_mesas": mesas_1v,
            "vuelta2_mesas": mesas_2v,
            "votos": votos_inserted,
            "partidos": len(_PARTY_MAP_2021_1V) + len(_PARTY_MAP_2021_2V),
            "fingerprint": current_fingerprint,
        }

    def total_mesas_2021(self, vuelta: int | None = None) -> int:
        with self._connect() as conn:
            if vuelta in (1, 2):
                row = conn.execute("SELECT COUNT(*) AS c FROM mesas_2021 WHERE vuelta = ?", (int(vuelta),)).fetchone()
            else:
                row = conn.execute("SELECT COUNT(*) AS c FROM mesas_2021").fetchone()
        return int(row["c"] if row else 0)

    def get_mesa_2021_from_local(self, codigo_mesa: str, vuelta: int | None = None) -> dict[str, Any] | None:
        with self._connect() as conn:
            if vuelta in (1, 2):
                m = conn.execute(
                    "SELECT * FROM mesas_2021 WHERE codigo_mesa = ? AND vuelta = ?",
                    (codigo_mesa, int(vuelta)),
                ).fetchone()
            else:
                m = conn.execute(
                    "SELECT * FROM mesas_2021 WHERE codigo_mesa = ? ORDER BY vuelta ASC LIMIT 1",
                    (codigo_mesa,),
                ).fetchone()
            if m is None:
                return None
            rows = conn.execute(
                """
                SELECT v.partido_id, v.votos, p.nombre_partido, p.candidato
                FROM votos_2021 v
                LEFT JOIN partidos_2021 p ON p.vuelta = v.vuelta AND p.partido_id = v.partido_id
                WHERE v.vuelta = ? AND v.codigo_mesa = ?
                ORDER BY v.votos DESC
                """,
                (int(m["vuelta"]), codigo_mesa),
            ).fetchall()
        return {
            "vuelta": int(m["vuelta"] or 0),
            "codigo_mesa": str(m["codigo_mesa"] or ""),
            "ubigeo": str(m["ubigeo"] or ""),
            "departamento": str(m["departamento"] or ""),
            "provincia": str(m["provincia"] or ""),
            "distrito": str(m["distrito"] or ""),
            "estado_acta": str(m["descrip_estado_acta"] or ""),
            "electores_habiles": int(m["n_elec_habil"] or 0),
            "votos_emitidos": int(m["votos_emitidos"] or 0),
            "votos_validos": int(m["votos_validos"] or 0),
            "blancos": int(m["votos_vb"] or 0),
            "nulos": int(m["votos_vn"] or 0),
            "impugnados": int(m["votos_vi"] or 0),
            "votos": [
                {
                    "partido_id": str(r["partido_id"] or ""),
                    "partido": str(r["nombre_partido"] or ""),
                    "candidato": str(r["candidato"] or ""),
                    "votos": int(r["votos"] or 0),
                }
                for r in rows
            ],
        }

    def _resolve_geo_filter_2021(self, vuelta: int, geo_query: str | None) -> tuple[str | None, str | None]:
        if not geo_query:
            return None, None
        target = _norm_text(str(geo_query))
        if not target:
            return None, None
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT DISTINCT departamento, provincia, distrito
                FROM mesas_2021
                WHERE vuelta = ?
                """,
                (vuelta,),
            ).fetchall()
        dept_map: dict[str, str] = {}
        prov_map: dict[str, str] = {}
        dist_map: dict[str, str] = {}
        for r in rows:
            d = str(r["departamento"] or "").strip()
            p = str(r["provincia"] or "").strip()
            t = str(r["distrito"] or "").strip()
            if d:
                dept_map[_norm_text(d)] = d
            if p:
                prov_map[_norm_text(p)] = p
            if t:
                dist_map[_norm_text(t)] = t

        if target in dept_map:
            return "departamento", dept_map[target]
        if target in prov_map:
            return "provincia", prov_map[target]
        if target in dist_map:
            return "distrito", dist_map[target]

        for k, v in dept_map.items():
            if target in k:
                return "departamento", v
        for k, v in prov_map.items():
            if target in k:
                return "provincia", v
        for k, v in dist_map.items():
            if target in k:
                return "distrito", v
        return None, None

    def aggregate_votes_2021(
        self,
        *,
        vuelta: int,
        geo_query: str | None = None,
        mesa_prefix: str | None = None,
        top_n: int = 10,
    ) -> dict[str, Any]:
        vuelta = 1 if int(vuelta) == 1 else 2
        top_n = max(1, min(int(top_n), 30))
        col, val = self._resolve_geo_filter_2021(vuelta, geo_query)
        # Use denorm fast-path only when no filters at all
        if self.denorm_available and not col and not mesa_prefix:
            try:
                conn = self._connect_denorm()
                rows = conn.execute("""
                    SELECT partido_id, nombre_partido, candidato, votos
                    FROM fact_votos_nacional
                    WHERE election_year=2021 AND vuelta=? AND es_especial=0
                    ORDER BY votos DESC
                """, (vuelta,)).fetchall()
                conn.close()
                return {
                    "vuelta": vuelta,
                    "nivel": "nacional",
                    "filtro": None,
                    "mesa_prefix": None,
                    "mesas": 0,
                    "votos_emitidos": 0,
                    "top": [
                        {
                            "partido_id": str(r["partido_id"] or ""),
                            "nombre_partido": str(r["nombre_partido"] or ""),
                            "candidato": str(r["candidato"] or ""),
                            "total_votos": int(r["votos"] or 0),
                        }
                        for r in rows
                    ],
                }
            except Exception as e:
                _logger.debug("denorm fast-path failed for aggregate_votes_2021, falling back to OLTP: %s", e)
        where = "WHERE v.vuelta = ?"
        params: list[Any] = [vuelta]
        where_m = "WHERE m.vuelta = ?"
        params_m: list[Any] = [vuelta]
        if col and val:
            where += f" AND m.{col} = ?"
            params.append(val)
            where_m += f" AND m.{col} = ?"
            params_m.append(val)
        if mesa_prefix:
            lp = self._mesa_prefix_like(mesa_prefix)
            pat = f"{lp}%"
            where += " AND m.codigo_mesa LIKE ?"
            params.append(pat)
            where_m += " AND m.codigo_mesa LIKE ?"
            params_m.append(pat)
        with self._connect() as conn:
            rows = conn.execute(
                f"""
                SELECT v.partido_id, COALESCE(p.nombre_partido,'') AS nombre_partido,
                       COALESCE(p.candidato,'') AS candidato, SUM(v.votos) AS total_votos
                FROM votos_2021 v
                JOIN mesas_2021 m ON m.vuelta = v.vuelta AND m.codigo_mesa = v.codigo_mesa
                LEFT JOIN partidos_2021 p ON p.vuelta = v.vuelta AND p.partido_id = v.partido_id
                {where}
                GROUP BY v.partido_id, p.nombre_partido, p.candidato
                ORDER BY total_votos DESC
                LIMIT ?
                """,
                [*params, top_n],
            ).fetchall()
            totals = conn.execute(
                f"SELECT COUNT(*) AS mesas, COALESCE(SUM(m.votos_emitidos),0) AS votos_emitidos FROM mesas_2021 m {where_m}",
                params_m,
            ).fetchone()
        return {
            "vuelta": vuelta,
            "nivel": col or ("mesa_prefix" if mesa_prefix else "nacional"),
            "filtro": val,
            "mesa_prefix": mesa_prefix,
            "mesas": int(totals["mesas"] or 0) if totals else 0,
            "votos_emitidos": int(totals["votos_emitidos"] or 0) if totals else 0,
            "top": [
                {
                    "partido_id": str(r["partido_id"] or ""),
                    "nombre_partido": str(r["nombre_partido"] or ""),
                    "candidato": str(r["candidato"] or ""),
                    "total_votos": int(r["total_votos"] or 0),
                }
                for r in rows
            ],
        }

    def get_candidate_votes_2021(
        self,
        *,
        vuelta: int,
        candidate_query: str,
        geo_query: str | None = None,
    ) -> dict[str, Any] | None:
        vuelta = 1 if int(vuelta) == 1 else 2
        cq = _norm_text(candidate_query)
        if not cq:
            return None
        with self._connect() as conn:
            party_rows = conn.execute(
                "SELECT partido_id, nombre_partido, candidato FROM partidos_2021 WHERE vuelta = ?",
                (vuelta,),
            ).fetchall()
        selected: sqlite3.Row | None = None
        for r in party_rows:
            cand = _norm_text(str(r["candidato"] or ""))
            par = _norm_text(str(r["nombre_partido"] or ""))
            if cq == cand or cq == par:
                selected = r
                break
        if selected is None:
            for r in party_rows:
                cand = _norm_text(str(r["candidato"] or ""))
                par = _norm_text(str(r["nombre_partido"] or ""))
                if cq in cand or cq in par:
                    selected = r
                    break
        if selected is None:
            return None

        col, val = self._resolve_geo_filter_2021(vuelta, geo_query)
        where = "WHERE v.vuelta = ? AND v.partido_id = ?"
        params: list[Any] = [vuelta, str(selected["partido_id"] or "")]
        if col and val:
            where += f" AND m.{col} = ?"
            params.append(val)
        with self._connect() as conn:
            row = conn.execute(
                f"""
                SELECT COALESCE(SUM(v.votos),0) AS votos
                FROM votos_2021 v
                JOIN mesas_2021 m ON m.vuelta = v.vuelta AND m.codigo_mesa = v.codigo_mesa
                {where}
                """,
                params,
            ).fetchone()
        return {
            "vuelta": vuelta,
            "partido_id": str(selected["partido_id"] or ""),
            "partido": str(selected["nombre_partido"] or ""),
            "candidato": str(selected["candidato"] or ""),
            "nivel": col or "nacional",
            "filtro": val,
            "votos": int(row["votos"] or 0) if row else 0,
        }

    # ────────────────────────────────────────────────────────────────────
    # RAW DATA EXPORT — para análisis estadístico arbitrario
    # ────────────────────────────────────────────────────────────────────

    _EXPORT_2021_MAX_LIMIT = 100_000

    def _build_geo_where_2021(
        self,
        *,
        vuelta: int,
        departamento: str | None,
        provincia: str | None,
        distrito: str | None,
        ubigeo_prefix: str | None,
        mesa_prefix: str | None,
    ) -> tuple[str, list[Any]]:
        """Common WHERE-clause builder for raw-data 2021 endpoints."""
        clauses: list[str] = ["vuelta = ?"]
        params: list[Any] = [int(vuelta)]
        if departamento:
            clauses.append("UPPER(departamento) = UPPER(?)")
            params.append(str(departamento).strip())
        if provincia:
            clauses.append("UPPER(provincia) = UPPER(?)")
            params.append(str(provincia).strip())
        if distrito:
            clauses.append("UPPER(distrito) = UPPER(?)")
            params.append(str(distrito).strip())
        if ubigeo_prefix:
            clauses.append("ubigeo LIKE ?")
            params.append(f"{str(ubigeo_prefix).strip()}%")
        if mesa_prefix:
            clauses.append("codigo_mesa LIKE ?")
            params.append(f"{str(mesa_prefix).strip()}%")
        return " AND ".join(clauses), params

    def export_mesas_2021(
        self,
        *,
        vuelta: int,
        departamento: str | None = None,
        provincia: str | None = None,
        distrito: str | None = None,
        ubigeo_prefix: str | None = None,
        mesa_prefix: str | None = None,
        limit: int = 1000,
        offset: int = 0,
    ) -> dict[str, Any]:
        """Raw mesa rows from 2021 (cabecera). Paginated.

        Returns a dict with:
            total      — full count matching filters
            offset     — current offset
            limit      — applied limit (capped)
            rows       — list of mesa dicts (cabecera fields only)
            schema     — list of field names in `rows`
        """
        v = 1 if int(vuelta) == 1 else 2
        limit_eff = max(1, min(int(limit), self._EXPORT_2021_MAX_LIMIT))
        offset_eff = max(0, int(offset))
        where, params = self._build_geo_where_2021(
            vuelta=v,
            departamento=departamento,
            provincia=provincia,
            distrito=distrito,
            ubigeo_prefix=ubigeo_prefix,
            mesa_prefix=mesa_prefix,
        )
        with self._connect() as conn:
            total_row = conn.execute(
                f"SELECT COUNT(*) AS c FROM mesas_2021 WHERE {where}", params
            ).fetchone()
            total = int(total_row["c"] or 0) if total_row else 0
            rows = conn.execute(
                f"""
                SELECT vuelta, codigo_mesa, ubigeo, departamento, provincia, distrito,
                       descrip_estado_acta AS estado_acta, tipo_observacion,
                       n_elec_habil AS electores_habiles,
                       votos_emitidos, votos_validos,
                       votos_vb AS blancos, votos_vn AS nulos, votos_vi AS impugnados
                FROM mesas_2021
                WHERE {where}
                ORDER BY codigo_mesa
                LIMIT ? OFFSET ?
                """,
                params + [limit_eff, offset_eff],
            ).fetchall()
        out_rows = [
            {
                "vuelta": int(r["vuelta"] or 0),
                "codigo_mesa": str(r["codigo_mesa"] or ""),
                "ubigeo": str(r["ubigeo"] or ""),
                "departamento": str(r["departamento"] or ""),
                "provincia": str(r["provincia"] or ""),
                "distrito": str(r["distrito"] or ""),
                "estado_acta": str(r["estado_acta"] or ""),
                "tipo_observacion": str(r["tipo_observacion"] or ""),
                "electores_habiles": int(r["electores_habiles"] or 0),
                "votos_emitidos": int(r["votos_emitidos"] or 0),
                "votos_validos": int(r["votos_validos"] or 0),
                "blancos": int(r["blancos"] or 0),
                "nulos": int(r["nulos"] or 0),
                "impugnados": int(r["impugnados"] or 0),
            }
            for r in rows
        ]
        schema = list(out_rows[0].keys()) if out_rows else [
            "vuelta", "codigo_mesa", "ubigeo", "departamento", "provincia", "distrito",
            "estado_acta", "tipo_observacion", "electores_habiles",
            "votos_emitidos", "votos_validos", "blancos", "nulos", "impugnados",
        ]
        return {
            "vuelta": v,
            "total": total,
            "offset": offset_eff,
            "limit": limit_eff,
            "returned": len(out_rows),
            "has_more": (offset_eff + len(out_rows)) < total,
            "schema": schema,
            "rows": out_rows,
        }

    def export_votos_2021(
        self,
        *,
        vuelta: int,
        partido_ids: list[str] | None = None,
        departamento: str | None = None,
        provincia: str | None = None,
        distrito: str | None = None,
        ubigeo_prefix: str | None = None,
        mesa_prefix: str | None = None,
        limit: int = 5000,
        offset: int = 0,
    ) -> dict[str, Any]:
        """Raw vote rows from 2021 (one row per mesa × partido). Paginated.

        Joins mesas_2021 to enrich each row with geo. Optional `partido_ids`
        filter restricts to specific partidos (e.g. ['PC','K']).
        """
        v = 1 if int(vuelta) == 1 else 2
        limit_eff = max(1, min(int(limit), self._EXPORT_2021_MAX_LIMIT))
        offset_eff = max(0, int(offset))
        # Reuse geo where on mesa side
        mesa_where, mesa_params = self._build_geo_where_2021(
            vuelta=v,
            departamento=departamento,
            provincia=provincia,
            distrito=distrito,
            ubigeo_prefix=ubigeo_prefix,
            mesa_prefix=mesa_prefix,
        )
        # Translate "vuelta" predicate to alias `m`
        mesa_where = mesa_where.replace("vuelta = ?", "m.vuelta = ?", 1) \
                               .replace("UPPER(departamento)", "UPPER(m.departamento)") \
                               .replace("UPPER(provincia)", "UPPER(m.provincia)") \
                               .replace("UPPER(distrito)", "UPPER(m.distrito)") \
                               .replace("ubigeo LIKE ?", "m.ubigeo LIKE ?") \
                               .replace("codigo_mesa LIKE ?", "m.codigo_mesa LIKE ?")
        # votos table needs its own vuelta predicate (use v.vuelta = ?) shared param
        full_where = "v.vuelta = ? AND " + mesa_where
        params: list[Any] = [v] + mesa_params
        if partido_ids:
            placeholders = ",".join("?" for _ in partido_ids)
            full_where += f" AND v.partido_id IN ({placeholders})"
            params += [str(p) for p in partido_ids]
        with self._connect() as conn:
            total_row = conn.execute(
                f"""
                SELECT COUNT(*) AS c
                FROM votos_2021 v
                JOIN mesas_2021 m ON m.vuelta = v.vuelta AND m.codigo_mesa = v.codigo_mesa
                WHERE {full_where}
                """,
                params,
            ).fetchone()
            total = int(total_row["c"] or 0) if total_row else 0
            rows = conn.execute(
                f"""
                SELECT v.vuelta, v.codigo_mesa, v.partido_id, v.votos,
                       p.nombre_partido, p.candidato,
                       m.ubigeo, m.departamento, m.provincia, m.distrito,
                       m.votos_validos AS mesa_validos
                FROM votos_2021 v
                JOIN mesas_2021 m ON m.vuelta = v.vuelta AND m.codigo_mesa = v.codigo_mesa
                LEFT JOIN partidos_2021 p ON p.vuelta = v.vuelta AND p.partido_id = v.partido_id
                WHERE {full_where}
                ORDER BY v.codigo_mesa, v.partido_id
                LIMIT ? OFFSET ?
                """,
                params + [limit_eff, offset_eff],
            ).fetchall()
        out_rows = [
            {
                "vuelta": int(r["vuelta"] or 0),
                "codigo_mesa": str(r["codigo_mesa"] or ""),
                "partido_id": str(r["partido_id"] or ""),
                "nombre_partido": str(r["nombre_partido"] or ""),
                "candidato": str(r["candidato"] or ""),
                "votos": int(r["votos"] or 0),
                "ubigeo": str(r["ubigeo"] or ""),
                "departamento": str(r["departamento"] or ""),
                "provincia": str(r["provincia"] or ""),
                "distrito": str(r["distrito"] or ""),
                "mesa_votos_validos": int(r["mesa_validos"] or 0),
            }
            for r in rows
        ]
        schema = list(out_rows[0].keys()) if out_rows else [
            "vuelta", "codigo_mesa", "partido_id", "nombre_partido", "candidato",
            "votos", "ubigeo", "departamento", "provincia", "distrito",
            "mesa_votos_validos",
        ]
        return {
            "vuelta": v,
            "total": total,
            "offset": offset_eff,
            "limit": limit_eff,
            "returned": len(out_rows),
            "has_more": (offset_eff + len(out_rows)) < total,
            "schema": schema,
            "rows": out_rows,
        }

    def export_partidos_2021(self, *, vuelta: int | None = None) -> dict[str, Any]:
        """Catalog of partidos and candidates (no votes). Small enough to return whole."""
        with self._connect() as conn:
            if vuelta in (1, 2):
                rows = conn.execute(
                    "SELECT vuelta, partido_id, nombre_partido, candidato "
                    "FROM partidos_2021 WHERE vuelta = ? ORDER BY partido_id",
                    (int(vuelta),),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT vuelta, partido_id, nombre_partido, candidato "
                    "FROM partidos_2021 ORDER BY vuelta, partido_id"
                ).fetchall()
        out_rows = [
            {
                "vuelta": int(r["vuelta"] or 0),
                "partido_id": str(r["partido_id"] or ""),
                "nombre_partido": str(r["nombre_partido"] or ""),
                "candidato": str(r["candidato"] or ""),
            }
            for r in rows
        ]
        return {
            "vuelta": int(vuelta) if vuelta in (1, 2) else None,
            "total": len(out_rows),
            "schema": ["vuelta", "partido_id", "nombre_partido", "candidato"],
            "rows": out_rows,
        }

    def summary_2021(self, *, vuelta: int) -> dict[str, Any]:
        """Aggregated nacional summary for one round.

        Useful as a quick sanity check or as the 'global denominator' for any
        downstream statistical calc (HHI, participation, etc.).
        """
        v = 1 if int(vuelta) == 1 else 2
        if self.denorm_available:
            try:
                conn = self._connect_denorm()
                rows = conn.execute("""
                    SELECT partido_id, nombre_partido, candidato,
                           votos, pct_votos_validos
                    FROM fact_votos_nacional
                    WHERE election_year=2021 AND vuelta=? AND es_especial=0
                    ORDER BY votos DESC
                """, (v,)).fetchall()
                conn.close()
                return [dict(r) for r in rows]
            except Exception as e:
                _logger.debug("denorm fast-path failed for summary_2021, falling back to OLTP: %s", e)
        with self._connect() as conn:
            agg = conn.execute(
                """
                SELECT COUNT(*) AS mesas,
                       SUM(n_elec_habil) AS electores_habiles,
                       SUM(votos_emitidos) AS votos_emitidos,
                       SUM(votos_validos) AS votos_validos,
                       SUM(votos_vb) AS blancos,
                       SUM(votos_vn) AS nulos,
                       SUM(votos_vi) AS impugnados
                FROM mesas_2021
                WHERE vuelta = ?
                """,
                (v,),
            ).fetchone()
            party = conn.execute(
                """
                SELECT v.partido_id, p.nombre_partido, p.candidato,
                       SUM(v.votos) AS total
                FROM votos_2021 v
                LEFT JOIN partidos_2021 p
                  ON p.vuelta = v.vuelta AND p.partido_id = v.partido_id
                WHERE v.vuelta = ?
                GROUP BY v.partido_id
                ORDER BY total DESC
                """,
                (v,),
            ).fetchall()
        electores = int(agg["electores_habiles"] or 0) if agg else 0
        emit = int(agg["votos_emitidos"] or 0) if agg else 0
        validos = int(agg["votos_validos"] or 0) if agg else 0
        return {
            "vuelta": v,
            "mesas": int(agg["mesas"] or 0) if agg else 0,
            "electores_habiles": electores,
            "votos_emitidos": emit,
            "votos_validos": validos,
            "votos_blancos": int(agg["blancos"] or 0) if agg else 0,
            "votos_nulos": int(agg["nulos"] or 0) if agg else 0,
            "votos_impugnados": int(agg["impugnados"] or 0) if agg else 0,
            "participacion_pct": (emit / electores * 100.0) if electores else 0.0,
            "validez_pct": (validos / emit * 100.0) if emit else 0.0,
            "por_partido": [
                {
                    "partido_id": str(r["partido_id"] or ""),
                    "nombre_partido": str(r["nombre_partido"] or ""),
                    "candidato": str(r["candidato"] or ""),
                    "total_votos": int(r["total"] or 0),
                    "pct_validos": (int(r["total"] or 0) / validos * 100.0) if validos else 0.0,
                }
                for r in party
            ],
        }

    # ════════════════════════════════════════════════════════════════════════
    # RAW DATA EXPORT — 2026 (1V y 2V) + analítica + listados/auditoría
    #
    # Mismo patrón que los _2021. Soportan filtros geo/mesa_prefix + paginación
    # + schema explícito. Permiten consumir raw data sin lógica de agregación.
    # ════════════════════════════════════════════════════════════════════════

    _EXPORT_2026_MAX_LIMIT = 100_000

    # ── Helpers compartidos ─────────────────────────────────────────────────

    def _build_geo_where_2026_1v(
        self,
        *,
        departamento: str | None,
        provincia: str | None,
        distrito: str | None,
        ubigeo_prefix: str | None,
        mesa_prefix: str | None,
        estado_acta: str | None,
    ) -> tuple[str, list[Any], bool]:
        """Build WHERE clause for 2026 1V mesas (joins ubigeo_onpe_api when needed).

        Returns (where_sql, params, needs_geo_join).
        """
        clauses: list[str] = []
        params: list[Any] = []
        needs_geo_join = False
        if departamento:
            clauses.append("UPPER(g.departamento) = UPPER(?)")
            params.append(str(departamento).strip())
            needs_geo_join = True
        if provincia:
            clauses.append("UPPER(g.provincia) = UPPER(?)")
            params.append(str(provincia).strip())
            needs_geo_join = True
        if distrito:
            clauses.append("UPPER(g.distrito) = UPPER(?)")
            params.append(str(distrito).strip())
            needs_geo_join = True
        if ubigeo_prefix:
            clauses.append("m.ubigeo LIKE ?")
            params.append(f"{str(ubigeo_prefix).strip()}%")
        if mesa_prefix:
            clauses.append("m.codigo_mesa LIKE ?")
            params.append(f"{str(mesa_prefix).strip()}%")
        if estado_acta:
            clauses.append("UPPER(m.estado_acta) = UPPER(?)")
            params.append(str(estado_acta).strip())
        where = " AND ".join(clauses) if clauses else "1=1"
        return where, params, needs_geo_join

    def _build_geo_where_2026_sv(
        self,
        *,
        departamento: str | None,
        provincia: str | None,
        distrito: str | None,
        ubigeo_prefix: str | None,
        mesa_prefix: str | None,
        codigo_estado_acta: str | None,
    ) -> tuple[str, list[Any], bool]:
        """Build WHERE clause for 2V mesas (joins ubicaciones_sv when needed)."""
        clauses: list[str] = []
        params: list[Any] = []
        needs_geo_join = False
        if departamento:
            clauses.append("UPPER(u.departamento) = UPPER(?)")
            params.append(str(departamento).strip())
            needs_geo_join = True
        if provincia:
            clauses.append("UPPER(u.provincia) = UPPER(?)")
            params.append(str(provincia).strip())
            needs_geo_join = True
        if distrito:
            clauses.append("UPPER(u.distrito) = UPPER(?)")
            params.append(str(distrito).strip())
            needs_geo_join = True
        if ubigeo_prefix:
            clauses.append("m.id_ubigeo LIKE ?")
            params.append(f"{str(ubigeo_prefix).strip()}%")
        if mesa_prefix:
            clauses.append("m.codigo_mesa LIKE ?")
            params.append(f"{str(mesa_prefix).strip()}%")
        if codigo_estado_acta:
            clauses.append("UPPER(m.codigo_estado_acta) = UPPER(?)")
            params.append(str(codigo_estado_acta).strip())
        where = " AND ".join(clauses) if clauses else "1=1"
        return where, params, needs_geo_join

    # ── 2026 1V — export ────────────────────────────────────────────────────

    def export_mesas_2026_1v(
        self,
        *,
        departamento: str | None = None,
        provincia: str | None = None,
        distrito: str | None = None,
        ubigeo_prefix: str | None = None,
        mesa_prefix: str | None = None,
        estado_acta: str | None = None,
        limit: int = 1000,
        offset: int = 0,
    ) -> dict[str, Any]:
        """Raw mesa rows from 2026 1V (`mesas_data`). Paginated."""
        limit_eff = max(1, min(int(limit), self._EXPORT_2026_MAX_LIMIT))
        offset_eff = max(0, int(offset))
        where, params, needs_geo = self._build_geo_where_2026_1v(
            departamento=departamento, provincia=provincia, distrito=distrito,
            ubigeo_prefix=ubigeo_prefix, mesa_prefix=mesa_prefix,
            estado_acta=estado_acta,
        )
        join_sql = (
            "LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6) "
            if needs_geo else ""
        )
        with self._connect() as conn:
            total_row = conn.execute(
                f"SELECT COUNT(*) AS c FROM mesas_data m {join_sql} WHERE {where}",
                params,
            ).fetchone()
            total = int(total_row["c"] or 0) if total_row else 0
            rows = conn.execute(
                f"""
                SELECT m.codigo_mesa, m.ubigeo, m.local_votacion,
                       m.electores_habiles, m.votos_emitidos, m.votos_validos,
                       m.blancos, m.nulos, m.impugnados, m.estado_acta,
                       g.departamento, g.provincia, g.distrito
                FROM mesas_data m
                LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                WHERE {where}
                ORDER BY m.codigo_mesa
                LIMIT ? OFFSET ?
                """,
                params + [limit_eff, offset_eff],
            ).fetchall()
        out_rows = [
            {
                "codigo_mesa": str(r["codigo_mesa"] or ""),
                "ubigeo": str(r["ubigeo"] or ""),
                "local_votacion": str(r["local_votacion"] or ""),
                "departamento": str(r["departamento"] or ""),
                "provincia": str(r["provincia"] or ""),
                "distrito": str(r["distrito"] or ""),
                "electores_habiles": int(r["electores_habiles"] or 0),
                "votos_emitidos": int(r["votos_emitidos"] or 0),
                "votos_validos": int(r["votos_validos"] or 0),
                "blancos": int(r["blancos"] or 0),
                "nulos": int(r["nulos"] or 0),
                "impugnados": int(r["impugnados"] or 0),
                "estado_acta": str(r["estado_acta"] or ""),
            }
            for r in rows
        ]
        schema = list(out_rows[0].keys()) if out_rows else [
            "codigo_mesa", "ubigeo", "local_votacion", "departamento", "provincia",
            "distrito", "electores_habiles", "votos_emitidos", "votos_validos",
            "blancos", "nulos", "impugnados", "estado_acta",
        ]
        return {
            "vuelta": 1,
            "total": total,
            "offset": offset_eff,
            "limit": limit_eff,
            "returned": len(out_rows),
            "has_more": (offset_eff + len(out_rows)) < total,
            "schema": schema,
            "rows": out_rows,
        }

    def export_votos_2026_1v(
        self,
        *,
        partido_ids: list[str] | None = None,
        departamento: str | None = None,
        provincia: str | None = None,
        distrito: str | None = None,
        ubigeo_prefix: str | None = None,
        mesa_prefix: str | None = None,
        estado_acta: str | None = None,
        limit: int = 5000,
        offset: int = 0,
    ) -> dict[str, Any]:
        """Raw vote rows from 2026 1V (`votos`), enriched with geo + partido."""
        limit_eff = max(1, min(int(limit), self._EXPORT_2026_MAX_LIMIT))
        offset_eff = max(0, int(offset))
        where, params, _ = self._build_geo_where_2026_1v(
            departamento=departamento, provincia=provincia, distrito=distrito,
            ubigeo_prefix=ubigeo_prefix, mesa_prefix=mesa_prefix,
            estado_acta=estado_acta,
        )
        if partido_ids:
            placeholders = ",".join("?" for _ in partido_ids)
            where = f"({where}) AND v.partido_id IN ({placeholders})"
            params = params + [str(p) for p in partido_ids]
        with self._connect() as conn:
            total_row = conn.execute(
                f"""
                SELECT COUNT(*) AS c
                FROM votos v
                JOIN mesas_data m ON m.codigo_mesa = v.codigo_mesa
                LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                WHERE {where}
                """,
                params,
            ).fetchone()
            total = int(total_row["c"] or 0) if total_row else 0
            rows = conn.execute(
                f"""
                SELECT v.codigo_mesa, v.partido_id, v.votos,
                       COALESCE(a.nombre,'') AS nombre_partido,
                       m.ubigeo, g.departamento, g.provincia, g.distrito,
                       m.votos_validos AS mesa_validos
                FROM votos v
                JOIN mesas_data m ON m.codigo_mesa = v.codigo_mesa
                LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                WHERE {where}
                ORDER BY v.codigo_mesa, v.partido_id
                LIMIT ? OFFSET ?
                """,
                params + [limit_eff, offset_eff],
            ).fetchall()
        out_rows = [
            {
                "codigo_mesa": str(r["codigo_mesa"] or ""),
                "partido_id": str(r["partido_id"] or ""),
                "nombre_partido": str(r["nombre_partido"] or ""),
                "votos": int(r["votos"] or 0),
                "ubigeo": str(r["ubigeo"] or ""),
                "departamento": str(r["departamento"] or ""),
                "provincia": str(r["provincia"] or ""),
                "distrito": str(r["distrito"] or ""),
                "mesa_votos_validos": int(r["mesa_validos"] or 0),
            }
            for r in rows
        ]
        schema = list(out_rows[0].keys()) if out_rows else [
            "codigo_mesa", "partido_id", "nombre_partido", "votos", "ubigeo",
            "departamento", "provincia", "distrito", "mesa_votos_validos",
        ]
        return {
            "vuelta": 1,
            "total": total,
            "offset": offset_eff,
            "limit": limit_eff,
            "returned": len(out_rows),
            "has_more": (offset_eff + len(out_rows)) < total,
            "schema": schema,
            "rows": out_rows,
        }

    def export_partidos_2026_1v(self) -> dict[str, Any]:
        """Catalog of partidos for 2026 1V (`agrupaciones`)."""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT partido_id, COALESCE(nombre,'') AS nombre, COALESCE(candidato,'') AS candidato "
                "FROM agrupaciones ORDER BY CAST(partido_id AS INTEGER)"
            ).fetchall()
        out_rows = [
            {
                "partido_id": str(r["partido_id"] or ""),
                "nombre_partido": str(r["nombre"] or ""),
                "candidato": str(r["candidato"] or ""),
                "is_candidate": str(r["partido_id"]) not in {"80", "81", "82"},
            }
            for r in rows
        ]
        return {
            "vuelta": 1,
            "total": len(out_rows),
            "candidatos": sum(1 for r in out_rows if r["is_candidate"]),
            "schema": ["partido_id", "nombre_partido", "candidato", "is_candidate"],
            "rows": out_rows,
        }

    def summary_2026_1v(self) -> dict[str, Any]:
        """Nacional aggregated summary for 2026 1V."""
        if self.denorm_available:
            try:
                conn = self._connect_denorm()
                rows = conn.execute("""
                    SELECT f.partido_id,
                           f.nombre_partido,
                           COALESCE(NULLIF(f.candidato, ''), a.candidato, '') AS candidato,
                           f.votos,
                           f.pct_votos_validos,
                           f.total_mesas,
                           f.mesas_contabilizadas,
                           f.total_electores_habiles,
                           f.total_votos_emitidos,
                           f.total_votos_validos
                    FROM fact_votos_nacional f
                    LEFT JOIN agrupaciones a ON a.partido_id = f.partido_id
                    WHERE f.election_year=2026 AND f.vuelta=1 AND f.es_especial=0
                    ORDER BY votos DESC
                """).fetchall()
                conn.close()
                return [dict(r) for r in rows]
            except Exception as e:
                _logger.debug("denorm fast-path failed for summary_2026_1v, falling back to OLTP: %s", e)
        with self._connect() as conn:
            agg = conn.execute(
                """
                SELECT COUNT(*) AS mesas,
                       SUM(electores_habiles) AS electores_habiles,
                       SUM(votos_emitidos) AS votos_emitidos,
                       SUM(votos_validos) AS votos_validos,
                       SUM(blancos) AS blancos,
                       SUM(nulos) AS nulos,
                       SUM(impugnados) AS impugnados
                FROM mesas_data
                """
            ).fetchone()
            party = conn.execute(
                """
                SELECT v.partido_id,
                       COALESCE(a.nombre, '') AS nombre_partido,
                       SUM(v.votos) AS total
                FROM votos v
                LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                GROUP BY v.partido_id
                ORDER BY total DESC
                """
            ).fetchall()
        electores = int(agg["electores_habiles"] or 0) if agg else 0
        emit = int(agg["votos_emitidos"] or 0) if agg else 0
        validos = int(agg["votos_validos"] or 0) if agg else 0
        return {
            "vuelta": 1,
            "mesas": int(agg["mesas"] or 0) if agg else 0,
            "electores_habiles": electores,
            "votos_emitidos": emit,
            "votos_validos": validos,
            "votos_blancos": int(agg["blancos"] or 0) if agg else 0,
            "votos_nulos": int(agg["nulos"] or 0) if agg else 0,
            "votos_impugnados": int(agg["impugnados"] or 0) if agg else 0,
            "participacion_pct": (emit / electores * 100.0) if electores else 0.0,
            "validez_pct": (validos / emit * 100.0) if emit else 0.0,
            "por_partido": [
                {
                    "partido_id": str(r["partido_id"] or ""),
                    "nombre_partido": str(r["nombre_partido"] or ""),
                    "total_votos": int(r["total"] or 0),
                    "pct_validos": (int(r["total"] or 0) / validos * 100.0) if validos else 0.0,
                }
                for r in party
            ],
        }

    # ── 2026 2V — export ────────────────────────────────────────────────────

    def export_mesas_2026_sv(
        self,
        *,
        departamento: str | None = None,
        provincia: str | None = None,
        distrito: str | None = None,
        ubigeo_prefix: str | None = None,
        mesa_prefix: str | None = None,
        codigo_estado_acta: str | None = None,
        limit: int = 1000,
        offset: int = 0,
    ) -> dict[str, Any]:
        """Raw mesa rows from 2026 SV (`mesas_sv`). Paginated."""
        limit_eff = max(1, min(int(limit), self._EXPORT_2026_MAX_LIMIT))
        offset_eff = max(0, int(offset))
        where, params, _ = self._build_geo_where_2026_sv(
            departamento=departamento, provincia=provincia, distrito=distrito,
            ubigeo_prefix=ubigeo_prefix, mesa_prefix=mesa_prefix,
            codigo_estado_acta=codigo_estado_acta,
        )
        with self._connect() as conn:
            total_row = conn.execute(
                f"""
                SELECT COUNT(*) AS c
                FROM mesas_sv m
                LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                WHERE {where}
                """,
                params,
            ).fetchone()
            total = int(total_row["c"] or 0) if total_row else 0
            rows = conn.execute(
                f"""
                SELECT m.codigo_mesa, m.id_ubigeo AS ubigeo, m.nombre_local,
                       m.electores_habiles, m.votos_emitidos, m.votos_validos,
                       m.codigo_estado_acta,
                       u.departamento, u.provincia, u.distrito, u.ciudad,
                       u.continente, u.pais
                FROM mesas_sv m
                LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                WHERE {where}
                ORDER BY m.codigo_mesa
                LIMIT ? OFFSET ?
                """,
                params + [limit_eff, offset_eff],
            ).fetchall()
        out_rows = [
            {
                "codigo_mesa": str(r["codigo_mesa"] or ""),
                "ubigeo": str(r["ubigeo"] or ""),
                "nombre_local": str(r["nombre_local"] or ""),
                "departamento": str(r["departamento"] or ""),
                "provincia": str(r["provincia"] or ""),
                "distrito": str(r["distrito"] or ""),
                "ciudad": str(r["ciudad"] or ""),
                "continente": str(r["continente"] or ""),
                "pais": str(r["pais"] or ""),
                "electores_habiles": int(r["electores_habiles"] or 0),
                "votos_emitidos": int(r["votos_emitidos"] or 0),
                "votos_validos": int(r["votos_validos"] or 0),
                "codigo_estado_acta": str(r["codigo_estado_acta"] or ""),
            }
            for r in rows
        ]
        schema = list(out_rows[0].keys()) if out_rows else [
            "codigo_mesa", "ubigeo", "nombre_local", "departamento", "provincia",
            "distrito", "ciudad", "continente", "pais", "electores_habiles",
            "votos_emitidos", "votos_validos", "codigo_estado_acta",
        ]
        return {
            "vuelta": 2,
            "total": total,
            "offset": offset_eff,
            "limit": limit_eff,
            "returned": len(out_rows),
            "has_more": (offset_eff + len(out_rows)) < total,
            "schema": schema,
            "rows": out_rows,
        }

    def export_votos_2026_sv(
        self,
        *,
        partido_ids: list[str] | None = None,
        departamento: str | None = None,
        provincia: str | None = None,
        distrito: str | None = None,
        ubigeo_prefix: str | None = None,
        mesa_prefix: str | None = None,
        codigo_estado_acta: str | None = None,
        limit: int = 5000,
        offset: int = 0,
    ) -> dict[str, Any]:
        """Raw vote rows from 2026 SV (`votos_sv`), enriched with geo + partido."""
        limit_eff = max(1, min(int(limit), self._EXPORT_2026_MAX_LIMIT))
        offset_eff = max(0, int(offset))
        where, params, _ = self._build_geo_where_2026_sv(
            departamento=departamento, provincia=provincia, distrito=distrito,
            ubigeo_prefix=ubigeo_prefix, mesa_prefix=mesa_prefix,
            codigo_estado_acta=codigo_estado_acta,
        )
        if partido_ids:
            placeholders = ",".join("?" for _ in partido_ids)
            where = f"({where}) AND v.partido_id IN ({placeholders})"
            params = params + [str(p) for p in partido_ids]
        with self._connect() as conn:
            total_row = conn.execute(
                f"""
                SELECT COUNT(*) AS c
                FROM votos_sv v
                JOIN mesas_sv m ON m.codigo_mesa = v.codigo_mesa
                LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                WHERE {where}
                """,
                params,
            ).fetchone()
            total = int(total_row["c"] or 0) if total_row else 0
            rows = conn.execute(
                f"""
                SELECT v.codigo_mesa, v.partido_id, v.votos,
                       COALESCE(a.nombre,'') AS nombre_partido,
                       m.id_ubigeo AS ubigeo,
                       u.departamento, u.provincia, u.distrito, u.ciudad,
                       u.continente, u.pais,
                       m.votos_validos AS mesa_validos
                FROM votos_sv v
                JOIN mesas_sv m ON m.codigo_mesa = v.codigo_mesa
                LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                LEFT JOIN agrupaciones_sv a ON a.partido_id = v.partido_id
                WHERE {where}
                ORDER BY v.codigo_mesa, v.partido_id
                LIMIT ? OFFSET ?
                """,
                params + [limit_eff, offset_eff],
            ).fetchall()
        out_rows = [
            {
                "codigo_mesa": str(r["codigo_mesa"] or ""),
                "partido_id": str(r["partido_id"] or ""),
                "nombre_partido": str(r["nombre_partido"] or ""),
                "votos": int(r["votos"] or 0),
                "ubigeo": str(r["ubigeo"] or ""),
                "departamento": str(r["departamento"] or ""),
                "provincia": str(r["provincia"] or ""),
                "distrito": str(r["distrito"] or ""),
                "ciudad": str(r["ciudad"] or ""),
                "continente": str(r["continente"] or ""),
                "pais": str(r["pais"] or ""),
                "mesa_votos_validos": int(r["mesa_validos"] or 0),
            }
            for r in rows
        ]
        schema = list(out_rows[0].keys()) if out_rows else [
            "codigo_mesa", "partido_id", "nombre_partido", "votos", "ubigeo",
            "departamento", "provincia", "distrito", "ciudad", "continente", "pais",
            "mesa_votos_validos",
        ]
        return {
            "vuelta": 2,
            "total": total,
            "offset": offset_eff,
            "limit": limit_eff,
            "returned": len(out_rows),
            "has_more": (offset_eff + len(out_rows)) < total,
            "schema": schema,
            "rows": out_rows,
        }

    def export_partidos_2026_sv(self) -> dict[str, Any]:
        """Catalog of partidos for 2026 SV (`agrupaciones_sv`)."""
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT partido_id, COALESCE(nombre,'') AS nombre "
                "FROM agrupaciones_sv ORDER BY CAST(partido_id AS INTEGER)"
            ).fetchall()
        out_rows = [
            {
                "partido_id": str(r["partido_id"] or ""),
                "nombre_partido": str(r["nombre"] or ""),
                "is_candidate": str(r["partido_id"]) not in {"80", "81", "82"},
            }
            for r in rows
        ]
        return {
            "vuelta": 2,
            "total": len(out_rows),
            "candidatos": sum(1 for r in out_rows if r["is_candidate"]),
            "schema": ["partido_id", "nombre_partido", "is_candidate"],
            "rows": out_rows,
        }

    def summary_2026_sv(self) -> dict[str, Any]:
        """Nacional aggregated summary for 2026 SV (cifras oficiales C only)."""
        if self.denorm_available:
            try:
                conn = self._connect_denorm()
                rows = conn.execute("""
                    SELECT partido_id, nombre_partido AS nombre_agrupacion,
                           candidato AS nombre_candidato,
                           votos, pct_votos_validos,
                           total_mesas, mesas_contabilizadas,
                           total_electores_habiles, total_votos_emitidos, total_votos_validos
                    FROM fact_votos_nacional
                    WHERE election_year=2026 AND vuelta=2 AND es_especial=0
                    ORDER BY votos DESC
                """).fetchall()
                conn.close()
                return [dict(r) for r in rows]
            except Exception as e:
                _logger.debug("denorm fast-path failed for summary_2026_sv, falling back to OLTP: %s", e)
        with self._connect() as conn:
            agg = conn.execute(
                """
                SELECT COUNT(*) AS mesas,
                       SUM(electores_habiles) AS electores_habiles,
                       SUM(votos_emitidos) AS votos_emitidos,
                       SUM(votos_validos) AS votos_validos
                FROM mesas_sv
                WHERE UPPER(codigo_estado_acta) = 'C'
                """
            ).fetchone()
            party = conn.execute(
                """
                SELECT v.partido_id,
                       COALESCE(a.nombre, '') AS nombre_partido,
                       SUM(v.votos) AS total
                FROM votos_sv v
                JOIN mesas_sv m ON m.codigo_mesa = v.codigo_mesa
                LEFT JOIN agrupaciones_sv a ON a.partido_id = v.partido_id
                WHERE UPPER(m.codigo_estado_acta) = 'C'
                GROUP BY v.partido_id
                ORDER BY total DESC
                """
            ).fetchall()
        electores = int(agg["electores_habiles"] or 0) if agg else 0
        emit = int(agg["votos_emitidos"] or 0) if agg else 0
        validos = int(agg["votos_validos"] or 0) if agg else 0
        return {
            "vuelta": 2,
            "mesas_contabilizadas": int(agg["mesas"] or 0) if agg else 0,
            "electores_habiles": electores,
            "votos_emitidos": emit,
            "votos_validos": validos,
            "participacion_pct": (emit / electores * 100.0) if electores else 0.0,
            "validez_pct": (validos / emit * 100.0) if emit else 0.0,
            "por_partido": [
                {
                    "partido_id": str(r["partido_id"] or ""),
                    "nombre_partido": str(r["nombre_partido"] or ""),
                    "total_votos": int(r["total"] or 0),
                    "pct_validos": (int(r["total"] or 0) / validos * 100.0) if validos else 0.0,
                }
                for r in party
            ],
        }

    # ── 2026 1V — geo aggregates (resultados_geo + cobertura) ───────────────

    def resultados_geo_2026_1v(
        self,
        *,
        nivel: str = "departamento",
        filtro: str | None = None,
        top_n: int = 5,
    ) -> dict[str, Any]:
        """Resultados 1V agregados por nivel geo (nacional/departamento/provincia/distrito).

        - `nivel` ∈ {"nacional","departamento","provincia","distrito"}.
        - `filtro` opcional: nombre exacto (case-insensitive) del nivel inmediato.
          Ej: nivel='departamento', filtro='LIMA' → top en Lima.
        """
        nivel_norm = (nivel or "").strip().lower()
        if nivel_norm not in {"nacional", "departamento", "provincia", "distrito"}:
            raise ValueError("nivel debe ser nacional, departamento, provincia o distrito")
        top_n_eff = max(1, min(int(top_n), 50))

        if self.denorm_available:
            result = self._resultados_geo_1v_denorm(nivel_norm, filtro, top_n_eff)
            if result is not None:
                return result

        clauses: list[str] = []
        params: list[Any] = []
        if nivel_norm != "nacional" and filtro:
            clauses.append(f"UPPER(g.{nivel_norm}) = UPPER(?)")
            params.append(str(filtro).strip())
        where = " AND ".join(clauses) if clauses else "1=1"

        with self._connect() as conn:
            coverage = conn.execute(
                f"""
                SELECT COUNT(DISTINCT m.codigo_mesa) AS mesas,
                       SUM(m.votos_emitidos) AS emit,
                       SUM(m.votos_validos) AS val
                FROM mesas_data m
                LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                WHERE {where}
                """,
                params,
            ).fetchone()
            party_rows = conn.execute(
                f"""
                SELECT v.partido_id,
                       COALESCE(a.nombre, '') AS nombre_partido,
                       SUM(v.votos) AS total
                FROM votos v
                JOIN mesas_data m ON m.codigo_mesa = v.codigo_mesa
                LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                WHERE {where} AND v.partido_id NOT IN ('80','81','82')
                GROUP BY v.partido_id
                ORDER BY total DESC
                LIMIT ?
                """,
                params + [top_n_eff],
            ).fetchall()
        mesas = int(coverage["mesas"] or 0) if coverage else 0
        emit = int(coverage["emit"] or 0) if coverage else 0
        val = int(coverage["val"] or 0) if coverage else 0
        return {
            "vuelta": 1,
            "nivel": nivel_norm,
            "filtro": filtro,
            "mesas": mesas,
            "votos_emitidos": emit,
            "votos_validos": val,
            "top": [
                {
                    "partido_id": str(r["partido_id"] or ""),
                    "nombre_partido": str(r["nombre_partido"] or ""),
                    "total_votos": int(r["total"] or 0),
                    "pct_validos": (int(r["total"] or 0) / val * 100.0) if val else 0.0,
                }
                for r in party_rows
            ],
        }

    def cobertura_2026_1v(self) -> dict[str, Any]:
        """Coverage por departamento para 2026 1V: % mesas con estado_acta válido."""
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT COALESCE(g.departamento, '?') AS departamento,
                       COUNT(*) AS total_mesas,
                       SUM(CASE WHEN UPPER(m.estado_acta) IN
                                ('CONTABILIZADA','CONTABILIZADAS','C') THEN 1 ELSE 0 END
                       ) AS contabilizadas
                FROM mesas_data m
                LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                GROUP BY departamento
                ORDER BY departamento
                """
            ).fetchall()
        out = []
        tot_mesas = 0
        tot_cont = 0
        for r in rows:
            t = int(r["total_mesas"] or 0)
            c = int(r["contabilizadas"] or 0)
            tot_mesas += t
            tot_cont += c
            out.append({
                "departamento": str(r["departamento"] or ""),
                "total_mesas": t,
                "contabilizadas": c,
                "pct_contabilizadas": (c / t * 100.0) if t else 0.0,
            })
        return {
            "vuelta": 1,
            "total_mesas": tot_mesas,
            "contabilizadas": tot_cont,
            "pct_contabilizadas": (tot_cont / tot_mesas * 100.0) if tot_mesas else 0.0,
            "por_departamento": out,
        }

    # ── Listados de catálogos ───────────────────────────────────────────────

    def list_departamentos(self) -> dict[str, Any]:
        """Catálogo de departamentos con provincias y distritos disponibles."""
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT departamento,
                       COUNT(DISTINCT provincia) AS n_provincias,
                       COUNT(DISTINCT distrito) AS n_distritos
                FROM ubigeo_reniec
                WHERE COALESCE(departamento,'') <> ''
                GROUP BY departamento
                ORDER BY departamento
                """
            ).fetchall()
        out = [
            {
                "departamento": str(r["departamento"] or ""),
                "provincias": int(r["n_provincias"] or 0),
                "distritos": int(r["n_distritos"] or 0),
            }
            for r in rows
        ]
        return {"total": len(out), "rows": out}

    def list_partidos(self, *, vuelta: int) -> dict[str, Any]:
        """Catálogo de partidos por vuelta 2026. Vuelta 1 = `agrupaciones`,
        Vuelta 2 = `agrupaciones_sv`."""
        v = 1 if int(vuelta) == 1 else 2
        if v == 1:
            return self.export_partidos_2026_1v()
        return self.export_partidos_2026_sv()

    def list_foreign_geo(self) -> dict[str, Any]:
        """Catálogo de países/ciudades con voto extranjero (foreign_catalog + ubicaciones_sv)."""
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT continente, pais, ciudad, COUNT(*) AS n_ubigeos
                FROM foreign_catalog
                GROUP BY continente, pais, ciudad
                ORDER BY continente, pais, ciudad
                """
            ).fetchall()
        out = [
            {
                "continente": str(r["continente"] or ""),
                "pais": str(r["pais"] or ""),
                "ciudad": str(r["ciudad"] or ""),
                "n_ubigeos": int(r["n_ubigeos"] or 0),
            }
            for r in rows
        ]
        # Aggregate counts by continente / pais
        by_continente: dict[str, int] = {}
        by_pais: dict[str, int] = {}
        for r in out:
            by_continente[r["continente"]] = by_continente.get(r["continente"], 0) + 1
            by_pais[r["pais"]] = by_pais.get(r["pais"], 0) + 1
        return {
            "total_ciudades": len(out),
            "total_paises": len(by_pais),
            "total_continentes": len(by_continente),
            "rows": out,
        }

    # ── Analítica genérica ───────────────────────────────────────────────────

    def top_candidato_geo(
        self,
        *,
        vuelta: int,
        partido_id: str | None = None,
        candidato_query: str | None = None,
        nivel: str = "distrito",
        top_n: int = 10,
    ) -> dict[str, Any]:
        """Top N geos (distrito/provincia/departamento) donde un candidato es más fuerte.

        - vuelta: 1 (2026 1V) o 2 (2026 SV).
        - partido_id OR candidato_query (al menos uno).
        - nivel: 'distrito' | 'provincia' | 'departamento'.
        """
        v = 1 if int(vuelta) == 1 else 2
        if not partido_id and not candidato_query:
            raise ValueError("Debe indicar partido_id o candidato_query")
        nivel_norm = (nivel or "distrito").strip().lower()
        if nivel_norm not in {"departamento", "provincia", "distrito"}:
            raise ValueError("nivel debe ser departamento, provincia o distrito")
        top_n_eff = max(1, min(int(top_n), 100))

        # Resolve partido_id from candidato query if needed
        if not partido_id and candidato_query:
            cq = _norm_text(candidato_query)
            with self._connect() as conn:
                if v == 1:
                    party_rows = conn.execute(
                        "SELECT partido_id, nombre FROM agrupaciones"
                    ).fetchall()
                else:
                    party_rows = conn.execute(
                        "SELECT partido_id, nombre FROM agrupaciones_sv"
                    ).fetchall()
            for r in party_rows:
                if cq in _norm_text(str(r["nombre"] or "")):
                    partido_id = str(r["partido_id"])
                    break
            if not partido_id:
                return {
                    "vuelta": v,
                    "partido_id": None,
                    "nivel": nivel_norm,
                    "top": [],
                    "note": f"No se encontró partido que coincida con '{candidato_query}'.",
                }

        if v == 1:
            sql = f"""
                SELECT g.{nivel_norm} AS geo,
                       SUM(v.votos) AS total,
                       COUNT(DISTINCT v.codigo_mesa) AS mesas
                FROM votos v
                JOIN mesas_data m ON m.codigo_mesa = v.codigo_mesa
                LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                WHERE v.partido_id = ? AND COALESCE(g.{nivel_norm},'') <> ''
                GROUP BY g.{nivel_norm}
                ORDER BY total DESC
                LIMIT ?
            """
        else:
            sql = f"""
                SELECT u.{nivel_norm} AS geo,
                       SUM(v.votos) AS total,
                       COUNT(DISTINCT v.codigo_mesa) AS mesas
                FROM votos_sv v
                JOIN mesas_sv m ON m.codigo_mesa = v.codigo_mesa
                LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                WHERE v.partido_id = ? AND COALESCE(u.{nivel_norm},'') <> ''
                GROUP BY u.{nivel_norm}
                ORDER BY total DESC
                LIMIT ?
            """
        with self._connect() as conn:
            rows = conn.execute(sql, (partido_id, top_n_eff)).fetchall()
        return {
            "vuelta": v,
            "partido_id": partido_id,
            "nivel": nivel_norm,
            "top": [
                {
                    "geo": str(r["geo"] or ""),
                    "votos": int(r["total"] or 0),
                    "mesas": int(r["mesas"] or 0),
                }
                for r in rows
            ],
        }

    def stats_participacion(
        self,
        *,
        vuelta: int,
        departamento: str | None = None,
    ) -> dict[str, Any]:
        """Distribución estadística de la participación (votos_emitidos/electores) por mesa.

        Útil para detectar dptos con baja/alta participación y outliers.
        """
        v = 1 if int(vuelta) == 1 else 2
        params: list[Any] = []
        if v == 1:
            sql = """
                SELECT m.votos_emitidos AS e, m.electores_habiles AS h
                FROM mesas_data m
                LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                WHERE COALESCE(m.electores_habiles,0) > 0
            """
            if departamento:
                sql += " AND UPPER(g.departamento) = UPPER(?)"
                params.append(str(departamento).strip())
        else:
            sql = """
                SELECT m.votos_emitidos AS e, m.electores_habiles AS h
                FROM mesas_sv m
                LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                WHERE COALESCE(m.electores_habiles,0) > 0
            """
            if departamento:
                sql += " AND UPPER(u.departamento) = UPPER(?)"
                params.append(str(departamento).strip())
        with self._connect() as conn:
            rows = conn.execute(sql, params).fetchall()
        if not rows:
            return {
                "vuelta": v,
                "departamento": departamento,
                "n_mesas": 0,
            }
        pcts = sorted((r["e"] or 0) / (r["h"] or 1) for r in rows)
        n = len(pcts)

        def _pct(p: float) -> float:
            i = max(0, min(n - 1, int(p * (n - 1))))
            return float(pcts[i])

        mean_val = sum(pcts) / n
        var = sum((x - mean_val) ** 2 for x in pcts) / max(1, (n - 1))
        std = var ** 0.5
        return {
            "vuelta": v,
            "departamento": departamento,
            "n_mesas": n,
            "participacion_pct": {
                "mean": mean_val * 100.0,
                "std_pp": std * 100.0,
                "min": pcts[0] * 100.0,
                "p10": _pct(0.10) * 100.0,
                "p25": _pct(0.25) * 100.0,
                "median": _pct(0.50) * 100.0,
                "p75": _pct(0.75) * 100.0,
                "p90": _pct(0.90) * 100.0,
                "max": pcts[-1] * 100.0,
            },
        }

    def audit_votos_consistency(
        self,
        *,
        vuelta: int,
        limit: int = 100,
    ) -> dict[str, Any]:
        """Audita mesas donde Σ(votos_partido) ≠ votos_validos en cabecera.

        Devuelve hasta `limit` discrepancias. n_inconsistentes = recuento total.
        Útil como check de integridad de datos.
        """
        v = 1 if int(vuelta) == 1 else 2
        limit_eff = max(1, min(int(limit), 1000))
        if v == 1:
            party_excl = "('80','81','82')"
            sql = f"""
                WITH sums AS (
                    SELECT v.codigo_mesa,
                           SUM(CASE WHEN v.partido_id NOT IN {party_excl}
                                    THEN v.votos ELSE 0 END) AS suma_partidos
                    FROM votos v
                    GROUP BY v.codigo_mesa
                )
                SELECT m.codigo_mesa, m.votos_validos AS cabecera,
                       s.suma_partidos AS suma_partidos,
                       (s.suma_partidos - m.votos_validos) AS diff
                FROM mesas_data m
                LEFT JOIN sums s ON s.codigo_mesa = m.codigo_mesa
                WHERE m.votos_validos <> COALESCE(s.suma_partidos, 0)
                ORDER BY ABS(s.suma_partidos - m.votos_validos) DESC
                LIMIT ?
            """
            total_sql = f"""
                WITH sums AS (
                    SELECT v.codigo_mesa,
                           SUM(CASE WHEN v.partido_id NOT IN {party_excl}
                                    THEN v.votos ELSE 0 END) AS suma_partidos
                    FROM votos v GROUP BY v.codigo_mesa
                )
                SELECT COUNT(*) AS c
                FROM mesas_data m
                LEFT JOIN sums s ON s.codigo_mesa = m.codigo_mesa
                WHERE m.votos_validos <> COALESCE(s.suma_partidos, 0)
            """
        else:
            sql = """
                WITH sums AS (
                    SELECT v.codigo_mesa,
                           SUM(CASE WHEN v.partido_id NOT IN ('80','81','82')
                                    THEN v.votos ELSE 0 END) AS suma_partidos
                    FROM votos_sv v
                    GROUP BY v.codigo_mesa
                )
                SELECT m.codigo_mesa, m.votos_validos AS cabecera,
                       s.suma_partidos AS suma_partidos,
                       (s.suma_partidos - m.votos_validos) AS diff
                FROM mesas_sv m
                LEFT JOIN sums s ON s.codigo_mesa = m.codigo_mesa
                WHERE m.votos_validos <> COALESCE(s.suma_partidos, 0)
                ORDER BY ABS(s.suma_partidos - m.votos_validos) DESC
                LIMIT ?
            """
            total_sql = """
                WITH sums AS (
                    SELECT v.codigo_mesa,
                           SUM(CASE WHEN v.partido_id NOT IN ('80','81','82')
                                    THEN v.votos ELSE 0 END) AS suma_partidos
                    FROM votos_sv v GROUP BY v.codigo_mesa
                )
                SELECT COUNT(*) AS c
                FROM mesas_sv m
                LEFT JOIN sums s ON s.codigo_mesa = m.codigo_mesa
                WHERE m.votos_validos <> COALESCE(s.suma_partidos, 0)
            """
        with self._connect() as conn:
            total_row = conn.execute(total_sql).fetchone()
            total = int(total_row["c"] or 0) if total_row else 0
            rows = conn.execute(sql, (limit_eff,)).fetchall()
        return {
            "vuelta": v,
            "n_inconsistentes": total,
            "limit": limit_eff,
            "rows": [
                {
                    "codigo_mesa": str(r["codigo_mesa"] or ""),
                    "votos_validos_cabecera": int(r["cabecera"] or 0),
                    "suma_partidos": int(r["suma_partidos"] or 0),
                    "diff": int(r["diff"] or 0),
                }
                for r in rows
            ],
        }

    def audit_coverage(self, *, vuelta: int) -> dict[str, Any]:
        """Matriz de cobertura: huecos por departamento (mesas sin votos hidratados)."""
        v = 1 if int(vuelta) == 1 else 2
        if v == 1:
            sql = """
                SELECT COALESCE(g.departamento, '?') AS departamento,
                       COUNT(DISTINCT m.codigo_mesa) AS total_mesas,
                       SUM(CASE WHEN EXISTS (
                            SELECT 1 FROM votos v WHERE v.codigo_mesa = m.codigo_mesa
                       ) THEN 1 ELSE 0 END) AS mesas_con_votos
                FROM mesas_data m
                LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                GROUP BY departamento
                ORDER BY departamento
            """
        else:
            sql = """
                SELECT COALESCE(u.departamento, '?') AS departamento,
                       COUNT(DISTINCT m.codigo_mesa) AS total_mesas,
                       SUM(CASE WHEN EXISTS (
                            SELECT 1 FROM votos_sv v WHERE v.codigo_mesa = m.codigo_mesa
                       ) THEN 1 ELSE 0 END) AS mesas_con_votos
                FROM mesas_sv m
                LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                GROUP BY departamento
                ORDER BY departamento
            """
        with self._connect() as conn:
            rows = conn.execute(sql).fetchall()
        out = []
        tot_mesas = 0
        tot_con = 0
        for r in rows:
            t = int(r["total_mesas"] or 0)
            c = int(r["mesas_con_votos"] or 0)
            tot_mesas += t
            tot_con += c
            out.append({
                "departamento": str(r["departamento"] or ""),
                "total_mesas": t,
                "mesas_con_votos": c,
                "huecos": max(0, t - c),
                "pct_hidratado": (c / t * 100.0) if t else 0.0,
            })
        return {
            "vuelta": v,
            "total_mesas": tot_mesas,
            "mesas_con_votos": tot_con,
            "huecos_totales": max(0, tot_mesas - tot_con),
            "pct_hidratado_global": (tot_con / tot_mesas * 100.0) if tot_mesas else 0.0,
            "por_departamento": out,
        }

    # ════════════════════════════════════════════════════════════════════════
    # GEO LOOKUP + LISTAR MESAS/LOCALES + COMPARACIONES CROSS-YEAR
    # ════════════════════════════════════════════════════════════════════════

    _YEARS_DISPONIBLES = (2021, 2026)

    def lookup_ubigeo(self, geo_name: str) -> dict[str, Any]:
        """Busca códigos ubigeo que coincidan con un nombre geográfico.

        Match exacto primero, luego fuzzy (substring). Devuelve hasta 50 matches
        ordenados por nivel: departamento → provincia → distrito.

        Args:
            geo_name: nombre del dpto/provincia/distrito (case + accent insensitive).

        Returns: {query, total, rows: [{ubigeo, nivel, departamento, provincia, distrito}]}
        """
        target = _norm_text(geo_name)
        if not target:
            return {"query": geo_name, "total": 0, "rows": []}

        # Fuente unica: ubigeo_reniec (ubigeo_onpe_api requiere sync explicito
        # via onpe_sync_domestic_catalog y suele estar vacio en arranque limpio).
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT ubigeo, distrito, provincia, departamento,
                       distrito_norm, provincia_norm, departamento_norm
                FROM ubigeo_reniec
                """
            ).fetchall()

        exact_dpto = []
        exact_prov = []
        exact_dist = []
        fuzzy_dpto = []
        fuzzy_prov = []
        fuzzy_dist = []

        seen_dpto: set[str] = set()
        seen_prov: set[str] = set()

        for r in rows:
            dn = str(r["departamento_norm"] or "")
            pn = str(r["provincia_norm"] or "")
            tn = str(r["distrito_norm"] or "")
            row_out = {
                "ubigeo": str(r["ubigeo"] or ""),
                "departamento": str(r["departamento"] or ""),
                "provincia": str(r["provincia"] or ""),
                "distrito": str(r["distrito"] or ""),
            }
            # Exact dpto match → reduce to one row per dpto
            if dn == target and dn not in seen_dpto:
                exact_dpto.append({**row_out, "nivel": "departamento"})
                seen_dpto.add(dn)
            elif pn == target and (dn, pn) not in seen_prov:
                exact_prov.append({**row_out, "nivel": "provincia"})
                seen_prov.add((dn, pn))
            elif tn == target:
                exact_dist.append({**row_out, "nivel": "distrito"})
            else:
                # Fuzzy (substring) — only collect if no exact found later
                if target in dn and dn not in seen_dpto:
                    fuzzy_dpto.append({**row_out, "nivel": "departamento"})
                    seen_dpto.add(dn)
                elif target in pn and (dn, pn) not in seen_prov:
                    fuzzy_prov.append({**row_out, "nivel": "provincia"})
                    seen_prov.add((dn, pn))
                elif target in tn:
                    fuzzy_dist.append({**row_out, "nivel": "distrito"})

        # Resolve aliases (e.g. "iquitos" → "maynas")
        if not (exact_dpto or exact_prov or exact_dist or fuzzy_dpto or fuzzy_prov or fuzzy_dist):
            alias_target = _CITY_ALIASES.get(target)
            if alias_target:
                alias_norm = _norm_text(alias_target)
                for r in rows:
                    pn = str(r["provincia_norm"] or "")
                    if pn == alias_norm:
                        exact_prov.append({
                            "ubigeo": str(r["ubigeo"] or ""),
                            "departamento": str(r["departamento"] or ""),
                            "provincia": str(r["provincia"] or ""),
                            "distrito": str(r["distrito"] or ""),
                            "nivel": "provincia",
                            "via_alias": alias_target,
                        })
                        break  # one per province match

        out = (exact_dpto + exact_prov + exact_dist +
               fuzzy_dpto + fuzzy_prov + fuzzy_dist)
        return {
            "query": geo_name,
            "query_norm": target,
            "total": len(out[:50]),
            "rows": out[:50],
        }

    def _geo_filter_ubigeo_list(
        self,
        *,
        departamento: str | None,
        provincia: str | None,
        distrito: str | None,
    ) -> set[str] | None:
        """Resolve geo filters to a set of ubigeos (None = no filter)."""
        if not (departamento or provincia or distrito):
            return None
        target = distrito or provincia or departamento
        nivel = "distrito" if distrito else ("provincia" if provincia else "departamento")
        with self._connect() as conn:
            rows = conn.execute(
                f"""
                SELECT ubigeo FROM ubigeo_reniec
                WHERE UPPER({nivel}) = UPPER(?)
                """,
                (str(target).strip(),),
            ).fetchall()
        return {str(r["ubigeo"]) for r in rows} if rows else set()

    def listar_mesas_por_geo(
        self,
        *,
        año: int,
        vuelta: int = 1,
        departamento: str | None = None,
        provincia: str | None = None,
        distrito: str | None = None,
        limit: int = 200,
        offset: int = 0,
    ) -> dict[str, Any]:
        """Lista mesas en una geo (distrito/prov/dpto) para un año/vuelta dado.

        Devuelve cabecera ligera (sin votos). Soporta paginación.
        Para años no disponibles, devuelve {available: false}.
        """
        año = int(año)
        if año not in self._YEARS_DISPONIBLES:
            return {
                "available": False,
                "año": año,
                "vuelta": int(vuelta),
                "reason": f"Año {año} no disponible en cache. "
                          f"Años hidratados: {self._YEARS_DISPONIBLES}",
            }
        v = 1 if int(vuelta) == 1 else 2
        limit_eff = max(1, min(int(limit), 5000))
        offset_eff = max(0, int(offset))

        if año == 2021:
            clauses: list[str] = ["m.vuelta = ?"]
            params: list[Any] = [v]
            if departamento:
                clauses.append("UPPER(m.departamento) = UPPER(?)")
                params.append(str(departamento).strip())
            if provincia:
                clauses.append("UPPER(m.provincia) = UPPER(?)")
                params.append(str(provincia).strip())
            if distrito:
                clauses.append("UPPER(m.distrito) = UPPER(?)")
                params.append(str(distrito).strip())
            where = " AND ".join(clauses)
            with self._connect() as conn:
                total = int((conn.execute(
                    f"SELECT COUNT(*) AS c FROM mesas_2021 m WHERE {where}",
                    params,
                ).fetchone() or {"c": 0})["c"])
                rows = conn.execute(
                    f"""
                    SELECT m.codigo_mesa, m.ubigeo,
                           m.departamento, m.provincia, m.distrito,
                           m.descrip_estado_acta AS estado_acta,
                           m.n_elec_habil AS electores_habiles,
                           m.votos_emitidos, m.votos_validos
                    FROM mesas_2021 m
                    WHERE {where}
                    ORDER BY m.codigo_mesa
                    LIMIT ? OFFSET ?
                    """,
                    params + [limit_eff, offset_eff],
                ).fetchall()
            out_rows = [
                {
                    "codigo_mesa": str(r["codigo_mesa"] or ""),
                    "ubigeo": str(r["ubigeo"] or ""),
                    "departamento": str(r["departamento"] or ""),
                    "provincia": str(r["provincia"] or ""),
                    "distrito": str(r["distrito"] or ""),
                    "estado_acta": str(r["estado_acta"] or ""),
                    "electores_habiles": int(r["electores_habiles"] or 0),
                    "votos_emitidos": int(r["votos_emitidos"] or 0),
                    "votos_validos": int(r["votos_validos"] or 0),
                }
                for r in rows
            ]
        else:  # 2026
            ubigeos = self._geo_filter_ubigeo_list(
                departamento=departamento, provincia=provincia, distrito=distrito,
            )
            if ubigeos is not None and not ubigeos:
                return {
                    "available": True,
                    "año": 2026,
                    "vuelta": v,
                    "total": 0,
                    "returned": 0,
                    "has_more": False,
                    "rows": [],
                    "warning": "No se encontraron ubigeos que coincidan con esos filtros geo.",
                }
            params: list[Any] = []
            clauses: list[str] = []
            if ubigeos is not None:
                placeholders = ",".join("?" for _ in ubigeos)
                if v == 1:
                    # mesas_data.ubigeo viene en 5 o 6 dígitos según dpto;
                    # ubigeos del filtro vienen de reniec (siempre 6). Normalizar.
                    clauses.append(
                        f"SUBSTR('00000' || m.ubigeo, -6) IN ({placeholders})"
                    )
                else:
                    clauses.append(f"m.id_ubigeo IN ({placeholders})")
                params.extend(ubigeos)
            where = " AND ".join(clauses) if clauses else "1=1"
            if v == 1:
                sql_count = f"SELECT COUNT(*) AS c FROM mesas_data m WHERE {where}"
                sql_select = f"""
                    SELECT m.codigo_mesa, m.ubigeo, m.local_votacion,
                           g.departamento, g.provincia, g.distrito,
                           m.estado_acta,
                           m.electores_habiles, m.votos_emitidos, m.votos_validos
                    FROM mesas_data m
                    LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                    WHERE {where}
                    ORDER BY m.codigo_mesa
                    LIMIT ? OFFSET ?
                """
            else:
                sql_count = f"SELECT COUNT(*) AS c FROM mesas_sv m WHERE {where}"
                sql_select = f"""
                    SELECT m.codigo_mesa, m.id_ubigeo AS ubigeo, m.nombre_local AS local_votacion,
                           u.departamento, u.provincia, u.distrito,
                           m.codigo_estado_acta AS estado_acta,
                           m.electores_habiles, m.votos_emitidos, m.votos_validos
                    FROM mesas_sv m
                    LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                    WHERE {where}
                    ORDER BY m.codigo_mesa
                    LIMIT ? OFFSET ?
                """
            with self._connect() as conn:
                total = int((conn.execute(sql_count, params).fetchone() or {"c": 0})["c"])
                rows = conn.execute(sql_select, params + [limit_eff, offset_eff]).fetchall()
            out_rows = [
                {
                    "codigo_mesa": str(r["codigo_mesa"] or ""),
                    "ubigeo": str(r["ubigeo"] or ""),
                    "local_votacion": str(r["local_votacion"] or ""),
                    "departamento": str(r["departamento"] or ""),
                    "provincia": str(r["provincia"] or ""),
                    "distrito": str(r["distrito"] or ""),
                    "estado_acta": str(r["estado_acta"] or ""),
                    "electores_habiles": int(r["electores_habiles"] or 0),
                    "votos_emitidos": int(r["votos_emitidos"] or 0),
                    "votos_validos": int(r["votos_validos"] or 0),
                }
                for r in rows
            ]

        return {
            "available": True,
            "año": año,
            "vuelta": v,
            "total": total,
            "offset": offset_eff,
            "limit": limit_eff,
            "returned": len(out_rows),
            "has_more": (offset_eff + len(out_rows)) < total,
            "rows": out_rows,
        }

    def listar_locales_por_geo(
        self,
        *,
        año: int,
        vuelta: int = 1,
        departamento: str | None = None,
        provincia: str | None = None,
        distrito: str | None = None,
        limit: int = 200,
    ) -> dict[str, Any]:
        """Lista locales de votación únicos en una geo, con #mesas y #electores por local.

        Para 2021 no hay campo de local explícito en el CSV — devolvemos por ubigeo+distrito.
        """
        año = int(año)
        if año not in self._YEARS_DISPONIBLES:
            return {
                "available": False, "año": año, "vuelta": int(vuelta),
                "reason": f"Año {año} no disponible.",
            }
        v = 1 if int(vuelta) == 1 else 2
        limit_eff = max(1, min(int(limit), 1000))

        if año == 2021:
            # 2021 no tiene nombre_local — agrupar por ubigeo + distrito
            return {
                "available": True,
                "año": año,
                "vuelta": v,
                "note": "El dataset 2021 (CSV oficial PCM) no incluye nombre de local de "
                        "votación. Se devuelven ubigeos como pseudo-locales.",
                "rows": [],
                "total": 0,
            }
        # 2026
        ubigeos = self._geo_filter_ubigeo_list(
            departamento=departamento, provincia=provincia, distrito=distrito,
        )
        params: list[Any] = []
        clauses: list[str] = []
        if ubigeos is not None:
            if not ubigeos:
                return {
                    "available": True, "año": año, "vuelta": v,
                    "total": 0, "rows": [],
                    "warning": "No se encontraron ubigeos que coincidan con esos filtros.",
                }
            placeholders = ",".join("?" for _ in ubigeos)
            if v == 1:
                clauses.append(
                    f"SUBSTR('00000' || m.ubigeo, -6) IN ({placeholders})"
                )
            else:
                clauses.append(f"m.id_ubigeo IN ({placeholders})")
            params.extend(ubigeos)
        where = " AND ".join(clauses) if clauses else "1=1"

        if v == 1:
            sql = f"""
                SELECT COALESCE(NULLIF(m.local_votacion,''), '(sin nombre)') AS local,
                       m.ubigeo,
                       g.departamento, g.provincia, g.distrito,
                       COUNT(*) AS n_mesas,
                       SUM(m.electores_habiles) AS electores,
                       SUM(m.votos_emitidos) AS emitidos
                FROM mesas_data m
                LEFT JOIN ubigeo_reniec g ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                WHERE {where}
                GROUP BY local, m.ubigeo
                ORDER BY g.departamento, g.provincia, g.distrito, local
                LIMIT ?
            """
        else:
            sql = f"""
                SELECT COALESCE(NULLIF(m.nombre_local,''), '(sin nombre)') AS local,
                       m.id_ubigeo AS ubigeo,
                       u.departamento, u.provincia, u.distrito,
                       COUNT(*) AS n_mesas,
                       SUM(m.electores_habiles) AS electores,
                       SUM(m.votos_emitidos) AS emitidos
                FROM mesas_sv m
                LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                WHERE {where}
                GROUP BY local, m.id_ubigeo
                ORDER BY u.departamento, u.provincia, u.distrito, local
                LIMIT ?
            """
        with self._connect() as conn:
            rows = conn.execute(sql, params + [limit_eff]).fetchall()
        out = [
            {
                "local_votacion": str(r["local"] or ""),
                "ubigeo": str(r["ubigeo"] or ""),
                "departamento": str(r["departamento"] or ""),
                "provincia": str(r["provincia"] or ""),
                "distrito": str(r["distrito"] or ""),
                "n_mesas": int(r["n_mesas"] or 0),
                "electores_habiles": int(r["electores"] or 0),
                "votos_emitidos": int(r["emitidos"] or 0),
            }
            for r in rows
        ]
        return {
            "available": True,
            "año": año,
            "vuelta": v,
            "total": len(out),
            "limit": limit_eff,
            "rows": out,
        }

    def mesa_geo_lookup(
        self,
        codigo_mesa: str,
        *,
        año: int = 2026,
        vuelta: int | None = None,
    ) -> dict[str, Any]:
        """Lookup ligero: dada una mesa, retorna SOLO geo + estado (sin votos)."""
        año = int(año)
        if año not in self._YEARS_DISPONIBLES:
            return {
                "found": False, "año": año,
                "reason": f"Año {año} no disponible.",
            }
        code = str(codigo_mesa).strip().zfill(6)
        if año == 2021:
            v = int(vuelta) if vuelta in (1, 2) else 1
            with self._connect() as conn:
                row = conn.execute(
                    "SELECT codigo_mesa, vuelta, ubigeo, departamento, provincia, distrito, "
                    "descrip_estado_acta AS estado_acta, n_elec_habil AS electores "
                    "FROM mesas_2021 WHERE codigo_mesa = ? AND vuelta = ?",
                    (code, v),
                ).fetchone()
            if not row:
                return {"found": False, "codigo_mesa": code, "año": 2021, "vuelta": v}
            return {
                "found": True,
                "año": 2021,
                "vuelta": int(row["vuelta"] or v),
                "codigo_mesa": str(row["codigo_mesa"] or ""),
                "ubigeo": str(row["ubigeo"] or ""),
                "departamento": str(row["departamento"] or ""),
                "provincia": str(row["provincia"] or ""),
                "distrito": str(row["distrito"] or ""),
                "local_votacion": None,
                "estado_acta": str(row["estado_acta"] or ""),
                "electores_habiles": int(row["electores"] or 0),
            }
        # 2026
        v = int(vuelta) if vuelta in (1, 2) else 1
        if v == 1:
            sql = """
                SELECT m.codigo_mesa, m.ubigeo, m.local_votacion, m.estado_acta,
                       m.electores_habiles,
                       g.departamento, g.provincia, g.distrito
                FROM mesas_data m
                LEFT JOIN ubigeo_reniec g
                  ON g.ubigeo = SUBSTR('00000' || m.ubigeo, -6)
                WHERE m.codigo_mesa = ?
            """
        else:
            sql = """
                SELECT m.codigo_mesa, m.id_ubigeo AS ubigeo,
                       m.nombre_local AS local_votacion,
                       m.codigo_estado_acta AS estado_acta,
                       m.electores_habiles,
                       u.departamento, u.provincia, u.distrito
                FROM mesas_sv m
                LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                WHERE m.codigo_mesa = ?
            """
        with self._connect() as conn:
            row = conn.execute(sql, (code,)).fetchone()
        if not row:
            return {"found": False, "codigo_mesa": code, "año": 2026, "vuelta": v}
        return {
            "found": True,
            "año": 2026,
            "vuelta": v,
            "codigo_mesa": str(row["codigo_mesa"] or ""),
            "ubigeo": str(row["ubigeo"] or ""),
            "departamento": str(row["departamento"] or ""),
            "provincia": str(row["provincia"] or ""),
            "distrito": str(row["distrito"] or ""),
            "local_votacion": str(row["local_votacion"] or ""),
            "estado_acta": str(row["estado_acta"] or ""),
            "electores_habiles": int(row["electores_habiles"] or 0),
        }

    def comparacion_mesa_2021(self, codigo_mesa: str) -> dict[str, Any]:
        """Compara una mesa entre 1ra y 2da vuelta de 2021."""
        code = str(codigo_mesa).strip().zfill(6)
        v1 = self.get_mesa_2021_from_local(code, vuelta=1)
        v2 = self.get_mesa_2021_from_local(code, vuelta=2)
        return {
            "codigo_mesa": code,
            "año": 2021,
            "primera_vuelta": v1,
            "segunda_vuelta": v2,
            "available_1v": v1 is not None,
            "available_2v": v2 is not None,
        }

    def comparacion_mesa_cross_year(
        self,
        codigo_mesa: str,
        *,
        año_a: int,
        año_b: int,
        vuelta_a: int = 1,
        vuelta_b: int = 1,
    ) -> dict[str, Any]:
        """Compara una mesa entre dos años distintos.

        Devuelve cabecera y top-5 candidatos en cada año. Marca available=false
        para años no hidratados (todo lo que no sea 2021/2026 hoy).
        """
        code = str(codigo_mesa).strip().zfill(6)

        def _fetch(año: int, vuelta: int) -> dict[str, Any]:
            if año not in self._YEARS_DISPONIBLES:
                return {
                    "año": año, "vuelta": vuelta, "available": False,
                    "reason": f"Año {año} no hidratado en cache.",
                }
            if año == 2021:
                bundle = self.get_mesa_2021_from_local(code, vuelta=vuelta)
                if bundle is None:
                    return {"año": año, "vuelta": vuelta, "available": True, "found": False}
                top = sorted(
                    bundle.get("votos", []), key=lambda x: -int(x.get("votos") or 0)
                )[:5]
                return {
                    "año": año, "vuelta": vuelta, "available": True, "found": True,
                    "cabecera": {
                        "departamento": bundle.get("departamento", ""),
                        "provincia": bundle.get("provincia", ""),
                        "distrito": bundle.get("distrito", ""),
                        "electores_habiles": bundle.get("electores_habiles", 0),
                        "votos_emitidos": bundle.get("votos_emitidos", 0),
                        "votos_validos": bundle.get("votos_validos", 0),
                    },
                    "top": [
                        {
                            "partido_id": v.get("partido_id"),
                            "candidato": v.get("candidato") or v.get("partido", ""),
                            "votos": int(v.get("votos") or 0),
                        }
                        for v in top
                    ],
                }
            # 2026
            if vuelta == 1:
                bundle = self.get_mesa_from_local(code)
                if bundle is None:
                    return {"año": año, "vuelta": vuelta, "available": True, "found": False}
                top = sorted(
                    [v for v in bundle.get("votos", [])
                     if str(v.get("partido_id")) not in {"80", "81", "82"}],
                    key=lambda x: -int(x.get("votos") or 0),
                )[:5]
                m = bundle.get("mesa_data", {})
                return {
                    "año": año, "vuelta": vuelta, "available": True, "found": True,
                    "cabecera": {
                        "departamento": m.get("departamento", ""),
                        "provincia": "",
                        "distrito": "",
                        "electores_habiles": m.get("electores_habiles", 0),
                        "votos_emitidos": m.get("votos_emitidos", 0),
                        "votos_validos": m.get("votos_validos", 0),
                    },
                    "top": [
                        {
                            "partido_id": v.get("partido_id"),
                            "nombre_partido": v.get("nombre_partido", ""),
                            "votos": int(v.get("votos") or 0),
                        }
                        for v in top
                    ],
                }
            # 2026 2V
            bundle = self.get_mesa_sv_from_local(code)
            if bundle is None:
                return {"año": año, "vuelta": vuelta, "available": True, "found": False}
            top = sorted(
                [v for v in bundle.get("votos", [])
                 if str(v.get("partido_id")) not in {"80", "81", "82"}],
                key=lambda x: -int(x.get("votos") or 0),
            )[:5]
            m = bundle.get("mesa_data", {})
            return {
                "año": año, "vuelta": vuelta, "available": True, "found": True,
                "cabecera": {
                    "departamento": m.get("departamento", ""),
                    "provincia": m.get("provincia", ""),
                    "distrito": m.get("distrito", ""),
                    "electores_habiles": m.get("electores_habiles", 0),
                    "votos_emitidos": m.get("votos_emitidos", 0),
                    "votos_validos": m.get("votos_validos", 0),
                },
                "top": [
                    {
                        "partido_id": v.get("partido_id"),
                        "nombre_partido": v.get("nombre_partido", ""),
                        "votos": int(v.get("votos") or 0),
                    }
                    for v in top
                ],
            }

        return {
            "codigo_mesa": code,
            "lado_a": _fetch(int(año_a), int(vuelta_a)),
            "lado_b": _fetch(int(año_b), int(vuelta_b)),
        }

    def comparacion_geo_cross_year(
        self,
        *,
        nivel: str,
        geo_name: str,
        año_a: int,
        año_b: int,
        vuelta_a: int = 1,
        vuelta_b: int = 1,
        top_n: int = 5,
    ) -> dict[str, Any]:
        """Compara top-N candidatos en una geo entre dos años distintos.

        nivel: 'departamento' | 'provincia' | 'distrito'.
        """
        nivel_n = (nivel or "departamento").strip().lower()
        if nivel_n not in {"departamento", "provincia", "distrito"}:
            raise ValueError("nivel debe ser departamento, provincia o distrito")
        top_n_eff = max(1, min(int(top_n), 20))

        def _fetch(año: int, vuelta: int) -> dict[str, Any]:
            if año not in self._YEARS_DISPONIBLES:
                return {
                    "año": año, "vuelta": vuelta, "available": False,
                    "reason": f"Año {año} no hidratado en cache.",
                }
            if año == 2021:
                with self._connect() as conn:
                    rows = conn.execute(
                        f"""
                        SELECT COALESCE(p.candidato, p.nombre_partido) AS nombre,
                               p.partido_id, SUM(v.votos) AS total,
                               COUNT(DISTINCT m.codigo_mesa) AS mesas,
                               SUM(m.votos_validos) AS validos
                        FROM mesas_2021 m
                        JOIN votos_2021 v
                          ON v.vuelta = m.vuelta AND v.codigo_mesa = m.codigo_mesa
                        LEFT JOIN partidos_2021 p
                          ON p.vuelta = v.vuelta AND p.partido_id = v.partido_id
                        WHERE m.vuelta = ? AND UPPER(m.{nivel_n}) = UPPER(?)
                        GROUP BY p.partido_id
                        ORDER BY total DESC
                        LIMIT ?
                        """,
                        (vuelta, geo_name, top_n_eff),
                    ).fetchall()
                total_validos = sum(int(r["total"] or 0) for r in rows)
                return {
                    "año": año, "vuelta": vuelta, "available": True,
                    "mesas": int(rows[0]["mesas"] or 0) if rows else 0,
                    "total_validos": total_validos,
                    "top": [
                        {
                            "partido_id": str(r["partido_id"] or ""),
                            "candidato": str(r["nombre"] or ""),
                            "votos": int(r["total"] or 0),
                            "pct": (int(r["total"] or 0) / total_validos * 100.0)
                                   if total_validos else 0.0,
                        }
                        for r in rows
                    ],
                }
            # 2026
            if vuelta == 1:
                ubigeos = self._geo_filter_ubigeo_list(
                    departamento=geo_name if nivel_n == "departamento" else None,
                    provincia=geo_name if nivel_n == "provincia" else None,
                    distrito=geo_name if nivel_n == "distrito" else None,
                )
                if not ubigeos:
                    return {
                        "año": año, "vuelta": vuelta, "available": True,
                        "warning": f"No se encontró '{geo_name}' como {nivel_n} en 2026.",
                        "top": [],
                    }
                placeholders = ",".join("?" for _ in ubigeos)
                sql = f"""
                    SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre,
                           SUM(v.votos) AS total
                    FROM votos v
                    JOIN mesas_data m ON m.codigo_mesa = v.codigo_mesa
                    LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                    WHERE SUBSTR('00000' || m.ubigeo, -6) IN ({placeholders})
                      AND v.partido_id NOT IN ('80','81','82')
                    GROUP BY v.partido_id
                    ORDER BY total DESC
                    LIMIT ?
                """
                with self._connect() as conn:
                    rows = conn.execute(sql, list(ubigeos) + [top_n_eff]).fetchall()
                    cov_sql = (
                        "SELECT COUNT(*) AS c, SUM(votos_validos) AS val FROM mesas_data m "
                        f"WHERE SUBSTR('00000' || m.ubigeo, -6) IN ({placeholders})"
                    )
                    cov_row = conn.execute(cov_sql, list(ubigeos)).fetchone()
                mesas = int(cov_row["c"] or 0) if cov_row else 0
                validos = int(cov_row["val"] or 0) if cov_row else 0
                return {
                    "año": año, "vuelta": vuelta, "available": True,
                    "mesas": mesas, "total_validos": validos,
                    "top": [
                        {
                            "partido_id": str(r["partido_id"] or ""),
                            "candidato": str(r["nombre"] or ""),
                            "votos": int(r["total"] or 0),
                            "pct": (int(r["total"] or 0) / validos * 100.0)
                                   if validos else 0.0,
                        }
                        for r in rows
                    ],
                }

            # 2026 2V: usar la misma geografía SV (ubicaciones_sv) que query_sv_geo.
            geo_col = "departamento" if nivel_n == "departamento" else ("provincia" if nivel_n == "provincia" else "distrito")
            with self._connect() as conn:
                geo_rows = conn.execute(
                    f"SELECT DISTINCT ubigeo FROM ubicaciones_sv WHERE UPPER({geo_col}) = UPPER(?)",
                    (geo_name,),
                ).fetchall()
            ubigeos_sv = {str(r["ubigeo"] or "") for r in geo_rows if str(r["ubigeo"] or "").strip()}
            if not ubigeos_sv:
                return {
                    "año": año, "vuelta": vuelta, "available": True,
                    "warning": f"No se encontró '{geo_name}' como {nivel_n} en 2026 2V.",
                    "top": [],
                }

            placeholders_sv = ",".join("?" for _ in ubigeos_sv)
            with self._connect() as conn:
                rows = conn.execute(
                    f"""
                    SELECT v.partido_id,
                           COALESCE(NULLIF(n.nombre_candidato,''), a.nombre, v.partido_id) AS nombre,
                           SUM(v.votos) AS total
                    FROM votos_sv v
                    JOIN mesas_sv m ON m.codigo_mesa = v.codigo_mesa
                    LEFT JOIN agrupaciones_sv a ON a.partido_id = v.partido_id
                    LEFT JOIN sv_resumen_nacional n ON n.partido_id = v.partido_id
                    WHERE m.id_ubigeo IN ({placeholders_sv})
                      AND v.partido_id NOT IN ('80','81','82')
                    GROUP BY v.partido_id, COALESCE(NULLIF(n.nombre_candidato,''), a.nombre, v.partido_id)
                    ORDER BY total DESC
                    LIMIT ?
                    """,
                    list(ubigeos_sv) + [top_n_eff],
                ).fetchall()
                cov_row = conn.execute(
                    f"""
                    SELECT COUNT(DISTINCT m.codigo_mesa) AS c,
                           SUM(v.votos) AS val
                    FROM mesas_sv m
                    JOIN votos_sv v ON v.codigo_mesa = m.codigo_mesa
                    WHERE m.id_ubigeo IN ({placeholders_sv})
                      AND v.partido_id NOT IN ('80','81','82')
                    """,
                    list(ubigeos_sv),
                ).fetchone()

            mesas = int(cov_row["c"] or 0) if cov_row else 0
            validos = int(cov_row["val"] or 0) if cov_row else 0
            return {
                "año": año, "vuelta": vuelta, "available": True,
                "mesas": mesas, "total_validos": validos,
                "top": [
                    {
                        "partido_id": str(r["partido_id"] or ""),
                        "candidato": str(r["nombre"] or ""),
                        "votos": int(r["total"] or 0),
                        "pct": (int(r["total"] or 0) / validos * 100.0)
                               if validos else 0.0,
                    }
                    for r in rows
                ],
            }

        return {
            "nivel": nivel_n,
            "geo_name": geo_name,
            "lado_a": _fetch(int(año_a), int(vuelta_a)),
            "lado_b": _fetch(int(año_b), int(vuelta_b)),
        }

    def aggregate_votes_by_party(self, ubigeos: set[str] | None = None) -> list[dict[str, Any]]:
        with self._connect() as conn:
            self._ensure_votes_by_ubigeo_partido_backfilled(conn)

            # Excluir siempre blancos (80), nulos (81) y viciados (82):
            # no son candidatos y distorsionan el ranking de partidos/candidatos.
            _NON_CANDIDATE_IDS = ("80", "81", "82")
            sql = (
                "SELECT v.partido_id AS partido_id, "
                "COALESCE(a.nombre, '') AS nombre_partido, "
                "SUM(v.total_votos) AS total_votos "
                "FROM votos_by_ubigeo_partido v "
                "LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id "
                f"WHERE v.partido_id NOT IN ({','.join('?' for _ in _NON_CANDIDATE_IDS)}) "
            )
            params: list[Any] = list(_NON_CANDIDATE_IDS)
            if ubigeos:
                placeholders = ",".join("?" for _ in ubigeos)
                sql += f"AND v.ubigeo IN ({placeholders}) "
                params.extend(sorted(ubigeos))
            sql += "GROUP BY v.partido_id, a.nombre ORDER BY total_votos DESC"

            rows = conn.execute(sql, params).fetchall()

        result: list[dict[str, Any]] = []
        for row in rows:
            result.append(
                {
                    "partido_id": str(row["partido_id"]),
                    "nombre_partido": str(row["nombre_partido"] or ""),
                    "total_votos": int(row["total_votos"] or 0),
                }
            )
        return result

    def count_mesas_by_ubigeos(self, ubigeos: set[str]) -> int:
        if not ubigeos:
            return 0

        placeholders = ",".join("?" for _ in ubigeos)
        sql = f"SELECT COUNT(*) AS c FROM mesas_data WHERE ubigeo IN ({placeholders})"
        with self._connect() as conn:
            row = conn.execute(sql, sorted(ubigeos)).fetchone()
        return int(row["c"] if row else 0)

    def find_ubigeos_by_prefix(self, prefix: str) -> list[str]:
        """Retorna ubigeos únicos de mesas_data que comienzan con el prefijo dado."""
        if not prefix:
            return []
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT DISTINCT ubigeo FROM mesas_data WHERE ubigeo LIKE ?",
                (prefix + "%",),
            ).fetchall()
        return [str(row["ubigeo"]) for row in rows if row["ubigeo"]]

    def find_domestic_ubigeos_by_geo_name(self, query: str) -> tuple[str, list[str]] | None:
        """Busca ubigeos en ubigeo_reniec + ubigeo_onpe_api (departamento > provincia > distrito).
        Retorna (nombre_geo, lista_ubigeos) o None si no se encuentran coincidencias.
        Primero resuelve alias de ciudades populares (pucallpa → coronel portillo, etc.),
        luego prueba el query completo y por tokens."""
        q_norm = _norm_text(query)
        if not q_norm:
            return None

        # 0. Alias de ciudad popular → nombre administrativo oficial
        alias_term = _CITY_ALIASES.get(q_norm)
        if alias_term is None:
            # Intentar encontrar alias dentro de tokens del query
            for word in sorted(q_norm.split(), key=len, reverse=True):
                if word in _CITY_ALIASES:
                    alias_term = _CITY_ALIASES[word]
                    break

        with self._connect() as conn:
            # Verificar si las tablas tienen datos
            count_reniec = int(
                (conn.execute("SELECT COUNT(*) AS c FROM ubigeo_reniec").fetchone() or {"c": 0})["c"]
            )
            count_onpe = int(
                (conn.execute("SELECT COUNT(*) AS c FROM ubigeo_onpe_api").fetchone() or {"c": 0})["c"]
            )
            if count_reniec == 0 and count_onpe == 0:
                return None

            def _search_table(table: str, term: str) -> tuple[str, list[str]] | None:
                like_sub = f"%{term}%" if len(term) >= 6 else f"{term}%"
                rows = conn.execute(
                    f"SELECT ubigeo, departamento FROM {table} WHERE departamento_norm LIKE ? LIMIT 2000",
                    (like_sub,),
                ).fetchall()
                if rows:
                    return str(rows[0]["departamento"]).lower(), [str(r["ubigeo"]).lstrip("0") for r in rows]
                rows = conn.execute(
                    f"SELECT ubigeo, provincia FROM {table} WHERE provincia_norm LIKE ? LIMIT 2000",
                    (like_sub,),
                ).fetchall()
                if rows:
                    return str(rows[0]["provincia"]).lower(), [str(r["ubigeo"]).lstrip("0") for r in rows]
                rows = conn.execute(
                    f"SELECT ubigeo, distrito FROM {table} WHERE distrito_norm LIKE ? LIMIT 500",
                    (like_sub,),
                ).fetchall()
                if rows:
                    return str(rows[0]["distrito"]).lower(), [str(r["ubigeo"]).lstrip("0") for r in rows]
                return None

            def _search(term: str) -> tuple[str, list[str]] | None:
                if count_reniec > 0:
                    result = _search_table("ubigeo_reniec", term)
                    if result:
                        return result
                if count_onpe > 0:
                    result = _search_table("ubigeo_onpe_api", term)
                    if result:
                        return result
                return None

            # 1. Si hay un alias de ciudad conocida, usar ese término directamente
            if alias_term:
                result = _search(alias_term)
                if result:
                    return result

            # 2. Intento con el query completo normalizado
            result = _search(q_norm)
            if result:
                return result

            # 3. Fallback: probar tokens individuales (≥4 chars), más largos primero
            _STOPWORDS = {
                "dame", "de", "del", "el", "en", "es", "la", "las", "los",
                "para", "que", "quienes", "son", "hay", "quien", "una", "uno",
                "mesas", "resumen", "top", "fueron", "gano", "cuantas", "cuanto",
                "saco", "votos", "voto", "nivel", "total", "suma", "dato", "gana",
                "vote", "nacional", "ganador", "ganaron", "obtuvo",
                "cuantos", "cuanta", "tiene", "tuvo", "habia", "hubo",
                "cual", "como", "donde", "cuando", "cada", "hubo",
            }
            tokens = sorted(
                [t for t in q_norm.split() if len(t) >= 4 and t not in _STOPWORDS],
                key=len,
                reverse=True,
            )
            for token in tokens:
                result = _search(token)
                if result:
                    return result
        return None

    def upsert_domestic_ubigeos_from_api(self, rows: list[dict[str, str]]) -> int:
        """Inserta o actualiza ubigeos domésticos obtenidos de la API ONPE en ubigeo_onpe_api.
        Cada fila debe tener: ubigeo, distrito, provincia, departamento."""
        if not rows:
            return 0
        now = self.now_iso()
        inserted = 0
        with self._connect() as conn:
            for row in rows:
                ubigeo = str(row.get("ubigeo") or "").strip()
                if not ubigeo:
                    continue
                distrito = str(row.get("distrito") or "").strip()
                provincia = str(row.get("provincia") or "").strip()
                departamento = str(row.get("departamento") or "").strip()
                conn.execute(
                    """
                    INSERT INTO ubigeo_onpe_api (
                        ubigeo, distrito, provincia, departamento,
                        distrito_norm, provincia_norm, departamento_norm, fetched_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(ubigeo) DO UPDATE SET
                        distrito=excluded.distrito,
                        provincia=excluded.provincia,
                        departamento=excluded.departamento,
                        distrito_norm=excluded.distrito_norm,
                        provincia_norm=excluded.provincia_norm,
                        departamento_norm=excluded.departamento_norm,
                        fetched_at=excluded.fetched_at
                    """,
                    (
                        ubigeo, distrito, provincia, departamento,
                        _norm_text(distrito), _norm_text(provincia), _norm_text(departamento),
                        now,
                    ),
                )
                inserted += 1
        return inserted

    def candidate_first_places_by_mesa_prefix(
        self,
        *,
        mesa_prefix: str,
        partido_ids: set[str],
        top_n: int = 10,
    ) -> dict[str, Any]:
        prefix = str(mesa_prefix or "").strip()
        if not prefix:
            raise ValueError("mesa_prefix no puede estar vacío")
        like_prefix = self._mesa_prefix_like(prefix)
        if not partido_ids:
            return {
                "mesa_prefix": prefix,
                "total_mesas_prefijo": 0,
                "mesas_con_votos": 0,
                "mesas_primero": 0,
                "lugares": [],
            }

        top_n = max(1, min(int(top_n), 50))

        with self._connect() as conn:
            # Use pre-computed totals if available (O(1))
            pt_row = conn.execute(
                "SELECT n_mesas, mesas_con_votos FROM mesa_prefix_totals WHERE prefix=?",
                (like_prefix,),
            ).fetchone() if 1 <= len(like_prefix) <= 4 else None

            if pt_row:
                total_mesas_prefijo = int(pt_row["n_mesas"] or 0)
                mesas_con_votos = int(pt_row["mesas_con_votos"] or 0)
            else:
                total_row = conn.execute(
                    "SELECT COUNT(*) AS c FROM mesas_data WHERE codigo_mesa LIKE ?",
                    (f"{like_prefix}%",),
                ).fetchone()
                total_mesas_prefijo = int(total_row["c"] if total_row else 0)
                mesas_con_votos = total_mesas_prefijo  # approx; full count not needed here

            placeholders = ",".join("?" for _ in partido_ids)

            # Fast path: use pre-computed mesa_winner table
            winner_check = conn.execute(
                "SELECT COUNT(*) AS c FROM mesa_winner LIMIT 1"
            ).fetchone()
            use_winner_table = winner_check and int(winner_check["c"] or 0) > 0

            if use_winner_table:
                rows = conn.execute(
                    f"""
                    SELECT
                        m.ubigeo,
                        COALESCE(m.local_votacion, '') AS local_votacion,
                        COALESCE(fc.continente, '') AS continente,
                        COALESCE(fc.pais, ul.pais, '') AS pais,
                        COALESCE(fc.ciudad, ul.ciudad, '') AS ciudad,
                        COALESCE(ul.departamento, '') AS departamento,
                        COUNT(*) AS mesas_primero
                    FROM mesa_winner mw
                    INNER JOIN mesas_data m ON m.codigo_mesa = mw.codigo_mesa
                    LEFT JOIN foreign_catalog fc ON fc.ubigeo = m.ubigeo
                    LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                    WHERE m.codigo_mesa LIKE ?
                      AND mw.partido_id IN ({placeholders})
                    GROUP BY m.ubigeo, m.local_votacion, fc.continente, fc.pais, fc.ciudad, ul.pais, ul.ciudad, ul.departamento
                    ORDER BY mesas_primero DESC, pais ASC, ciudad ASC, m.ubigeo ASC
                    LIMIT ?
                    """,
                    [f"{like_prefix}%", *sorted(partido_ids), top_n],
                ).fetchall()

                mesas_primero_row = conn.execute(
                    f"""SELECT COUNT(*) AS c FROM mesa_winner mw
                        INNER JOIN mesas_data m ON m.codigo_mesa = mw.codigo_mesa
                        WHERE m.codigo_mesa LIKE ? AND mw.partido_id IN ({placeholders})""",
                    [f"{like_prefix}%", *sorted(partido_ids)],
                ).fetchone()
                mesas_primero = int(mesas_primero_row["c"] if mesas_primero_row else 0)
            else:
                # Fallback: slow CTE scan
                rows = conn.execute(
                    f"""
                    WITH mesas_pref AS (
                        SELECT m.codigo_mesa, m.ubigeo, m.local_votacion
                        FROM mesas_data m WHERE m.codigo_mesa LIKE ?
                    ),
                    mesa_top AS (
                        SELECT v.codigo_mesa, MAX(v.votos) AS max_votos
                        FROM votos v INNER JOIN mesas_pref mp ON mp.codigo_mesa = v.codigo_mesa
                        GROUP BY v.codigo_mesa
                    ),
                    candidate_hits AS (
                        SELECT DISTINCT mp.codigo_mesa, mp.ubigeo, mp.local_votacion
                        FROM mesas_pref mp
                        INNER JOIN mesa_top mt ON mt.codigo_mesa = mp.codigo_mesa
                        INNER JOIN votos v ON v.codigo_mesa = mt.codigo_mesa
                            AND v.votos = mt.max_votos AND v.partido_id IN ({placeholders})
                    )
                    SELECT
                        ch.ubigeo, COALESCE(ch.local_votacion,'') AS local_votacion,
                        COALESCE(fc.continente,'') AS continente,
                        COALESCE(fc.pais,ul.pais,'') AS pais,
                        COALESCE(fc.ciudad,ul.ciudad,'') AS ciudad,
                        COALESCE(ul.departamento,'') AS departamento,
                        COUNT(*) AS mesas_primero
                    FROM candidate_hits ch
                    LEFT JOIN foreign_catalog fc ON fc.ubigeo = ch.ubigeo
                    LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = ch.ubigeo
                    GROUP BY ch.ubigeo, ch.local_votacion, fc.continente, fc.pais, fc.ciudad, ul.pais, ul.ciudad, ul.departamento
                    ORDER BY mesas_primero DESC LIMIT ?
                    """,
                    [f"{like_prefix}%", *sorted(partido_ids), top_n],
                ).fetchall()
                mesas_primero = sum(int(r["mesas_primero"] or 0) for r in rows)

        lugares: list[dict[str, Any]] = []
        for row in rows:
            lugares.append(
                {
                    "ubigeo": str(row["ubigeo"] or ""),
                    "local_votacion": str(row["local_votacion"] or ""),
                    "continente": str(row["continente"] or ""),
                    "pais": str(row["pais"] or ""),
                    "ciudad": str(row["ciudad"] or ""),
                    "departamento": str(row["departamento"] or ""),
                    "mesas_primero": int(row["mesas_primero"] or 0),
                }
            )

        return {
            "mesa_prefix": prefix,
            "total_mesas_prefijo": total_mesas_prefijo,
            "mesas_con_votos": mesas_con_votos,
            "mesas_primero": mesas_primero,
            "lugares": lugares,
        }

    def all_first_places_by_prefix(self, mesa_prefix: str, top_n: int = 20) -> dict[str, Any]:
        """Ranking completo de candidatos por mesas donde quedaron primero en el prefijo dado."""
        prefix = str(mesa_prefix or "").strip()
        if not prefix:
            raise ValueError("mesa_prefix no puede estar vacío")
        top_n = max(1, min(int(top_n), 50))
        like_prefix = self._mesa_prefix_like(prefix)

        with self._connect() as conn:
            # Use pre-computed tables if available
            pt_row = conn.execute(
                "SELECT n_mesas, mesas_con_votos FROM mesa_prefix_totals WHERE prefix=?",
                (like_prefix,),
            ).fetchone() if 1 <= len(like_prefix) <= 4 else None

            total_mesas = int(pt_row["n_mesas"] or 0) if pt_row else int(
                (conn.execute(
                    "SELECT COUNT(*) AS c FROM mesas_data WHERE codigo_mesa LIKE ?",
                    (f"{like_prefix}%",),
                ).fetchone() or {"c": 0})["c"]
            )
            mesas_con_votos = int(pt_row["mesas_con_votos"] or 0) if pt_row else total_mesas

            winner_check = conn.execute("SELECT COUNT(*) AS c FROM mesa_winner LIMIT 1").fetchone()
            if winner_check and int(winner_check["c"] or 0) > 0:
                rows = conn.execute(
                    """
                    SELECT mw.partido_id, COALESCE(a.nombre,'') AS nombre_partido,
                           COUNT(*) AS mesas_primero
                    FROM mesa_winner mw
                    INNER JOIN mesas_data m ON m.codigo_mesa = mw.codigo_mesa
                    LEFT JOIN agrupaciones a ON a.partido_id = mw.partido_id
                    WHERE m.codigo_mesa LIKE ?
                      AND mw.partido_id NOT IN ('80','81','82')
                    GROUP BY mw.partido_id, a.nombre
                    ORDER BY mesas_primero DESC LIMIT ?
                    """,
                    (f"{like_prefix}%", top_n),
                ).fetchall()
            else:
                rows = conn.execute(
                    """
                    WITH mesas_pref AS (SELECT codigo_mesa FROM mesas_data WHERE codigo_mesa LIKE ?),
                    mesa_max AS (
                        SELECT v.codigo_mesa, MAX(v.votos) AS max_votos
                        FROM votos v INNER JOIN mesas_pref mp ON mp.codigo_mesa = v.codigo_mesa
                        WHERE v.votos > 0 GROUP BY v.codigo_mesa
                    ),
                    winners AS (
                        SELECT v.codigo_mesa, v.partido_id FROM votos v
                        INNER JOIN mesa_max mm ON mm.codigo_mesa = v.codigo_mesa AND mm.max_votos = v.votos
                        WHERE v.partido_id NOT IN ('80','81','82')
                    )
                    SELECT w.partido_id, COALESCE(a.nombre,'') AS nombre_partido,
                           COUNT(DISTINCT w.codigo_mesa) AS mesas_primero
                    FROM winners w LEFT JOIN agrupaciones a ON a.partido_id = w.partido_id
                    GROUP BY w.partido_id, a.nombre ORDER BY mesas_primero DESC LIMIT ?
                    """,
                    (f"{like_prefix}%", top_n),
                ).fetchall()
        ranking = [
            {
                "partido_id": str(row["partido_id"]),
                "nombre_partido": str(row["nombre_partido"] or ""),
                "mesas_primero": int(row["mesas_primero"] or 0),
            }
            for row in rows
        ]
        return {
            "mesa_prefix": prefix,
            "total_mesas": total_mesas,
            "mesas_con_votos": mesas_con_votos,
            "ranking": ranking,
        }

    def describe_mesa_prefix(self, mesa_prefix: str, top_n_locations: int = 10) -> dict[str, Any]:
        """Describe las mesas de un prefijo: cuántas hay, dónde están, cuántos votos."""
        prefix = str(mesa_prefix or "").strip()
        if not prefix:
            raise ValueError("mesa_prefix no puede estar vacío")
        top_n_locations = max(1, min(int(top_n_locations), 50))
        with self._connect() as conn:
            total_row = conn.execute(
                "SELECT COUNT(*) AS c FROM mesas_data WHERE codigo_mesa LIKE ?",
                (f"{prefix}%",),
            ).fetchone()
            total_mesas = int(total_row["c"] if total_row else 0)
            votos_row = conn.execute(
                """
                SELECT SUM(m.votos_emitidos) AS total_votos, SUM(m.electores_habiles) AS total_electores
                FROM mesas_data m
                WHERE m.codigo_mesa LIKE ?
                """,
                (f"{prefix}%",),
            ).fetchone()
            total_votos = int((votos_row["total_votos"] or 0) if votos_row else 0)
            total_electores = int((votos_row["total_electores"] or 0) if votos_row else 0)
            loc_rows = conn.execute(
                """
                SELECT m.ubigeo, m.local_votacion,
                       COALESCE(fc.pais, '') AS pais,
                       COALESCE(fc.ciudad, '') AS ciudad,
                       COUNT(*) AS num_mesas,
                       SUM(m.votos_emitidos) AS votos_emitidos,
                       SUM(m.electores_habiles) AS electores_habiles
                FROM mesas_data m
                LEFT JOIN foreign_catalog fc ON fc.ubigeo = m.ubigeo
                WHERE m.codigo_mesa LIKE ?
                GROUP BY m.ubigeo, m.local_votacion, fc.pais, fc.ciudad
                ORDER BY num_mesas DESC
                LIMIT ?
                """,
                (f"{prefix}%", top_n_locations),
            ).fetchall()
        locations = [
            {
                "ubigeo": str(row["ubigeo"] or ""),
                "local_votacion": str(row["local_votacion"] or ""),
                "pais": str(row["pais"] or ""),
                "ciudad": str(row["ciudad"] or ""),
                "num_mesas": int(row["num_mesas"] or 0),
                "votos_emitidos": int(row["votos_emitidos"] or 0),
                "electores_habiles": int(row["electores_habiles"] or 0),
            }
            for row in loc_rows
        ]

        # Extra: breakdown by departamento (via ubigeo_location_cache)
        with self._connect() as conn:
            depto_rows = conn.execute(
                """
                SELECT
                  COALESCE(u.departamento, 'Sin departamento') AS depto,
                  COUNT(DISTINCT m.codigo_mesa) AS n_mesas,
                  SUM(m.electores_habiles) AS total_eh,
                  SUM(m.votos_emitidos) AS total_votos
                FROM mesas_data m
                LEFT JOIN ubigeo_location_cache u ON m.ubigeo = u.ubigeo
                WHERE m.codigo_mesa LIKE ?
                GROUP BY depto ORDER BY n_mesas DESC LIMIT 20
                """,
                (f"{prefix}%",),
            ).fetchall()
            eh_stats = conn.execute(
                """SELECT MIN(electores_habiles) as min_eh, MAX(electores_habiles) as max_eh,
                   ROUND(AVG(electores_habiles),1) as avg_eh
                   FROM mesas_data WHERE codigo_mesa LIKE ?""",
                (f"{prefix}%",),
            ).fetchone()

        departamentos = [
            {
                "departamento": str(r["depto"]),
                "n_mesas": int(r["n_mesas"] or 0),
                "total_electores_habiles": int(r["total_eh"] or 0),
                "total_votos_emitidos": int(r["total_votos"] or 0),
            }
            for r in depto_rows
        ]
        eh_min = int(eh_stats["min_eh"] or 0) if eh_stats else 0
        eh_max = int(eh_stats["max_eh"] or 0) if eh_stats else 0
        eh_avg = float(eh_stats["avg_eh"] or 0) if eh_stats else 0.0

        return {
            "mesa_prefix": prefix,
            "total_mesas": total_mesas,
            "total_votos_emitidos": total_votos,
            "total_electores_habiles": total_electores,
            "electores_min": eh_min,
            "electores_max": eh_max,
            "electores_avg": eh_avg,
            "locations": locations,
            "departamentos": departamentos,
        }

    def summarize_mesa_prefix(self, mesa_prefix: str, sample_size: int = 5) -> dict[str, Any]:
        """Resumen factual de un segmento de mesas: métricas, países, ciudades y muestra."""
        lp = self._mesa_prefix_like(mesa_prefix)
        pat = f"{lp}%"
        with self._connect() as conn:
            # Fast path for totals via pre-computed table
            counts = None
            if 1 <= len(lp) <= 4:
                row_pt = conn.execute(
                    "SELECT n_mesas, mesas_con_votos, votos_emitidos, votos_validos FROM mesa_prefix_totals WHERE prefix=?",
                    (lp,),
                ).fetchone()
                if row_pt:
                    counts = row_pt

            if counts is None:
                counts = conn.execute(
                    """SELECT COUNT(*) AS n_mesas,
                              SUM(CASE WHEN votos_emitidos>0 THEN 1 ELSE 0 END) AS mesas_con_votos,
                              COALESCE(SUM(votos_emitidos),0) AS votos_emitidos,
                              COALESCE(SUM(votos_validos),0) AS votos_validos
                       FROM mesas_data WHERE codigo_mesa LIKE ?""",
                    (pat,),
                ).fetchone()

            # Agrupar por ciudad/país (join con ubigeo_location_cache y foreign_catalog)
            city_rows = conn.execute(
                """SELECT COALESCE(ul.ciudad, fc.ciudad, '') AS geo_ciudad,
                          COALESCE(ul.pais, fc.pais, '') AS geo_pais,
                          COALESCE(ul.departamento, '') AS geo_departamento,
                          COUNT(*) AS num_mesas,
                          COALESCE(SUM(m.votos_emitidos), 0) AS votos_emitidos
                   FROM mesas_data m
                   LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                   LEFT JOIN foreign_catalog fc ON fc.ubigeo = m.ubigeo
                   WHERE m.codigo_mesa LIKE ?
                   GROUP BY geo_ciudad, geo_pais, geo_departamento
                   ORDER BY num_mesas DESC
                   LIMIT 20""",
                (pat,),
            ).fetchall()
            # Muestra de mesas con ubicación
            sample_rows = conn.execute(
                """SELECT m.codigo_mesa, m.ubigeo, m.local_votacion,
                          m.votos_emitidos, m.votos_validos, m.estado_acta,
                          COALESCE(ul.ciudad, fc.ciudad, '') AS geo_ciudad,
                          COALESCE(ul.pais, fc.pais, '') AS geo_pais,
                          COALESCE(ul.departamento, '') AS geo_departamento
                   FROM mesas_data m
                   LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                   LEFT JOIN foreign_catalog fc ON fc.ubigeo = m.ubigeo
                   WHERE m.codigo_mesa LIKE ?
                   ORDER BY m.codigo_mesa
                   LIMIT ?""",
                (pat, max(1, int(sample_size))),
            ).fetchall()

        # counts may come from mesa_prefix_totals (n_mesas) or fallback (total_mesas)
        def _ci(key1: str, key2: str) -> int:
            if counts is None:
                return 0
            try:
                v = counts[key1]
            except (IndexError, KeyError):
                try:
                    v = counts[key2]
                except (IndexError, KeyError):
                    v = 0
            return int(v or 0)

        total_mesas = _ci("n_mesas", "total_mesas")
        mesas_con_votos = _ci("mesas_con_votos", "mesas_con_votos")
        votos_emitidos = _ci("votos_emitidos", "votos_emitidos")
        votos_validos = _ci("votos_validos", "votos_validos")

        top_ciudades = [
            {
                "ciudad": row["geo_ciudad"],
                "pais": row["geo_pais"],
                "departamento": row["geo_departamento"],
                "num_mesas": row["num_mesas"],
                "votos_emitidos": row["votos_emitidos"],
            }
            for row in city_rows
        ]
        total_ciudades = len({(r["geo_ciudad"], r["geo_pais"]) for r in city_rows if r["geo_ciudad"] or r["geo_pais"]})
        total_paises = len({r["geo_pais"] for r in city_rows if r["geo_pais"]})

        sample = [
            {
                "codigo_mesa": row["codigo_mesa"],
                "ubigeo": row["ubigeo"],
                "local_votacion": row["local_votacion"],
                "votos_emitidos": row["votos_emitidos"],
                "votos_validos": row["votos_validos"],
                "estado_acta": row["estado_acta"],
                "ciudad": row["geo_ciudad"],
                "pais": row["geo_pais"],
                "departamento": row["geo_departamento"],
            }
            for row in sample_rows
        ]

        return {
            "mesa_prefix": mesa_prefix,
            "total_mesas": total_mesas,
            "mesas_con_votos": mesas_con_votos,
            "votos_emitidos": votos_emitidos,
            "votos_validos": votos_validos,
            "total_ciudades": total_ciudades,
            "total_paises": total_paises,
            "top_ciudades": top_ciudades,
            "sample": sample,
        }

    def get_top_candidates_for_prefix(
        self, mesa_prefix: str, top_n: int = 5, *, exclude_blancos_nulos: bool = True
    ) -> list[dict[str, Any]]:
        """Ranking de candidatos para un prefijo de mesa usando tabla pre-computada (O(1)).

        Fallback a scan de votos si la tabla no tiene datos para ese prefijo.
        """
        lp = self._mesa_prefix_like(mesa_prefix)
        top_n = max(1, min(int(top_n), 50))

        with self._connect() as conn:
            plen = len(lp)
            rows = None
            if 1 <= plen <= 4:
                rows = conn.execute(
                    """SELECT ps.partido_id, a.nombre, ps.total_votos, ps.n_mesas
                       FROM mesa_prefix_party_summary ps
                       LEFT JOIN agrupaciones a ON a.partido_id = ps.partido_id
                       WHERE ps.prefix = ?
                       ORDER BY ps.total_votos DESC""",
                    (lp,),
                ).fetchall()

            if not rows:
                # Fallback: scan votos (slow but correct)
                pat = f"{lp}%"
                rows = conn.execute(
                    """SELECT v.partido_id, a.nombre, SUM(v.votos) AS total_votos,
                              COUNT(DISTINCT v.codigo_mesa) AS n_mesas
                       FROM votos v LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                       WHERE v.codigo_mesa LIKE ?
                       GROUP BY v.partido_id ORDER BY total_votos DESC""",
                    (pat,),
                ).fetchall()

        results = []
        rank = 1
        for row in rows:
            nombre = row["nombre"] or row["partido_id"] or ""
            if exclude_blancos_nulos:
                nom_low = nombre.lower()
                if any(k in nom_low for k in ("blanco", "nulo", "impugnad")):
                    continue
            results.append({
                "rank": rank,
                "partido_id": row["partido_id"],
                "nombre": nombre,
                "votos": int(row["total_votos"] or 0),
                "n_mesas": int(row["n_mesas"] or 0),
            })
            rank += 1
            if rank > top_n:
                break
        return results

    def get_coverage_metrics(
        self,
        *,
        prefix: str | None = None,
        ubigeos: set[str] | None = None,
    ) -> dict[str, Any]:
        """Métricas de cobertura para un segmento.

        Para queries por prefijo de 1-4 dígitos usa mesa_prefix_totals (O(1)).
        Para queries por ubigeo usa mesas_data directamente.
        """
        if not prefix and not ubigeos:
            return {
                "total_mesas": 0, "mesas_con_votos": 0,
                "votos_emitidos": 0, "votos_validos": 0, "coverage_pct": 0.0,
            }
        with self._connect() as conn:
            if prefix:
                lp = self._mesa_prefix_like(prefix)
                # Fast path: use pre-aggregated table if available
                plen = len(lp)
                if 1 <= plen <= 4:
                    row = conn.execute(
                        """SELECT n_mesas, mesas_con_votos, votos_emitidos, votos_validos
                           FROM mesa_prefix_totals WHERE prefix = ?""",
                        (lp,),
                    ).fetchone()
                    if row:
                        total_mesas = int(row["n_mesas"] or 0)
                        cov = int(row["mesas_con_votos"] or 0)
                        votos_emitidos = int(row["votos_emitidos"] or 0)
                        votos_validos = int(row["votos_validos"] or 0)
                        coverage_pct = round(cov / total_mesas * 100, 1) if total_mesas > 0 else 0.0
                        return {
                            "total_mesas": total_mesas,
                            "mesas_con_votos": cov,
                            "votos_emitidos": votos_emitidos,
                            "votos_validos": votos_validos,
                            "coverage_pct": coverage_pct,
                        }
                # Fallback: LIKE scan on mesas_data (still fast via PK index)
                pat = f"{lp}%"
                row2 = conn.execute(
                    """SELECT COUNT(*) AS n, SUM(CASE WHEN votos_emitidos>0 THEN 1 ELSE 0 END) AS cov,
                              COALESCE(SUM(votos_emitidos),0) AS ve, COALESCE(SUM(votos_validos),0) AS vv
                       FROM mesas_data WHERE codigo_mesa LIKE ?""", (pat,)
                ).fetchone()
                total_mesas = int((row2["n"] if row2 else 0) or 0)
                cov = int((row2["cov"] if row2 else 0) or 0)
                votos_emitidos = int((row2["ve"] if row2 else 0) or 0)
                votos_validos = int((row2["vv"] if row2 else 0) or 0)
            else:
                ph = ",".join("?" for _ in ubigeos)  # type: ignore[arg-type]
                args = sorted(ubigeos)  # type: ignore[arg-type]
                row2 = conn.execute(
                    f"""SELECT COUNT(*) AS n,
                               SUM(CASE WHEN votos_emitidos>0 THEN 1 ELSE 0 END) AS cov,
                               COALESCE(SUM(votos_emitidos),0) AS ve,
                               COALESCE(SUM(votos_validos),0) AS vv
                        FROM mesas_data WHERE ubigeo IN ({ph})""", args
                ).fetchone()
                total_mesas = int((row2["n"] if row2 else 0) or 0)
                cov = int((row2["cov"] if row2 else 0) or 0)
                votos_emitidos = int((row2["ve"] if row2 else 0) or 0)
                votos_validos = int((row2["vv"] if row2 else 0) or 0)
        coverage_pct = round(cov / total_mesas * 100, 1) if total_mesas > 0 else 0.0
        return {
            "total_mesas": total_mesas,
            "mesas_con_votos": cov,
            "votos_emitidos": votos_emitidos,
            "votos_validos": votos_validos,
            "coverage_pct": coverage_pct,
        }

    def get_uncovered_mesas(
        self,
        *,
        prefix: str | None = None,
        ubigeos: set[str] | None = None,
        limit: int = 20,
    ) -> list[str]:
        """Retorna códigos de mesas sin datos de votos en el segmento dado."""
        limit = max(1, min(int(limit), 200))
        with self._connect() as conn:
            if prefix:
                lp = self._mesa_prefix_like(prefix)
                rows = conn.execute(
                    """SELECT m.codigo_mesa FROM mesas_data m
                       WHERE m.codigo_mesa LIKE ?
                       AND NOT EXISTS (
                           SELECT 1 FROM votos v WHERE v.codigo_mesa = m.codigo_mesa AND v.votos > 0
                       )
                       ORDER BY m.codigo_mesa LIMIT ?""",
                    (f"{lp}%", limit),
                ).fetchall()
            elif ubigeos:
                ph = ",".join("?" for _ in ubigeos)
                rows = conn.execute(
                    f"""SELECT m.codigo_mesa FROM mesas_data m
                        WHERE m.ubigeo IN ({ph})
                        AND NOT EXISTS (
                            SELECT 1 FROM votos v WHERE v.codigo_mesa = m.codigo_mesa AND v.votos > 0
                        )
                        ORDER BY m.codigo_mesa LIMIT ?""",
                    [*sorted(ubigeos), limit],
                ).fetchall()
            else:
                return []
        return [str(row["codigo_mesa"]) for row in rows]


        prefix = str(mesa_prefix or "").strip()
        if not prefix:
            raise ValueError("mesa_prefix no puede estar vacío")
        like_prefix = self._mesa_prefix_like(prefix)

        sample_size = max(1, min(int(sample_size), 20))
        like_value = f"{like_prefix}%"

        with self._connect() as conn:
            total_row = conn.execute(
                "SELECT COUNT(*) AS c FROM mesas_data WHERE codigo_mesa LIKE ?",
                (like_value,),
            ).fetchone()
            total_mesas = int(total_row["c"] if total_row else 0)

            votos_row = conn.execute(
                """
                SELECT COUNT(DISTINCT v.codigo_mesa) AS c
                FROM votos v
                INNER JOIN mesas_data m ON m.codigo_mesa = v.codigo_mesa
                WHERE m.codigo_mesa LIKE ?
                """,
                (like_value,),
            ).fetchone()
            mesas_con_votos = int(votos_row["c"] if votos_row else 0)

            foreign_row = conn.execute(
                """
                SELECT COUNT(*) AS c
                FROM mesas_data m
                INNER JOIN foreign_catalog f ON f.ubigeo = m.ubigeo
                WHERE m.codigo_mesa LIKE ?
                """,
                (like_value,),
            ).fetchone()
            mesas_con_match_foreign_catalog = int(foreign_row["c"] if foreign_row else 0)

            totals_row = conn.execute(
                """
                SELECT
                    COALESCE(SUM(COALESCE(m.votos_emitidos, 0)), 0) AS votos_emitidos,
                    COALESCE(SUM(COALESCE(m.votos_validos, 0)), 0) AS votos_validos
                FROM mesas_data m
                WHERE m.codigo_mesa LIKE ?
                """,
                (like_value,),
            ).fetchone()
            votos_emitidos = int(totals_row["votos_emitidos"] if totals_row else 0)
            votos_validos = int(totals_row["votos_validos"] if totals_row else 0)

            geo_row = conn.execute(
                """
                SELECT
                    COUNT(DISTINCT CASE WHEN COALESCE(f.pais, ul.pais, '') <> '' THEN COALESCE(f.pais, ul.pais) END) AS paises,
                    COUNT(DISTINCT CASE WHEN COALESCE(f.ciudad, ul.ciudad, '') <> '' THEN COALESCE(f.ciudad, ul.ciudad) END) AS ciudades
                FROM mesas_data m
                LEFT JOIN foreign_catalog f ON f.ubigeo = m.ubigeo
                LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                WHERE m.codigo_mesa LIKE ?
                """,
                (like_value,),
            ).fetchone()
            total_paises = int(geo_row["paises"] if geo_row else 0)
            total_ciudades = int(geo_row["ciudades"] if geo_row else 0)

            top_city_rows = conn.execute(
                """
                SELECT
                    COALESCE(f.pais, ul.pais, '') AS pais,
                    COALESCE(f.ciudad, ul.ciudad, '') AS ciudad,
                    COALESCE(ul.departamento, '') AS departamento,
                    COUNT(*) AS mesas,
                    COALESCE(SUM(COALESCE(m.votos_emitidos, 0)), 0) AS votos_emitidos,
                    COALESCE(SUM(COALESCE(m.votos_validos, 0)), 0) AS votos_validos
                FROM mesas_data m
                LEFT JOIN foreign_catalog f ON f.ubigeo = m.ubigeo
                LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                WHERE m.codigo_mesa LIKE ?
                GROUP BY COALESCE(f.pais, ul.pais, ''), COALESCE(f.ciudad, ul.ciudad, ''), COALESCE(ul.departamento, '')
                ORDER BY mesas DESC, votos_validos DESC, pais ASC, ciudad ASC
                LIMIT 5
                """,
                (like_value,),
            ).fetchall()

            sample_rows = conn.execute(
                """
                SELECT
                    m.codigo_mesa,
                    COALESCE(m.ubigeo, '') AS ubigeo,
                    COALESCE(m.local_votacion, '') AS local_votacion,
                    COALESCE(m.estado_acta, '') AS estado_acta,
                    CASE WHEN f.ubigeo IS NULL THEN 0 ELSE 1 END AS foreign_catalog_match,
                    COALESCE(f.continente, '') AS continente,
                    COALESCE(f.pais, ul.pais, '') AS pais,
                    COALESCE(f.ciudad, ul.ciudad, '') AS ciudad,
                    COALESCE(ul.departamento, '') AS departamento
                FROM mesas_data m
                LEFT JOIN foreign_catalog f ON f.ubigeo = m.ubigeo
                LEFT JOIN ubigeo_location_cache ul ON ul.ubigeo = m.ubigeo
                WHERE m.codigo_mesa LIKE ?
                ORDER BY m.codigo_mesa ASC
                LIMIT ?
                """,
                (like_value, sample_size),
            ).fetchall()

        sample: list[dict[str, Any]] = []
        for row in sample_rows:
            sample.append(
                {
                    "codigo_mesa": str(row["codigo_mesa"] or ""),
                    "ubigeo": str(row["ubigeo"] or ""),
                    "local_votacion": str(row["local_votacion"] or ""),
                    "estado_acta": str(row["estado_acta"] or ""),
                    "foreign_catalog_match": bool(int(row["foreign_catalog_match"] or 0)),
                    "continente": str(row["continente"] or ""),
                    "pais": str(row["pais"] or ""),
                    "ciudad": str(row["ciudad"] or ""),
                    "departamento": str(row["departamento"] or ""),
                }
            )

        top_ciudades: list[dict[str, Any]] = []
        for row in top_city_rows:
            top_ciudades.append(
                {
                    "pais": str(row["pais"] or ""),
                    "ciudad": str(row["ciudad"] or ""),
                    "departamento": str(row["departamento"] or ""),
                    "mesas": int(row["mesas"] or 0),
                    "votos_emitidos": int(row["votos_emitidos"] or 0),
                    "votos_validos": int(row["votos_validos"] or 0),
                }
            )

        return {
            "mesa_prefix": prefix,
            "total_mesas": total_mesas,
            "mesas_con_votos": mesas_con_votos,
            "votos_emitidos": votos_emitidos,
            "votos_validos": votos_validos,
            "total_paises": total_paises,
            "total_ciudades": total_ciudades,
            "mesas_con_match_foreign_catalog": mesas_con_match_foreign_catalog,
            "mesas_sin_match_foreign_catalog": max(0, total_mesas - mesas_con_match_foreign_catalog),
            "top_ciudades": top_ciudades,
            "sample": sample,
        }

    # ═══════════════════════════════════════════════════════════════════════════
    # ─── SEGUNDA VUELTA METHODS ─────────────────────────────────────────────
    # ═══════════════════════════════════════════════════════════════════════════

    def get_sv_sync_meta(self, key: str) -> str | None:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT value FROM sv_sync_meta WHERE key = ?", (key,)
            ).fetchone()
        return str(row["value"]) if row else None

    def set_sv_sync_meta(self, key: str, value: str) -> None:
        now = self.now_iso()
        with self._connect() as conn:
            conn.execute(
                "INSERT INTO sv_sync_meta (key, value, updated_at) VALUES (?, ?, ?) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at",
                (key, value, now),
            )

    def bootstrap_sv_mesas(self, sv_output_dir: Path) -> int:
        """Load/UPSERT mesas_sv from mesas_data.txt. Returns rows_affected."""
        mesas_file = sv_output_dir / "mesas_data.txt"
        if not mesas_file.exists():
            return 0
        now = self.now_iso()
        count = 0
        with self._connect() as conn:
            with mesas_file.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    cm = str(row.get("codigo_mesa", "")).strip()
                    if not cm:
                        continue
                    conn.execute(
                        """INSERT INTO mesas_sv (
                            codigo_mesa, id_ubigeo, nombre_local, id_ambito,
                            electores_habiles, votos_emitidos, votos_validos,
                            total_asistentes, codigo_estado_acta, fetched_at
                        ) VALUES (?,?,?,?,?,?,?,?,?,?)
                        ON CONFLICT(codigo_mesa) DO UPDATE SET
                            id_ubigeo=excluded.id_ubigeo,
                            nombre_local=excluded.nombre_local,
                            id_ambito=excluded.id_ambito,
                            electores_habiles=excluded.electores_habiles,
                            votos_emitidos=excluded.votos_emitidos,
                            votos_validos=excluded.votos_validos,
                            total_asistentes=excluded.total_asistentes,
                            codigo_estado_acta=excluded.codigo_estado_acta,
                            fetched_at=excluded.fetched_at""",
                        (
                            cm,
                            str(row.get("id_ubigeo", "")).strip(),
                            str(row.get("nombre_local_votacion", "")).strip(),
                            int(str(row.get("id_ambito_geografico", "1")).strip() or 1),
                            int(str(row.get("electores_habiles", "0")).strip() or 0),
                            int(str(row.get("votos_emitidos", "0")).strip() or 0),
                            int(str(row.get("votos_validos", "0")).strip() or 0),
                            int(str(row.get("total_asistentes", "0")).strip() or 0),
                            str(row.get("codigo_estado_acta", "")).strip(),
                            now,
                        ),
                    )
                    count += 1
        return count

    def bootstrap_sv_votos(self, sv_output_dir: Path) -> int:
        """Load/UPSERT votos_sv from votos.txt. Returns rows_affected."""
        votos_file = sv_output_dir / "votos.txt"
        if not votos_file.exists():
            return 0
        now = self.now_iso()
        count = 0
        with self._connect() as conn:
            with votos_file.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    cm = str(row.get("codigo_mesa", "")).strip()
                    pid = str(row.get("partido_id", "")).strip()
                    votos_raw = str(row.get("votos", "0")).strip()
                    if not cm or not pid or not votos_raw.lstrip("-").isdigit():
                        continue
                    conn.execute(
                        """INSERT INTO votos_sv (codigo_mesa, partido_id, votos, fetched_at)
                        VALUES (?,?,?,?)
                        ON CONFLICT(codigo_mesa, partido_id) DO UPDATE SET
                            votos=excluded.votos, fetched_at=excluded.fetched_at""",
                        (cm, pid, int(votos_raw), now),
                    )
                    count += 1
        return count

    def bootstrap_sv_agrupaciones(self, sv_output_dir: Path) -> int:
        agr_file = sv_output_dir / "agrupaciones.txt"
        if not agr_file.exists():
            return 0
        now = self.now_iso()
        count = 0
        with self._connect() as conn:
            with agr_file.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    pid = str(row.get("partido_id", "")).strip()
                    if not pid:
                        continue
                    conn.execute(
                        """INSERT INTO agrupaciones_sv (partido_id, nombre, fetched_at)
                        VALUES (?,?,?)
                        ON CONFLICT(partido_id) DO UPDATE SET nombre=excluded.nombre, fetched_at=excluded.fetched_at""",
                        (pid, str(row.get("nombre", "")).strip(), now),
                    )
                    count += 1
        return count

    def bootstrap_sv_ubicaciones(self, sv_output_dir: Path) -> int:
        ub_file = sv_output_dir / "ubicaciones.txt"
        if not ub_file.exists():
            return 0
        now = self.now_iso()
        count = 0
        with self._connect() as conn:
            with ub_file.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    ubigeo = str(row.get("ubigeo", "")).strip()
                    if not ubigeo:
                        continue
                    conn.execute(
                        """INSERT INTO ubicaciones_sv (ubigeo, ambito, departamento, provincia, distrito, continente, pais, ciudad, fetched_at)
                        VALUES (?,?,?,?,?,?,?,?,?)
                        ON CONFLICT(ubigeo) DO UPDATE SET
                            ambito=excluded.ambito, departamento=excluded.departamento,
                            provincia=excluded.provincia, distrito=excluded.distrito,
                            continente=excluded.continente, pais=excluded.pais,
                            ciudad=excluded.ciudad, fetched_at=excluded.fetched_at""",
                        (
                            ubigeo,
                            str(row.get("ambito", "")).strip(),
                            str(row.get("departamento", "")).strip(),
                            str(row.get("provincia", "")).strip(),
                            str(row.get("distrito", "")).strip(),
                            str(row.get("continente", "")).strip(),
                            str(row.get("pais", "")).strip(),
                            str(row.get("ciudad", "")).strip(),
                            now,
                        ),
                    )
                    count += 1
        return count

    def bootstrap_sv_reasignados(self, sv_output_dir: Path) -> int:
        reass_file = sv_output_dir / "locales_reasignados_segunda_vuelta_2026.txt"
        if not reass_file.exists():
            return 0
        count = 0
        with self._connect() as conn:
            with reass_file.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    nro_raw = str(row.get("nro", "")).strip()
                    if not nro_raw.isdigit():
                        continue
                    conn.execute(
                        """INSERT INTO locales_reasignados_sv (
                            nro, odpe, dpto, provincia, distrito, ccpp,
                            nombre_local_original, nombre_local_nuevo, motivo, mesas_afectadas, estado_parseo
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
                        ON CONFLICT(nro) DO UPDATE SET
                            odpe=excluded.odpe, dpto=excluded.dpto, provincia=excluded.provincia,
                            distrito=excluded.distrito, ccpp=excluded.ccpp,
                            nombre_local_original=excluded.nombre_local_original,
                            nombre_local_nuevo=excluded.nombre_local_nuevo,
                            motivo=excluded.motivo, mesas_afectadas=excluded.mesas_afectadas,
                            estado_parseo=excluded.estado_parseo""",
                        (
                            int(nro_raw),
                            str(row.get("odpe", "")).strip(),
                            str(row.get("dpto", "")).strip(),
                            str(row.get("provincia", "")).strip(),
                            str(row.get("distrito", "")).strip(),
                            str(row.get("ccpp", "")).strip(),
                            str(row.get("nombre_local_votacion", "")).strip(),
                            str(row.get("nombre_local_votacion_nuevo", "")).strip(),
                            str(row.get("motivo", "")).strip(),
                            int(str(row.get("mesas_a_reasignar", "0")).strip() or 0),
                            str(row.get("estado_parseo", "")).strip(),
                        ),
                    )
                    count += 1
        return count

    def bootstrap_resumen_sv(self, sv_resumen_dir: Path) -> dict[str, int]:
        """Load all 4 resumen files transactionally. Full replace."""

        def _safe_int(value: Any) -> int:
            raw = str(value or "").strip()
            if not raw:
                return 0
            try:
                return int(float(raw))
            except ValueError:
                return 0

        def _safe_float(value: Any) -> float:
            raw = str(value or "").strip()
            if not raw:
                return 0.0
            try:
                return float(raw)
            except ValueError:
                return 0.0

        def _infer_pid(row: dict[str, Any], fallback_index: int) -> str:
            pid = str(row.get("partido_id", "")).strip()
            if pid:
                return pid
            label = _norm_text(
                str(row.get("nombre_agrupacion_politica", "")).strip()
                or str(row.get("nombre_candidato", "")).strip()
            )
            explicit = {
                "fuerza popular": "8",
                "keiko sofia fujimori higuchi": "8",
                "juntos por el peru": "10",
                "roberto helbert sanchez palomino": "10",
                "votos en blanco": "80",
                "votos nulos": "81",
                "votos impugnados": "82",
            }
            return explicit.get(label, f"sv_nacional_{fallback_index}")

        now = self.now_iso()
        counts: dict[str, int] = {"nacional": 0, "departamentos": 0, "provincias": 0, "cobertura": 0}

        nac_file = sv_resumen_dir / "resumen_nacional.txt"
        nacionales: list[tuple[Any, ...]] = []
        if nac_file.exists():
            with nac_file.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for idx, row in enumerate(reader, start=1):
                    pid = _infer_pid(row, idx)
                    nacionales.append((
                        pid,
                        str(row.get("nombre_candidato", "")).strip(),
                        str(row.get("nombre_agrupacion_politica", "")).strip(),
                        _safe_int(row.get("votos_validos", "0")),
                        _safe_float(row.get("pct_votos_validos", "0")),
                        _safe_float(row.get("pct_votos_emitidos", "0")),
                        _safe_float(row.get("actas_contabilizadas_pct", "0")),
                        _safe_int(row.get("contabilizadas", "0")),
                        _safe_int(row.get("total_actas", "0")),
                        _safe_float(row.get("participacion_ciudadana", "0")),
                        str(row.get("fecha_actualizacion", "")).strip(),
                        str(row.get("fuente", "")).strip(),
                        now,
                    ))

        dept_file = sv_resumen_dir / "resumen_departamentos.txt"
        departamentos: list[tuple[Any, ...]] = []
        if dept_file.exists():
            with dept_file.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    ubigeo = str(row.get("ubigeo", "")).strip()
                    pid = str(row.get("partido_id", "")).strip()
                    if not ubigeo or not pid:
                        continue
                    departamentos.append((
                        ubigeo, pid,
                        str(row.get("nombre_candidato", "")).strip(),
                        str(row.get("nombre_agrupacion_politica", "")).strip(),
                        _safe_int(row.get("votos_validos", "0")),
                        _safe_float(row.get("pct_votos_validos", "0")),
                        _safe_float(row.get("pct_votos_emitidos", "0")),
                        _safe_int(row.get("total_votos_validos_geo", "0")),
                        _safe_int(row.get("total_votos_emitidos_geo", "0")),
                        str(row.get("fuente", "")).strip(),
                        now,
                    ))

        prov_file = sv_resumen_dir / "resumen_provincias.txt"
        provincias: list[tuple[Any, ...]] = []
        if prov_file.exists():
            with prov_file.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    ubigeo = str(row.get("ubigeo", "")).strip()
                    pid = str(row.get("partido_id", "")).strip()
                    if not ubigeo or not pid:
                        continue
                    provincias.append((
                        ubigeo, pid,
                        str(row.get("nombre_candidato", "")).strip(),
                        str(row.get("nombre_agrupacion_politica", "")).strip(),
                        "",
                        _safe_int(row.get("votos_validos", "0")),
                        _safe_float(row.get("pct_votos_validos", "0")),
                        _safe_float(row.get("pct_votos_emitidos", "0")),
                        _safe_int(row.get("total_votos_validos_geo", "0")),
                        _safe_int(row.get("total_votos_emitidos_geo", "0")),
                        str(row.get("fuente", "")).strip(),
                        now,
                    ))

        cob_file = sv_resumen_dir / "resumen_cobertura_departamentos.txt"
        coberturas: list[tuple[Any, ...]] = []
        if cob_file.exists():
            with cob_file.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    ubigeo = str(row.get("ubigeo", "")).strip()
                    if not ubigeo:
                        continue
                    coberturas.append((
                        ubigeo,
                        str(row.get("nombre_departamento", "")).strip(),
                        _safe_int(row.get("actas_contabilizadas", "0")),
                        _safe_float(row.get("pct_actas_contabilizadas", "0")),
                        str(row.get("fuente", "")).strip(),
                        now,
                    ))

        with self._connect() as conn:
            try:
                conn.execute("BEGIN IMMEDIATE")
                conn.execute("DELETE FROM sv_resumen_nacional")
                if nacionales:
                    conn.executemany(
                        """INSERT INTO sv_resumen_nacional
                        (partido_id, nombre_candidato, nombre_agrupacion, votos_validos,
                         pct_votos_validos, pct_votos_emitidos, actas_contabilizadas_pct,
                         contabilizadas, total_actas, participacion_ciudadana,
                         fecha_actualizacion, fuente, loaded_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        nacionales,
                    )
                    counts["nacional"] = len(nacionales)

                conn.execute("DELETE FROM sv_resumen_departamentos")
                if departamentos:
                    conn.executemany(
                        """INSERT INTO sv_resumen_departamentos
                        (ubigeo, partido_id, nombre_candidato, nombre_agrupacion,
                         votos_validos, pct_votos_validos, pct_votos_emitidos,
                         total_votos_validos_geo, total_votos_emitidos_geo, fuente, loaded_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                        departamentos,
                    )
                    counts["departamentos"] = len(departamentos)

                conn.execute("DELETE FROM sv_resumen_provincias")
                if provincias:
                    conn.executemany(
                        """INSERT INTO sv_resumen_provincias
                        (ubigeo, partido_id, nombre_candidato, nombre_agrupacion, nombre_geo,
                         votos_validos, pct_votos_validos, pct_votos_emitidos,
                         total_votos_validos_geo, total_votos_emitidos_geo, fuente, loaded_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                        provincias,
                    )
                    # Populate nombre_geo: ubigeo_reniec/ubicaciones_sv have district-level
                    # ubigeos, while sv_resumen_provincias has province/country ubigeos.
                    # Match on first 4 chars (province prefix) to find a representative row.
                    conn.execute(
                        """UPDATE sv_resumen_provincias
                        SET nombre_geo = COALESCE(
                            NULLIF((
                                SELECT CASE
                                    WHEN substr(sv_resumen_provincias.ubigeo, 1, 1) = '9'
                                        THEN COALESCE(NULLIF(u.pais, ''), NULLIF(u.ciudad, ''), NULLIF(u.continente, ''))
                                    ELSE COALESCE(NULLIF(u.provincia, ''), NULLIF(u.departamento, ''), NULLIF(u.distrito, ''))
                                END
                                FROM ubicaciones_sv u
                                WHERE u.ubigeo LIKE SUBSTR(sv_resumen_provincias.ubigeo, 1, 4) || '%'
                                LIMIT 1
                            ), ''),
                            NULLIF(nombre_geo, '')
                        )
                        WHERE nombre_geo IS NULL OR nombre_geo = ''"""
                    )
                    counts["provincias"] = len(provincias)

                conn.execute("DELETE FROM sv_resumen_cobertura")
                if coberturas:
                    conn.executemany(
                        """INSERT INTO sv_resumen_cobertura
                        (ubigeo, nombre_departamento, actas_contabilizadas,
                         pct_actas_contabilizadas, fuente, loaded_at)
                        VALUES (?,?,?,?,?,?)""",
                        coberturas,
                    )
                    counts["cobertura"] = len(coberturas)
                conn.commit()
            except Exception:
                conn.rollback()
                raise

        return counts

    def populate_sv_nombre_geo(self) -> int:
        """Populate nombre_geo in sv_resumen_provincias for existing rows.

        Uses LIKE on first 4 chars of ubigeo to match district-level rows in
        ubicaciones_sv (which don't have exact province-level ubigeo matches).
        Returns number of rows updated.
        """
        with self._connect() as conn:
            result = conn.execute(
                """UPDATE sv_resumen_provincias
                SET nombre_geo = COALESCE(
                    NULLIF((
                        SELECT CASE
                            WHEN substr(sv_resumen_provincias.ubigeo, 1, 1) = '9'
                                THEN COALESCE(NULLIF(u.pais, ''), NULLIF(u.ciudad, ''), NULLIF(u.continente, ''))
                            ELSE COALESCE(NULLIF(u.provincia, ''), NULLIF(u.departamento, ''), NULLIF(u.distrito, ''))
                        END
                        FROM ubicaciones_sv u
                        WHERE u.ubigeo LIKE SUBSTR(sv_resumen_provincias.ubigeo, 1, 4) || '%'
                        LIMIT 1
                    ), ''),
                    NULLIF(nombre_geo, '')
                )
                WHERE nombre_geo IS NULL OR nombre_geo = ''"""
            )
            return result.rowcount


        """Rebuild sv_agg_distrito and sv_agg_ciudad via CTAS from raw tables."""
        now = self.now_iso()
        with self._connect() as conn:
            try:
                conn.execute("BEGIN IMMEDIATE")
                conn.execute("DROP TABLE IF EXISTS _sv_base")
                conn.execute(
                    """
                    CREATE TEMP TABLE _sv_base AS
                    SELECT
                        m.codigo_mesa,
                        m.id_ubigeo AS ubigeo,
                        COALESCE(u.departamento, '') AS departamento,
                        COALESCE(u.provincia, '') AS provincia,
                        COALESCE(u.distrito, '') AS distrito,
                        COALESCE(NULLIF(u.ciudad, ''), NULLIF(u.distrito, ''), NULLIF(u.provincia, ''), '') AS ciudad,
                        m.codigo_estado_acta,
                        v.partido_id,
                        COALESCE(a.nombre, COALESCE(v.partido_id, '')) AS nombre_candidato,
                        COALESCE(v.votos, 0) AS votos
                    FROM mesas_sv m
                    LEFT JOIN ubicaciones_sv u ON u.ubigeo = m.id_ubigeo
                    LEFT JOIN votos_sv v ON v.codigo_mesa = m.codigo_mesa
                    LEFT JOIN agrupaciones_sv a ON a.partido_id = v.partido_id
                    """
                )

                conn.execute("DROP TABLE IF EXISTS sv_agg_distrito_new")
                conn.execute(
                    """CREATE TABLE sv_agg_distrito_new (
                        ubigeo TEXT NOT NULL,
                        partido_id TEXT NOT NULL,
                        nombre_candidato TEXT,
                        votos INTEGER,
                        total_mesas INTEGER,
                        mesas_contabilizadas INTEGER,
                        rebuilt_at TEXT NOT NULL,
                        PRIMARY KEY (ubigeo, partido_id)
                    )"""
                )
                conn.execute(
                    f"""
                    INSERT INTO sv_agg_distrito_new
                    (ubigeo, partido_id, nombre_candidato, votos, total_mesas, mesas_contabilizadas, rebuilt_at)
                    SELECT
                        ubigeo,
                        partido_id,
                        MAX(nombre_candidato) AS nombre_candidato,
                        SUM(votos) AS votos,
                        COUNT(DISTINCT codigo_mesa) AS total_mesas,
                        COUNT(DISTINCT CASE WHEN codigo_estado_acta='C' THEN codigo_mesa END) AS mesas_contabilizadas,
                        '{now}' AS rebuilt_at
                    FROM _sv_base
                    WHERE partido_id IS NOT NULL AND partido_id != '' AND ubigeo != ''
                    GROUP BY ubigeo, partido_id
                    """
                )
                conn.execute("DROP TABLE IF EXISTS sv_agg_distrito")
                conn.execute("ALTER TABLE sv_agg_distrito_new RENAME TO sv_agg_distrito")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_sv_agg_distrito_ubigeo ON sv_agg_distrito (ubigeo)")

                conn.execute("DROP TABLE IF EXISTS sv_agg_ciudad_new")
                conn.execute(
                    """CREATE TABLE sv_agg_ciudad_new (
                        ubigeo TEXT NOT NULL,
                        ciudad TEXT NOT NULL,
                        partido_id TEXT NOT NULL,
                        nombre_candidato TEXT,
                        votos INTEGER,
                        total_mesas INTEGER,
                        mesas_contabilizadas INTEGER,
                        rebuilt_at TEXT NOT NULL,
                        PRIMARY KEY (ubigeo, ciudad, partido_id)
                    )"""
                )
                conn.execute(
                    f"""
                    INSERT INTO sv_agg_ciudad_new
                    (ubigeo, ciudad, partido_id, nombre_candidato, votos, total_mesas, mesas_contabilizadas, rebuilt_at)
                    SELECT
                        ubigeo,
                        ciudad,
                        partido_id,
                        MAX(nombre_candidato) AS nombre_candidato,
                        SUM(votos) AS votos,
                        COUNT(DISTINCT codigo_mesa) AS total_mesas,
                        COUNT(DISTINCT CASE WHEN codigo_estado_acta='C' THEN codigo_mesa END) AS mesas_contabilizadas,
                        '{now}' AS rebuilt_at
                    FROM _sv_base
                    WHERE partido_id IS NOT NULL AND partido_id != '' AND ciudad != ''
                    GROUP BY ubigeo, ciudad, partido_id
                    """
                )
                conn.execute("DROP TABLE IF EXISTS sv_agg_ciudad")
                conn.execute("ALTER TABLE sv_agg_ciudad_new RENAME TO sv_agg_ciudad")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_sv_agg_ciudad_ubigeo ON sv_agg_ciudad (ubigeo)")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_sv_agg_ciudad_nombre ON sv_agg_ciudad (ciudad)")

                n_dist = conn.execute("SELECT COUNT(*) AS c FROM sv_agg_distrito").fetchone()["c"]
                n_city = conn.execute("SELECT COUNT(*) AS c FROM sv_agg_ciudad").fetchone()["c"]
                conn.commit()
            except Exception:
                conn.rollback()
                raise

        return {"distrito": int(n_dist), "ciudad": int(n_city)}

    def get_mesa_sv_from_local(self, codigo_mesa: str) -> dict[str, Any] | None:
        """Build SV mesa bundle from local tables."""
        with self._connect() as conn:
            mesa_row = conn.execute(
                "SELECT * FROM mesas_sv WHERE codigo_mesa = ?", (codigo_mesa,)
            ).fetchone()
            if mesa_row is None:
                return None

            votos_rows = conn.execute(
                """SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre_partido, v.votos
                   FROM votos_sv v
                   LEFT JOIN agrupaciones_sv a ON a.partido_id = v.partido_id
                   WHERE v.codigo_mesa = ?
                   ORDER BY v.votos DESC""",
                (codigo_mesa,),
            ).fetchall()

            ub = str(mesa_row["id_ubigeo"] or "")
            loc_row = conn.execute(
                "SELECT departamento, provincia, distrito, ciudad FROM ubicaciones_sv WHERE ubigeo = ?",
                (ub,),
            ).fetchone()

        mesa_data: dict[str, Any] = {
            "codigo_mesa": codigo_mesa,
            "ubigeo": ub,
            "nombre_local": str(mesa_row["nombre_local"] or ""),
            "electores_habiles": int(mesa_row["electores_habiles"] or 0),
            "votos_emitidos": int(mesa_row["votos_emitidos"] or 0),
            "votos_validos": int(mesa_row["votos_validos"] or 0),
            "codigo_estado_acta": str(mesa_row["codigo_estado_acta"] or ""),
            "id_eleccion": 10,
        }
        if loc_row:
            mesa_data.update({
                "departamento": str(loc_row["departamento"] or ""),
                "provincia": str(loc_row["provincia"] or ""),
                "distrito": str(loc_row["distrito"] or ""),
                "ciudad": str(loc_row["ciudad"] or ""),
            })

        votos = [
            {"partido_id": str(r["partido_id"]), "nombre_partido": str(r["nombre_partido"]), "votos": int(r["votos"] or 0)}
            for r in votos_rows
        ]

        return {
            "codigo_mesa": codigo_mesa,
            "found": True,
            "mesa_data": mesa_data,
            "agrupaciones": [{"partido_id": v["partido_id"], "nombre": v["nombre_partido"]} for v in votos],
            "votos": votos,
            "source": "local_db_sv",
            "id_eleccion": 10,
        }

    def query_sv_nacional(self) -> list[dict[str, Any]]:
        if self.denorm_available:
            try:
                conn = self._connect_denorm()
                rows = conn.execute("""
                    SELECT partido_id, nombre_partido AS nombre_agrupacion,
                           candidato AS nombre_candidato,
                           votos AS votos_validos, pct_votos_validos,
                           mesas_contabilizadas AS contabilizadas,
                           total_mesas AS total_actas,
                           total_electores_habiles, total_votos_emitidos
                    FROM fact_votos_nacional
                    WHERE election_year=2026 AND vuelta=2 AND es_especial=0
                    ORDER BY votos DESC LIMIT 20
                """).fetchall()
                conn.close()
                result = []
                for r in rows:
                    d = dict(r)
                    total_actas = d.get("total_actas") or 0
                    contabilizadas = d.get("contabilizadas") or 0
                    d["actas_contabilizadas_pct"] = round(contabilizadas / total_actas * 100, 2) if total_actas else 0.0
                    electores = d.get("total_electores_habiles") or 0
                    emitidos = d.get("total_votos_emitidos") or 0
                    d["participacion_ciudadana"] = round(emitidos / electores * 100, 2) if electores else 0.0
                    result.append(d)
                return result
            except Exception as e:
                _logger.debug("denorm fast-path failed for query_sv_nacional, falling back to OLTP: %s", e)
        with self._connect() as conn:
            rows = conn.execute(
                """SELECT partido_id, nombre_candidato, nombre_agrupacion, votos_validos,
                   pct_votos_validos, pct_votos_emitidos, actas_contabilizadas_pct,
                   contabilizadas, total_actas, participacion_ciudadana, fecha_actualizacion
                   FROM sv_resumen_nacional
                   ORDER BY votos_validos DESC"""
            ).fetchall()
        return [dict(r) for r in rows]

    def query_sv_geo(self, nivel: str, ubigeo: str | None = None, nombre: str | None = None, top_n: int = 10) -> list[dict[str, Any]]:
        """
        Query SV results by geographic level.
        nivel: 'nacional' | 'continente' | 'pais_exterior' | 'departamento' | 'provincia' | 'distrito' | 'ciudad'
        """
        if self.denorm_available:
            try:
                return self._query_sv_geo_denorm(nivel, ubigeo, nombre, top_n)
            except Exception as e:
                _logger.debug("denorm sv_geo fast-path failed: %s", e)
        with self._connect() as conn:
            if nivel == "nacional":
                rows = conn.execute(
                    "SELECT partido_id, nombre_candidato, nombre_agrupacion, votos_validos, pct_votos_validos FROM sv_resumen_nacional ORDER BY votos_validos DESC"
                ).fetchall()
            elif nivel == "continente":
                if ubigeo:
                    rows = conn.execute(
                        "SELECT ubigeo, partido_id, nombre_candidato, nombre_agrupacion, votos_validos, pct_votos_validos FROM sv_resumen_departamentos WHERE ubigeo = ? ORDER BY votos_validos DESC",
                        (ubigeo,),
                    ).fetchall()
                else:
                    rows = conn.execute(
                        "SELECT ubigeo, partido_id, nombre_candidato, nombre_agrupacion, votos_validos, pct_votos_validos FROM sv_resumen_departamentos WHERE CAST(ubigeo AS TEXT) >= '910000' ORDER BY ubigeo, votos_validos DESC"
                    ).fetchall()
            elif nivel == "departamento":
                if ubigeo:
                    rows = conn.execute(
                        "SELECT ubigeo, partido_id, nombre_candidato, nombre_agrupacion, votos_validos, pct_votos_validos, total_votos_validos_geo FROM sv_resumen_departamentos WHERE ubigeo = ? ORDER BY votos_validos DESC",
                        (ubigeo,),
                    ).fetchall()
                else:
                    # Fetch with nombre_departamento for Python-side accent-insensitive filtering
                    all_dept_rows = conn.execute(
                        """SELECT d.ubigeo, d.partido_id, d.nombre_candidato, d.nombre_agrupacion,
                                  d.votos_validos, d.pct_votos_validos,
                                  COALESCE(c.nombre_departamento,'') AS nombre_departamento
                           FROM sv_resumen_departamentos d
                           LEFT JOIN sv_resumen_cobertura c ON c.ubigeo = d.ubigeo
                           WHERE CAST(d.ubigeo AS TEXT) < '910000'
                           ORDER BY d.ubigeo, d.votos_validos DESC
                           LIMIT ?""",
                        (top_n * 30,),
                    ).fetchall()
                    if nombre:
                        norm_n = _norm_text(nombre)
                        rows = [r for r in all_dept_rows if norm_n in _norm_text(str(r["nombre_departamento"] or ""))]
                    else:
                        rows = all_dept_rows
            elif nivel == "provincia":
                if ubigeo:
                    rows = conn.execute(
                        "SELECT ubigeo, partido_id, nombre_candidato, nombre_agrupacion, nombre_geo, votos_validos, pct_votos_validos FROM sv_resumen_provincias WHERE ubigeo = ? ORDER BY votos_validos DESC",
                        (ubigeo,),
                    ).fetchall()
                elif nombre:
                    rows = conn.execute(
                        "SELECT ubigeo, partido_id, nombre_candidato, nombre_agrupacion, nombre_geo, votos_validos, pct_votos_validos FROM sv_resumen_provincias WHERE nombre_geo LIKE ? ORDER BY ubigeo, votos_validos DESC LIMIT ?",
                        (f"%{nombre}%", top_n * 5),
                    ).fetchall()
                else:
                    rows = conn.execute(
                        "SELECT ubigeo, partido_id, nombre_candidato, nombre_agrupacion, nombre_geo, votos_validos, pct_votos_validos FROM sv_resumen_provincias WHERE CAST(ubigeo AS TEXT) < '910000' ORDER BY ubigeo, votos_validos DESC LIMIT ?",
                        (top_n * 5,),
                    ).fetchall()
            elif nivel == "pais_exterior":
                if nombre:
                    rows = conn.execute(
                        "SELECT ubigeo, partido_id, nombre_candidato, nombre_agrupacion, nombre_geo, votos_validos, pct_votos_validos FROM sv_resumen_provincias WHERE CAST(ubigeo AS TEXT) >= '910000' AND nombre_geo LIKE ? ORDER BY votos_validos DESC",
                        (f"%{nombre}%",),
                    ).fetchall()
                else:
                    rows = conn.execute(
                        "SELECT ubigeo, partido_id, nombre_candidato, nombre_agrupacion, nombre_geo, votos_validos, pct_votos_validos FROM sv_resumen_provincias WHERE CAST(ubigeo AS TEXT) >= '910000' ORDER BY ubigeo, votos_validos DESC LIMIT ?",
                        (top_n * 5,),
                    ).fetchall()
            elif nivel == "distrito":
                if ubigeo:
                    rows = conn.execute(
                        "SELECT ubigeo, partido_id, nombre_candidato, votos, total_mesas, mesas_contabilizadas FROM sv_agg_distrito WHERE ubigeo = ? ORDER BY votos DESC",
                        (ubigeo,),
                    ).fetchall()
                else:
                    rows = conn.execute(
                        "SELECT ubigeo, partido_id, nombre_candidato, votos, total_mesas FROM sv_agg_distrito ORDER BY votos DESC LIMIT ?",
                        (top_n * 3,),
                    ).fetchall()
            elif nivel == "ciudad":
                if nombre:
                    rows = conn.execute(
                        "SELECT ubigeo, ciudad, partido_id, nombre_candidato, votos, total_mesas FROM sv_agg_ciudad WHERE ciudad LIKE ? ORDER BY votos DESC LIMIT ?",
                        (f"%{nombre}%", top_n * 3),
                    ).fetchall()
                elif ubigeo:
                    rows = conn.execute(
                        "SELECT ubigeo, ciudad, partido_id, nombre_candidato, votos, total_mesas FROM sv_agg_ciudad WHERE ubigeo = ? ORDER BY votos DESC",
                        (ubigeo,),
                    ).fetchall()
                else:
                    rows = conn.execute(
                        "SELECT ubigeo, ciudad, partido_id, nombre_candidato, votos, total_mesas FROM sv_agg_ciudad ORDER BY votos DESC LIMIT ?",
                        (top_n * 3,),
                    ).fetchall()
            else:
                rows = []

        return [dict(r) for r in rows]

    def get_sv_cobertura(self) -> list[dict[str, Any]]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT ubigeo, nombre_departamento, actas_contabilizadas, pct_actas_contabilizadas FROM sv_resumen_cobertura ORDER BY ubigeo"
            ).fetchall()
        return [dict(r) for r in rows]

    def get_sv_reasignados(self, dpto: str | None = None, motivo: str | None = None) -> list[dict[str, Any]]:
        with self._connect() as conn:
            if dpto:
                rows = conn.execute(
                    "SELECT * FROM locales_reasignados_sv WHERE UPPER(dpto) LIKE ? ORDER BY nro",
                    (f"%{dpto.upper()}%",),
                ).fetchall()
            elif motivo:
                rows = conn.execute(
                    "SELECT * FROM locales_reasignados_sv WHERE UPPER(motivo) LIKE ? ORDER BY nro",
                    (f"%{motivo.upper()}%",),
                ).fetchall()
            else:
                rows = conn.execute("SELECT * FROM locales_reasignados_sv ORDER BY nro").fetchall()
        return [dict(r) for r in rows]

    def total_mesas_sv_local(self) -> int:
        with self._connect() as conn:
            row = conn.execute("SELECT COUNT(*) AS c FROM mesas_sv").fetchone()
        return int((row or {"c": 0})["c"])

    # -----------------------------------------------------------------
    # Estado de actas — primera vuelta (1V, idEleccion=10)
    # -----------------------------------------------------------------
    # En la 1V el estado del acta en `mesas_data` se guarda como string libre
    # ("Contabilizada", "Observada", "Pendiente"). Esta función entrega la
    # misma estructura que `get_sv_estado_actas` para que la tool de servidor
    # pueda decidir entre 1V y 2V por `id_eleccion` sin duplicar lógica.

    _ESTADO_1V_DESCRIPCIONES = {
        "C": "Contabilizada",
        "E": "Observada / para envío al JEE",
        "P": "Pendiente",
    }

    @staticmethod
    def _normalize_estado_acta_1v(raw: str | None) -> str:
        norm = (raw or "").strip().upper()
        if norm in {"CONTABILIZADA", "C"}:
            return "C"
        if norm in {
            "OBSERVADA",
            "E",
            "EN PROCESO",
            "OBSERVADO",
            "PARA ENVIO AL JEE",
            "PARA ENVÍO AL JEE",
        }:
            return "E"
        if norm in {"PENDIENTE", "P", "NO INSTALADA"}:
            return "P"
        return norm or "?"

    # -----------------------------------------------------------------
    # Totales nacionales 1V — usados por margen_pase y claim_verifier
    # -----------------------------------------------------------------
    def get_totales_nacionales_1v(self) -> dict[str, Any]:
        """Totales nacionales agregados desde `mesas_data` (1V).

        Devuelve los denominadores oficiales (padrón, emitidos, válidos,
        blancos, nulos) que permiten verificar cualquier claim porcentual.
        """
        if self.denorm_available:
            try:
                conn = self._connect_denorm()
                row = conn.execute("""
                    SELECT total_mesas, mesas_contabilizadas,
                           total_electores_habiles, total_votos_emitidos, total_votos_validos
                    FROM fact_votos_nacional
                    WHERE election_year=2026 AND vuelta=1 AND es_especial=0
                    LIMIT 1
                """).fetchone()
                conn.close()
                if row:
                    return dict(row)
            except Exception as e:
                _logger.debug("denorm fast-path failed for get_totales_nacionales_1v, falling back to OLTP: %s", e)
        with self._connect() as conn:
            row = conn.execute(
                """SELECT COUNT(*) AS mesas,
                          SUM(electores_habiles) AS electores_habiles,
                          SUM(votos_emitidos)    AS votos_emitidos,
                          SUM(votos_validos)     AS votos_validos,
                          SUM(blancos)           AS blancos,
                          SUM(nulos)             AS nulos,
                          SUM(impugnados)        AS impugnados
                   FROM mesas_data"""
            ).fetchone()
            fa_row = conn.execute(
                "SELECT MAX(fetched_at) AS f FROM mesas_data"
            ).fetchone()

        def _col(r: sqlite3.Row | None, key: str) -> int:
            if r is None:
                return 0
            try:
                val = r[key]
            except (IndexError, KeyError):
                return 0
            return int(val or 0)

        electores = _col(row, "electores_habiles")
        emitidos = _col(row, "votos_emitidos")
        validos = _col(row, "votos_validos")
        blancos = _col(row, "blancos")
        nulos = _col(row, "nulos")
        impug = _col(row, "impugnados")
        fecha = ""
        if fa_row is not None:
            try:
                fecha = str(fa_row["f"] or "")
            except (IndexError, KeyError):
                fecha = ""

        def _pct(n: int, d: int) -> float:
            return round(100.0 * n / d, 4) if d else 0.0

        return {
            "id_eleccion": 10,
            "mesas": _col(row, "mesas"),
            "electores_habiles": electores,
            "votos_emitidos": emitidos,
            "votos_validos": validos,
            "votos_blancos": blancos,
            "votos_nulos": nulos,
            "votos_impugnados": impug,
            "participacion_pct": _pct(emitidos, electores),
            "ausentismo_total": max(0, electores - emitidos),
            "ausentismo_pct": _pct(max(0, electores - emitidos), electores),
            "validos_pct_emitidos": _pct(validos, emitidos),
            "blancos_pct_emitidos": _pct(blancos, emitidos),
            "nulos_pct_emitidos": _pct(nulos, emitidos),
            "fecha_actualizacion": fecha,
        }

    def get_top_partidos_1v(self, top_n: int = 10) -> list[dict[str, Any]]:
        """Top-N partidos a nivel nacional en 1V (incluye blancos/nulos)."""
        top_n = max(1, min(int(top_n or 10), 50))
        if self.denorm_available:
            try:
                conn = self._connect_denorm()
                rows = conn.execute("""
                    SELECT partido_id, nombre_partido, candidato, votos, pct_votos_validos
                    FROM fact_votos_nacional
                    WHERE election_year=2026 AND vuelta=1 AND es_especial=0
                    ORDER BY votos DESC LIMIT ?
                """, (top_n,)).fetchall()
                conn.close()
                return [dict(r) for r in rows]
            except Exception as e:
                _logger.debug("denorm fast-path failed for get_top_partidos_1v, falling back to OLTP: %s", e)
        with self._connect() as conn:
            rows = conn.execute(
                """SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre,
                          SUM(v.votos) AS total
                   FROM votos v
                   LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                   GROUP BY v.partido_id ORDER BY total DESC LIMIT ?""",
                (top_n,),
            ).fetchall()
        return [
            {
                "partido_id": str(r["partido_id"]),
                "nombre": str(r["nombre"] or ""),
                "votos": int(r["total"] or 0),
            }
            for r in rows
        ]

    # -----------------------------------------------------------------
    # Margen para pasar a 2da vuelta — útil para refutar claims
    # cuantitativos del tipo "nos faltaron N votos" / "perdimos X%".
    # -----------------------------------------------------------------
    @staticmethod
    def _is_partido_real(partido_id: str, nombre: str) -> bool:
        pid = str(partido_id or "").strip()
        nom_norm = _norm_text(nombre)
        if pid in {"80", "81", "82"}:  # blanco / nulo / impugnado en SV
            return False
        if any(kw in nom_norm for kw in ("blanco", "nulo", "impugnad")):
            return False
        return True

    def get_margen_pase(
        self,
        partido: str,
        id_eleccion: int = 10,
        top_n: int = 8,
    ) -> dict[str, Any]:
        """Calcula la brecha del partido vs los puestos vecinos en 1V.

        Args:
            partido: partido_id (ej. "35") o nombre (ej. "renovacion popular",
                "lopez aliaga"). Se hace matching accent-insensitive.
            id_eleccion: 10 (1V) — único soportado hoy. El concepto de
                "pase a 2V" no aplica a SV.
            top_n: cantidad de competidores a mostrar en `ranking_top` (1-25).

        Devuelve:
            - candidato_objetivo: {partido_id, nombre, votos, posicion}
            - margen_vs_anterior: {posicion, partido, diferencia_votos,
                                   pct_padron, pct_emitidos, pct_validos}
            - margen_vs_lider: ídem (posición 1)
            - ranking_top: top-N partidos con votos absolutos y porcentajes
            - denominadores: {padron, emitidos, validos}
            - claim_helper: cifras de referencia para validar "X% = Y votos"
        """
        if int(id_eleccion or 10) != 10:
            raise ValueError(
                "get_margen_pase solo aplica a 1V (id_eleccion=10). "
                "Para SV use get_sv_estado_actas (`escenario_jee_aceptadas`)."
            )

        top_n = max(1, min(int(top_n or 8), 25))

        with self._connect() as conn:
            ranking_rows = conn.execute(
                """SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre,
                          SUM(v.votos) AS total
                   FROM votos v
                   LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                   GROUP BY v.partido_id
                   ORDER BY total DESC"""
            ).fetchall()

        ranking_real: list[dict[str, Any]] = []
        rank = 1
        for r in ranking_rows:
            pid = str(r["partido_id"])
            nombre = str(r["nombre"] or "")
            if not self._is_partido_real(pid, nombre):
                continue
            ranking_real.append({
                "rank": rank,
                "partido_id": pid,
                "nombre": nombre,
                "votos": int(r["total"] or 0),
            })
            rank += 1

        if not ranking_real:
            raise ValueError(
                "No hay datos de candidatos en SQLite — corre onpe_bootstrap_snapshot primero."
            )

        target_query = (partido or "").strip()
        if not target_query:
            raise ValueError("partido es obligatorio (id o nombre)")

        target_idx: int | None = None
        target_q_norm = _norm_text(target_query)

        # 1) match por partido_id exacto
        for i, row in enumerate(ranking_real):
            if row["partido_id"] == target_query:
                target_idx = i
                break

        # 2) match por nombre (substring accent-insensitive)
        if target_idx is None:
            for i, row in enumerate(ranking_real):
                if target_q_norm and target_q_norm in _norm_text(row["nombre"]):
                    target_idx = i
                    break

        # 3) alias frecuentes
        if target_idx is None:
            alias_map = {
                "lopez aliaga": "renovacion popular",
                "rafael lopez aliaga": "renovacion popular",
                "keiko fujimori": "fuerza popular",
                "keiko": "fuerza popular",
                "fujimori": "fuerza popular",
                "sanchez": "juntos por el peru",
                "roberto sanchez": "juntos por el peru",
                "nieto": "partido del buen gobierno",
                "belmont": "partido civico obras",
            }
            for alias, real_name in alias_map.items():
                if alias in target_q_norm:
                    real_norm = _norm_text(real_name)
                    for i, row in enumerate(ranking_real):
                        if real_norm in _norm_text(row["nombre"]):
                            target_idx = i
                            break
                    if target_idx is not None:
                        break

        if target_idx is None:
            raise ValueError(
                f"Partido '{partido}' no encontrado. Top-3: "
                + ", ".join(f"{r['nombre']} (id={r['partido_id']})" for r in ranking_real[:3])
            )

        totales = self.get_totales_nacionales_1v()
        padron = totales["electores_habiles"]
        emitidos = totales["votos_emitidos"]
        validos = totales["votos_validos"]

        target_row = ranking_real[target_idx]

        def _margen(other_idx: int) -> dict[str, Any]:
            other = ranking_real[other_idx]
            diff = other["votos"] - target_row["votos"]
            def _pct(d: int) -> float:
                return round(100.0 * diff / d, 6) if d else 0.0
            return {
                "rank": other["rank"],
                "partido_id": other["partido_id"],
                "nombre": other["nombre"],
                "votos_objetivo": other["votos"],
                "diferencia_votos": diff,
                "pct_padron": _pct(padron),
                "pct_emitidos": _pct(emitidos),
                "pct_validos": _pct(validos),
            }

        margen_anterior = _margen(target_idx - 1) if target_idx > 0 else None
        margen_lider = _margen(0) if target_idx > 0 else None

        # Vista compacta del ranking con porcentajes
        ranking_top = []
        for r in ranking_real[: top_n]:
            v = r["votos"]
            ranking_top.append({
                "rank": r["rank"],
                "partido_id": r["partido_id"],
                "nombre": r["nombre"],
                "votos": v,
                "pct_padron": round(100.0 * v / padron, 4) if padron else 0.0,
                "pct_emitidos": round(100.0 * v / emitidos, 4) if emitidos else 0.0,
                "pct_validos": round(100.0 * v / validos, 4) if validos else 0.0,
            })

        # Ayudante para validar claims tipo "1.2% = 100 mil votos"
        def _ref(pct: float) -> dict[str, int]:
            return {
                "pct_padron_=>_votos": int(round(pct / 100 * padron)),
                "pct_emitidos_=>_votos": int(round(pct / 100 * emitidos)),
                "pct_validos_=>_votos": int(round(pct / 100 * validos)),
            }

        return {
            "id_eleccion": 10,
            "candidato_objetivo": {
                "partido_id": target_row["partido_id"],
                "nombre": target_row["nombre"],
                "votos": target_row["votos"],
                "posicion": target_row["rank"],
                "pct_padron": round(100.0 * target_row["votos"] / padron, 4) if padron else 0.0,
                "pct_emitidos": round(100.0 * target_row["votos"] / emitidos, 4) if emitidos else 0.0,
                "pct_validos": round(100.0 * target_row["votos"] / validos, 4) if validos else 0.0,
            },
            "margen_vs_anterior": margen_anterior,
            "margen_vs_lider": margen_lider,
            "ranking_top": ranking_top,
            "denominadores": {
                "padron_habil": padron,
                "votos_emitidos": emitidos,
                "votos_validos": validos,
            },
            "claim_helper": {
                "0.5%_equivale_a": _ref(0.5),
                "1.0%_equivale_a": _ref(1.0),
                "1.2%_equivale_a": _ref(1.2),
                "2.0%_equivale_a": _ref(2.0),
                "5.0%_equivale_a": _ref(5.0),
            },
            "fecha_actualizacion": totales["fecha_actualizacion"],
            "nota_metodologia": (
                "Ranking excluye votos en blanco/nulos/impugnados (no compiten "
                "por curul). Los denominadores incluyen el padrón hábil "
                "completo, votos emitidos (incluye blancos+nulos) y votos "
                "válidos (sólo a candidatos)."
            ),
        }

    def get_estado_actas_1v(
        self,
        ubigeo_prefix: str | None = None,
        top_geo: int = 10,
    ) -> dict[str, Any]:
        """Estado de actas para PRIMERA VUELTA (idEleccion=10).

        Estructura idéntica a `get_sv_estado_actas` para uniformidad,
        pero leyendo de `mesas_data`, `votos` y `agrupaciones`. En 1V
        el escrutinio cerró al 100% (`Contabilizada`), por lo que las
        listas de "observadas" y "pendientes" vienen vacías; la utilidad
        principal es CONTRADECIR claims tipo "faltan N mesas/votos".
        """
        like: str | None = None
        if ubigeo_prefix:
            ubigeo_prefix_clean = str(ubigeo_prefix).strip()
            if ubigeo_prefix_clean:
                # mesas_data guarda ubigeos ONPE con leading-zero strippeado
                # ('040101' → '40101'). Para que el LIKE matchee, normalizamos.
                ub_strip = ubigeo_prefix_clean.lstrip("0") or ubigeo_prefix_clean
                like = f"{ub_strip}%"
                ubigeo_prefix = ubigeo_prefix_clean

        with self._connect() as conn:
            # 1) Conteos por estado (incluyendo NULL)
            if like:
                estado_rows = conn.execute(
                    """SELECT COALESCE(estado_acta,'') AS estado,
                              COUNT(*) AS mesas,
                              SUM(electores_habiles) AS electores_habiles,
                              SUM(votos_emitidos) AS votos_emitidos,
                              SUM(votos_validos) AS votos_validos
                       FROM mesas_data WHERE ubigeo LIKE ?
                       GROUP BY estado_acta""",
                    (like,),
                ).fetchall()
            else:
                estado_rows = conn.execute(
                    """SELECT COALESCE(estado_acta,'') AS estado,
                              COUNT(*) AS mesas,
                              SUM(electores_habiles) AS electores_habiles,
                              SUM(votos_emitidos) AS votos_emitidos,
                              SUM(votos_validos) AS votos_validos
                       FROM mesas_data GROUP BY estado_acta"""
                ).fetchall()

            por_codigo: dict[str, dict[str, Any]] = {}
            for r in estado_rows:
                cod = self._normalize_estado_acta_1v(r["estado"])
                bucket = por_codigo.setdefault(
                    cod,
                    {
                        "codigo": cod,
                        "descripcion": self._ESTADO_1V_DESCRIPCIONES.get(cod, cod),
                        "mesas": 0,
                        "electores_habiles": 0,
                        "votos_emitidos": 0,
                        "votos_validos": 0,
                        "estados_originales": [],
                    },
                )
                bucket["mesas"] += int(r["mesas"] or 0)
                bucket["electores_habiles"] += int(r["electores_habiles"] or 0)
                bucket["votos_emitidos"] += int(r["votos_emitidos"] or 0)
                bucket["votos_validos"] += int(r["votos_validos"] or 0)
                if r["estado"]:
                    bucket["estados_originales"].append(str(r["estado"]))

            por_estado = sorted(
                por_codigo.values(), key=lambda x: x["mesas"], reverse=True
            )

            totales = {
                "mesas": sum(it["mesas"] for it in por_estado),
                "contabilizadas_C": por_codigo.get("C", {}).get("mesas", 0),
                "observadas_E": por_codigo.get("E", {}).get("mesas", 0),
                "pendientes_P": por_codigo.get("P", {}).get("mesas", 0),
                "electores_habiles": sum(it["electores_habiles"] for it in por_estado),
                "votos_emitidos": sum(it["votos_emitidos"] for it in por_estado),
                "votos_validos": sum(it["votos_validos"] for it in por_estado),
            }

            # 2) Votos en mesas no-contabilizadas por partido (E/P)
            if like:
                no_c_rows = conn.execute(
                    """SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre,
                              SUM(v.votos) AS total
                       FROM votos v
                       JOIN mesas_data m ON m.codigo_mesa = v.codigo_mesa
                       LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                       WHERE UPPER(COALESCE(m.estado_acta,'')) NOT IN ('CONTABILIZADA','C')
                         AND m.ubigeo LIKE ?
                       GROUP BY v.partido_id ORDER BY total DESC""",
                    (like,),
                ).fetchall()
            else:
                no_c_rows = conn.execute(
                    """SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre,
                              SUM(v.votos) AS total
                       FROM votos v
                       JOIN mesas_data m ON m.codigo_mesa = v.codigo_mesa
                       LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                       WHERE UPPER(COALESCE(m.estado_acta,'')) NOT IN ('CONTABILIZADA','C')
                       GROUP BY v.partido_id ORDER BY total DESC"""
                ).fetchall()

            votos_no_contabilizados = [
                {
                    "partido_id": str(r["partido_id"]),
                    "nombre": str(r["nombre"] or ""),
                    "votos": int(r["total"] or 0),
                }
                for r in no_c_rows
                if int(r["total"] or 0) > 0
            ]

            # 3) Top departamentos con mesas no-contabilizadas
            geo_top_no_c: list[dict[str, Any]] = []
            if not like and top_geo and int(top_geo) > 0:
                geo_rows = conn.execute(
                    """SELECT SUBSTR(ubigeo,1,2) AS dpto,
                              COUNT(*) AS mesas,
                              SUM(electores_habiles) AS electores
                       FROM mesas_data
                       WHERE UPPER(COALESCE(estado_acta,'')) NOT IN ('CONTABILIZADA','C')
                       GROUP BY dpto ORDER BY mesas DESC LIMIT ?""",
                    (int(top_geo),),
                ).fetchall()
                for r in geo_rows:
                    geo_top_no_c.append({
                        "dpto_prefix": str(r["dpto"] or ""),
                        "mesas_no_C": int(r["mesas"] or 0),
                        "electores": int(r["electores"] or 0),
                    })

            # 4) Última fecha de actualización (usa mesa_cache si está disponible)
            row_fa = conn.execute(
                "SELECT MAX(fetched_at) AS f FROM mesas_data"
            ).fetchone()
            fecha_actualizacion = str((row_fa or {"f": ""})["f"] or "")

        # Sanity-check de cierre del escrutinio
        mesas_total = totales["mesas"]
        contab = totales["contabilizadas_C"]
        pct_contab = round(100.0 * contab / mesas_total, 4) if mesas_total else 0.0
        escrutinio_cerrado = (mesas_total > 0 and contab == mesas_total)

        return {
            "id_eleccion": 10,
            "filtro": {"ubigeo_prefix": ubigeo_prefix},
            "fecha_actualizacion": fecha_actualizacion,
            "totales": totales,
            "por_estado": por_estado,
            "votos_no_contabilizados": votos_no_contabilizados,
            "geo_top_no_contabilizadas": geo_top_no_c,
            "pct_contabilizadas": pct_contab,
            "escrutinio_cerrado": escrutinio_cerrado,
            "nota_metodologia": (
                "En primera vuelta el 100% de actas suele cerrar como "
                "Contabilizada (no existe escenario JEE pendiente como en SV). "
                "Esta tool sirve principalmente para REFUTAR claims de "
                "'mesas sin contar' / 'votos desaparecidos'."
            ),
        }

    def get_sv_estado_actas(
        self,
        ubigeo_prefix: str | None = None,
        top_geo: int = 10,
    ) -> dict[str, Any]:
        """Resumen de estados de acta SV (C/E/P) y escenario "si JEE acepta todas las E".

        - C = Contabilizada
        - E = Para envío al JEE (observada)
        - P = Pendiente (aún no procesada)

        Args:
            ubigeo_prefix: Filtra por prefijo de ubigeo (2 dígitos = departamento,
                6 dígitos = distrito). None = nacional.
            top_geo: Top departamentos a incluir en `geo_top_jee` (solo cuando no
                se filtra por ubigeo). 0 desactiva el listado.

        Returns dict con:
            - totales: {mesas, contabilizadas_C, para_envio_jee_E, pendientes_P,
                        electores_habiles, votos_emitidos}
            - por_estado: lista [{codigo, descripcion, mesas, electores_habiles,
                                  votos_emitidos, votos_validos}]
            - votos_jee_pendientes: lista [{partido_id, nombre, votos}] (E)
            - escenario_jee_aceptadas: {
                actual: [{partido_id, nombre, votos, pct_validos}],
                con_jee_aceptadas: [{partido_id, nombre, votos, pct_validos}],
                margen_actual: {lider, ventaja, ventaja_pp},
                margen_si_aceptadas: {lider, ventaja, ventaja_pp},
              }
            - geo_top_jee: top departamentos con mesas E (solo si no se filtra)
            - fecha_actualizacion: timestamp ONPE de sv_resumen_nacional
            - filtro: {ubigeo_prefix}
        """
        descripciones_estado = {
            "C": "Contabilizada",
            "E": "Para envío al JEE",
            "P": "Pendiente",
        }

        like = None
        if ubigeo_prefix:
            ubigeo_prefix = str(ubigeo_prefix).strip()
            if ubigeo_prefix:
                like = f"{ubigeo_prefix}%"

        with self._connect() as conn:
            # 1) Conteos por estado
            if like:
                estado_rows = conn.execute(
                    """SELECT codigo_estado_acta, COUNT(*) AS mesas,
                       SUM(electores_habiles) AS electores_habiles,
                       SUM(votos_emitidos) AS votos_emitidos,
                       SUM(votos_validos) AS votos_validos
                       FROM mesas_sv WHERE id_ubigeo LIKE ?
                       GROUP BY codigo_estado_acta""",
                    (like,),
                ).fetchall()
            else:
                estado_rows = conn.execute(
                    """SELECT codigo_estado_acta, COUNT(*) AS mesas,
                       SUM(electores_habiles) AS electores_habiles,
                       SUM(votos_emitidos) AS votos_emitidos,
                       SUM(votos_validos) AS votos_validos
                       FROM mesas_sv
                       GROUP BY codigo_estado_acta"""
                ).fetchall()

            por_estado: list[dict[str, Any]] = []
            por_codigo: dict[str, dict[str, int]] = {}
            for r in estado_rows:
                cod = str(r["codigo_estado_acta"] or "").strip().upper()
                item = {
                    "codigo": cod,
                    "descripcion": descripciones_estado.get(cod, cod),
                    "mesas": int(r["mesas"] or 0),
                    "electores_habiles": int(r["electores_habiles"] or 0),
                    "votos_emitidos": int(r["votos_emitidos"] or 0),
                    "votos_validos": int(r["votos_validos"] or 0),
                }
                por_estado.append(item)
                por_codigo[cod] = item

            totales = {
                "mesas": sum(it["mesas"] for it in por_estado),
                "contabilizadas_C": por_codigo.get("C", {}).get("mesas", 0),
                "para_envio_jee_E": por_codigo.get("E", {}).get("mesas", 0),
                "pendientes_P": por_codigo.get("P", {}).get("mesas", 0),
                "electores_habiles": sum(it["electores_habiles"] for it in por_estado),
                "votos_emitidos": sum(it["votos_emitidos"] for it in por_estado),
            }

            # 2) Votos en mesas E por partido
            if like:
                jee_rows = conn.execute(
                    """SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre,
                       SUM(v.votos) AS total_votos
                       FROM votos_sv v
                       JOIN mesas_sv m ON m.codigo_mesa = v.codigo_mesa
                       LEFT JOIN agrupaciones_sv a ON a.partido_id = v.partido_id
                       WHERE m.codigo_estado_acta = 'E' AND m.id_ubigeo LIKE ?
                       GROUP BY v.partido_id
                       ORDER BY total_votos DESC""",
                    (like,),
                ).fetchall()
            else:
                jee_rows = conn.execute(
                    """SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre,
                       SUM(v.votos) AS total_votos
                       FROM votos_sv v
                       JOIN mesas_sv m ON m.codigo_mesa = v.codigo_mesa
                       LEFT JOIN agrupaciones_sv a ON a.partido_id = v.partido_id
                       WHERE m.codigo_estado_acta = 'E'
                       GROUP BY v.partido_id
                       ORDER BY total_votos DESC"""
                ).fetchall()

            votos_jee_pendientes = [
                {
                    "partido_id": str(r["partido_id"]),
                    "nombre": str(r["nombre"] or ""),
                    "votos": int(r["total_votos"] or 0),
                }
                for r in jee_rows
            ]

            # 3) Escenario actual (C) vs aceptadas (C+E) por partido
            if like:
                escenario_rows = conn.execute(
                    """SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre,
                       SUM(CASE WHEN m.codigo_estado_acta='C' THEN v.votos ELSE 0 END) AS contabilizado,
                       SUM(CASE WHEN m.codigo_estado_acta='E' THEN v.votos ELSE 0 END) AS jee_pendiente
                       FROM votos_sv v
                       JOIN mesas_sv m ON m.codigo_mesa = v.codigo_mesa
                       LEFT JOIN agrupaciones_sv a ON a.partido_id = v.partido_id
                       WHERE m.id_ubigeo LIKE ?
                       GROUP BY v.partido_id""",
                    (like,),
                ).fetchall()
            else:
                escenario_rows = conn.execute(
                    """SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre,
                       SUM(CASE WHEN m.codigo_estado_acta='C' THEN v.votos ELSE 0 END) AS contabilizado,
                       SUM(CASE WHEN m.codigo_estado_acta='E' THEN v.votos ELSE 0 END) AS jee_pendiente
                       FROM votos_sv v
                       JOIN mesas_sv m ON m.codigo_mesa = v.codigo_mesa
                       LEFT JOIN agrupaciones_sv a ON a.partido_id = v.partido_id
                       GROUP BY v.partido_id"""
                ).fetchall()

            # 4) Top departamentos con mesas E (solo nacional)
            geo_top_jee: list[dict[str, Any]] = []
            if not like and top_geo and int(top_geo) > 0:
                # Map de prefijo (2-3 dígitos) → nombre departamento/continente
                dpto_name_map: dict[str, str] = {}
                for r in conn.execute(
                    "SELECT ubigeo, nombre_departamento FROM sv_resumen_cobertura"
                ).fetchall():
                    ub_raw = str(r["ubigeo"] or "")
                    nombre_d = str(r["nombre_departamento"] or "")
                    if not ub_raw:
                        continue
                    # Departamentos peruanos: 6 dígitos, prefijo 2 dígitos. Continentes
                    # exterior (91-95): prefijo 2 dígitos también.
                    dpto_name_map[ub_raw[:2]] = nombre_d

                geo_rows = conn.execute(
                    """SELECT SUBSTR(id_ubigeo,1,2) AS dpto,
                       COUNT(*) AS mesas_E,
                       SUM(electores_habiles) AS electores_E
                       FROM mesas_sv WHERE codigo_estado_acta = 'E'
                       GROUP BY dpto ORDER BY mesas_E DESC LIMIT ?""",
                    (int(top_geo),),
                ).fetchall()

                # Votos E por departamento × partido finalista (8 Keiko, 10 Sánchez)
                votos_dpto_rows = conn.execute(
                    """SELECT SUBSTR(m.id_ubigeo,1,2) AS dpto, v.partido_id,
                       SUM(v.votos) AS votos
                       FROM votos_sv v
                       JOIN mesas_sv m ON m.codigo_mesa = v.codigo_mesa
                       WHERE m.codigo_estado_acta = 'E' AND v.partido_id IN ('8','10')
                       GROUP BY dpto, v.partido_id"""
                ).fetchall()
                votos_dpto: dict[tuple[str, str], int] = {
                    (str(r["dpto"]), str(r["partido_id"])): int(r["votos"] or 0)
                    for r in votos_dpto_rows
                }

                for r in geo_rows:
                    dpto = str(r["dpto"])
                    geo_top_jee.append({
                        "dpto_prefix": dpto,
                        "nombre": dpto_name_map.get(dpto, ""),
                        "mesas_E": int(r["mesas_E"] or 0),
                        "electores_E": int(r["electores_E"] or 0),
                        "votos_E_keiko": votos_dpto.get((dpto, "8"), 0),
                        "votos_E_sanchez": votos_dpto.get((dpto, "10"), 0),
                    })

            # 5) Última fecha de actualización oficial ONPE
            row_fa = conn.execute(
                "SELECT MAX(fecha_actualizacion) AS f FROM sv_resumen_nacional"
            ).fetchone()
            fecha_actualizacion = str((row_fa or {"f": ""})["f"] or "")

        # Construir escenario
        actual: list[dict[str, Any]] = []
        aceptadas: list[dict[str, Any]] = []
        for r in escenario_rows:
            pid = str(r["partido_id"])
            nombre = str(r["nombre"] or "")
            c = int(r["contabilizado"] or 0)
            e = int(r["jee_pendiente"] or 0)
            actual.append({"partido_id": pid, "nombre": nombre, "votos": c})
            aceptadas.append({"partido_id": pid, "nombre": nombre, "votos": c + e})

        # Solo candidatos (partido_id != 80/81/82) cuentan para % válidos
        def _pct_validos(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
            total_validos = sum(it["votos"] for it in rows if it["partido_id"] not in ("80", "81", "82"))
            out = []
            for it in rows:
                if it["partido_id"] in ("80", "81", "82"):
                    pct = 0.0
                else:
                    pct = round(it["votos"] / total_validos * 100, 3) if total_validos > 0 else 0.0
                out.append({**it, "pct_validos": pct})
            return sorted(out, key=lambda x: x["votos"], reverse=True)

        actual_pct = _pct_validos(actual)
        aceptadas_pct = _pct_validos(aceptadas)

        def _margen(rows: list[dict[str, Any]]) -> dict[str, Any]:
            candidatos = [r for r in rows if r["partido_id"] not in ("80", "81", "82")]
            if len(candidatos) < 2:
                return {"lider": None, "ventaja": 0, "ventaja_pp": 0.0}
            primero, segundo = candidatos[0], candidatos[1]
            return {
                "lider": primero["partido_id"],
                "lider_nombre": primero["nombre"],
                "ventaja": primero["votos"] - segundo["votos"],
                "ventaja_pp": round(primero["pct_validos"] - segundo["pct_validos"], 3),
            }

        return {
            "filtro": {"ubigeo_prefix": ubigeo_prefix},
            "fecha_actualizacion": fecha_actualizacion,
            "totales": totales,
            "por_estado": por_estado,
            "votos_jee_pendientes": votos_jee_pendientes,
            "escenario_jee_aceptadas": {
                "actual": actual_pct,
                "con_jee_aceptadas": aceptadas_pct,
                "margen_actual": _margen(actual_pct),
                "margen_si_aceptadas": _margen(aceptadas_pct),
            },
            "geo_top_jee": geo_top_jee,
        }

    def get_comparacion_mesa(self, codigo_mesa: str) -> dict[str, Any]:
        """Compare 1V vs 2V results for the same mesa."""
        with self._connect() as conn:
            m1 = conn.execute("SELECT * FROM mesas_data WHERE codigo_mesa = ?", (codigo_mesa,)).fetchone()
            v1 = conn.execute(
                "SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre, v.votos FROM votos v LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id WHERE v.codigo_mesa = ? ORDER BY v.votos DESC",
                (codigo_mesa,),
            ).fetchall()
            m2 = conn.execute("SELECT * FROM mesas_sv WHERE codigo_mesa = ?", (codigo_mesa,)).fetchone()
            v2 = conn.execute(
                "SELECT v.partido_id, COALESCE(a.nombre,'') AS nombre, v.votos FROM votos_sv v LEFT JOIN agrupaciones_sv a ON a.partido_id = v.partido_id WHERE v.codigo_mesa = ? ORDER BY v.votos DESC",
                (codigo_mesa,),
            ).fetchall()

        result: dict[str, Any] = {
            "codigo_mesa": codigo_mesa,
            "primera_vuelta": None,
            "segunda_vuelta": None,
        }
        if m1:
            result["primera_vuelta"] = {
                "electores_habiles": int(m1["electores_habiles"] or 0),
                "votos_emitidos": int(m1["votos_emitidos"] or 0),
                "votos_validos": int(m1["votos_validos"] or 0),
                "estado_acta": str(m1["estado_acta"] or ""),
                "votos": [{"partido_id": str(r["partido_id"]), "nombre": str(r["nombre"]), "votos": int(r["votos"] or 0)} for r in v1],
            }
        if m2:
            result["segunda_vuelta"] = {
                "electores_habiles": int(m2["electores_habiles"] or 0),
                "votos_emitidos": int(m2["votos_emitidos"] or 0),
                "votos_validos": int(m2["votos_validos"] or 0),
                "codigo_estado_acta": str(m2["codigo_estado_acta"] or ""),
                "votos": [{"partido_id": str(r["partido_id"]), "nombre": str(r["nombre"]), "votos": int(r["votos"] or 0)} for r in v2],
            }
        return result

    def get_comparacion_geo(self, ubigeo_prefix: str) -> dict[str, Any]:
        """Compare 1V vs 2V aggregates for a geo prefix.

        Normalizes dept/province-level ubigeo codes so they match district-level
        ubigeos stored in votos_by_ubigeo_partido and sv_agg_distrito:
          '140000' → '14'  (Lima Metropolitana dept)
          '040000' → '04'  (Arequipa dept)
          '150100' → '1501' (Lima province)
          '140110' → '140110' (specific district, unchanged)
        """
        # Normalize: strip trailing zeros for dept/province codes so LIKE
        # matches district-level ubigeos (e.g., '14%' matches '140110')
        raw = str(ubigeo_prefix or "").strip()
        if len(raw) >= 6 and raw[2:] == "0000":       # dept code like '140000'
            normalized = raw[:2]
        elif len(raw) >= 6 and raw[4:] == "00" and raw[2:4] != "00":  # province like '150100'
            normalized = raw[:4]
        else:
            normalized = raw                           # district or short prefix

        # 1V mesas_data/votos_by_ubigeo_partido stores ubigeos with leading zeros stripped
        # (ONPE internal format): '040101' → '40101', '100101' stays '100101'
        # 2V mesas_sv/sv_resumen_provincias uses RENIEC 6-digit format with leading zeros
        normalized_1v = normalized.lstrip("0") or normalized
        with self._connect() as conn:
            v1_rows = conn.execute(
                """SELECT vup.partido_id, COALESCE(a.nombre,'') AS nombre, SUM(vup.total_votos) AS total_votos
                   FROM votos_by_ubigeo_partido vup
                   LEFT JOIN agrupaciones a ON a.partido_id = vup.partido_id
                   WHERE vup.ubigeo LIKE ?
                   GROUP BY vup.partido_id
                   ORDER BY total_votos DESC""",
                (f"{normalized_1v}%",),
            ).fetchall()
            m1_count = conn.execute(
                "SELECT COUNT(*) AS c FROM mesas_data WHERE ubigeo LIKE ?",
                (f"{normalized_1v}%",),
            ).fetchone()["c"]

            v2_rows = conn.execute(
                """SELECT partido_id, COALESCE(nombre_candidato,'') AS nombre, SUM(votos_validos) AS total_votos
                   FROM sv_resumen_provincias
                   WHERE ubigeo LIKE ?
                   GROUP BY partido_id
                   ORDER BY total_votos DESC""",
                (f"{normalized}%",),
            ).fetchall()
            if not v2_rows:
                v2_rows = conn.execute(
                    """SELECT partido_id, nombre_candidato AS nombre, SUM(votos) AS total_votos
                       FROM sv_agg_distrito
                       WHERE ubigeo LIKE ?
                       GROUP BY partido_id
                       ORDER BY total_votos DESC""",
                    (f"{normalized}%",),
                ).fetchall()
            m2_count = conn.execute(
                "SELECT COUNT(*) AS c FROM mesas_sv WHERE id_ubigeo LIKE ?",
                (f"{normalized}%",),
            ).fetchone()["c"]

        return {
            "ubigeo_prefix": ubigeo_prefix,
            "primera_vuelta": {
                "mesas": int(m1_count),
                "votos": [{"partido_id": str(r["partido_id"]), "nombre": str(r["nombre"]), "total_votos": int(r["total_votos"] or 0)} for r in v1_rows],
            },
            "segunda_vuelta": {
                "mesas": int(m2_count),
                "votos": [{"partido_id": str(r["partido_id"]), "nombre": str(r["nombre"]), "total_votos": int(r["total_votos"] or 0)} for r in v2_rows],
            },
        }

    def seed_transfer_map(self) -> int:
        """Seed voto_transfer_map from TRANSFER_MAP in knowledge_base."""
        from onpe_mcp.knowledge_base import TRANSFER_MAP

        now = self.now_iso()
        count = 0
        with self._connect() as conn:
            for nombre_norm, (pk, ps, pb, fuente) in TRANSFER_MAP.items():
                conn.execute(
                    """INSERT INTO voto_transfer_map (partido_nombre_norm, peso_keiko, peso_sanchez, peso_bn, fuente, loaded_at)
                    VALUES (?,?,?,?,?,?)
                    ON CONFLICT(partido_nombre_norm) DO UPDATE SET
                        peso_keiko=excluded.peso_keiko, peso_sanchez=excluded.peso_sanchez,
                        peso_bn=excluded.peso_bn, fuente=excluded.fuente, loaded_at=excluded.loaded_at""",
                    (nombre_norm, pk, ps, pb, fuente, now),
                )
                count += 1
        return count

    def rebuild_proyeccion_sv(self) -> dict[str, Any]:
        """Build proyeccion_sv_by_ubigeo using float accumulation + NNLS weights."""
        from onpe_mcp.knowledge_base import get_transfer

        now = self.now_iso()
        with self._connect() as conn:
            rows_1v = conn.execute(
                """SELECT m.ubigeo, v.partido_id, COALESCE(a.nombre,'') AS nombre_partido, SUM(v.votos) AS votos
                   FROM votos v
                   JOIN mesas_data m ON m.codigo_mesa = v.codigo_mesa
                   LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                   WHERE v.partido_id NOT IN ('80','81','82')
                   GROUP BY m.ubigeo, v.partido_id""",
            ).fetchall()

        ubigeo_proj: dict[str, dict[str, float]] = {}
        ubigeo_total: dict[str, int] = {}
        for row in rows_1v:
            ubigeo = str(row["ubigeo"] or "")
            partido = str(row["nombre_partido"] or "")
            votos = int(row["votos"] or 0)
            pk, ps, pb, _ = get_transfer(partido)

            if ubigeo not in ubigeo_proj:
                ubigeo_proj[ubigeo] = {"keiko": 0.0, "sanchez": 0.0, "bn": 0.0, "total": 0.0}
                ubigeo_total[ubigeo] = 0

            ubigeo_proj[ubigeo]["keiko"] += votos * pk
            ubigeo_proj[ubigeo]["sanchez"] += votos * ps
            ubigeo_proj[ubigeo]["bn"] += votos * pb
            ubigeo_proj[ubigeo]["total"] += votos
            ubigeo_total[ubigeo] += votos

        proj_rows = []
        for ubigeo, proj in ubigeo_proj.items():
            total = ubigeo_total[ubigeo]
            pk_int = round(proj["keiko"])
            ps_int = round(proj["sanchez"])
            pb_int = round(proj["bn"])
            abs_int = max(0, total - pk_int - ps_int - pb_int)
            proj_rows.append((ubigeo, total, pk_int, ps_int, pb_int, abs_int, now))

        with self._connect() as conn:
            conn.execute("DELETE FROM proyeccion_sv_by_ubigeo")
            conn.executemany(
                """INSERT INTO proyeccion_sv_by_ubigeo
                (ubigeo, votos_1v_total, votos_proyectados_keiko, votos_proyectados_sanchez,
                 votos_proyectados_bn, votos_abstencion_estimada, rebuilt_at)
                VALUES (?,?,?,?,?,?,?)""",
                proj_rows,
            )

        return {"ubigeos_processed": len(proj_rows)}

    def get_proyeccion_sv(self, ubigeo_prefix: str | None = None) -> list[dict[str, Any]]:
        """Get transfer projection. If ubigeo_prefix given, filter. If None, aggregate nationally."""
        with self._connect() as conn:
            if ubigeo_prefix:
                rows = conn.execute(
                    """SELECT ubigeo, votos_1v_total, votos_proyectados_keiko, votos_proyectados_sanchez,
                       votos_proyectados_bn, votos_abstencion_estimada FROM proyeccion_sv_by_ubigeo
                       WHERE ubigeo LIKE ?""",
                    (f"{ubigeo_prefix}%",),
                ).fetchall()
            else:
                rows = conn.execute(
                    """SELECT 'nacional' AS ubigeo,
                       SUM(votos_1v_total) AS votos_1v_total,
                       SUM(votos_proyectados_keiko) AS votos_proyectados_keiko,
                       SUM(votos_proyectados_sanchez) AS votos_proyectados_sanchez,
                       SUM(votos_proyectados_bn) AS votos_proyectados_bn,
                       SUM(votos_abstencion_estimada) AS votos_abstencion_estimada
                       FROM proyeccion_sv_by_ubigeo"""
                ).fetchall()
        return [dict(r) for r in rows]

    def get_proyeccion_sv_by_mesa_prefix(
        self,
        mesa_prefix: str,
        *,
        top_partidos: int = 15,
    ) -> dict[str, Any]:
        """Calcula proyección 1V→2V para el bloque de mesas cuyo código empieza con `mesa_prefix`.

        Aplica TRANSFER_MAP nacional (NNLS calibrado 86K mesas) a los votos 1V del bloque
        y compara con los votos 2V observados. Útil para preguntas como
        "transferencia en mesas 900K".

        Args:
            mesa_prefix: prefijo del código de mesa (ej: '9', '900', '9001'). Normalizado externamente.
            top_partidos: número de partidos emisores 1V a detallar.

        Returns:
            dict con totales 1V/2V, predicción NNLS, error, breakdown por partido y notas.
        """
        from onpe_mcp.knowledge_base import get_transfer

        prefix = str(mesa_prefix or "").strip()
        if not prefix or not prefix.isdigit():
            raise ValueError(f"mesa_prefix debe ser numérico no vacío, got {mesa_prefix!r}")

        like = f"{prefix}%"

        with self._connect() as conn:
            # Totales 1V por partido en el bloque
            rows_1v = conn.execute(
                """SELECT a.partido_id, COALESCE(a.nombre,'') AS nombre, SUM(v.votos) AS total
                   FROM votos v
                   LEFT JOIN agrupaciones a ON a.partido_id = v.partido_id
                   WHERE v.codigo_mesa LIKE ? AND v.votos IS NOT NULL
                   GROUP BY a.partido_id, a.nombre
                   HAVING total > 0
                   ORDER BY total DESC""",
                (like,),
            ).fetchall()

            # Cobertura del bloque en 1V
            cov_1v = conn.execute(
                """SELECT COUNT(*) AS mesas,
                          COALESCE(SUM(electores_habiles),0) AS electores,
                          COALESCE(SUM(votos_emitidos),0) AS emitidos,
                          COALESCE(SUM(votos_validos),0) AS validos
                   FROM mesas_data WHERE codigo_mesa LIKE ?""",
                (like,),
            ).fetchone()

            # Observación 2V real
            obs_2v = conn.execute(
                """SELECT
                     COALESCE(SUM(CASE WHEN partido_id='8'  THEN votos END),0) AS keiko,
                     COALESCE(SUM(CASE WHEN partido_id='10' THEN votos END),0) AS sanchez,
                     COALESCE(SUM(CASE WHEN partido_id='80' THEN votos END),0) AS blancos,
                     COALESCE(SUM(CASE WHEN partido_id='81' THEN votos END),0) AS nulos
                   FROM votos_sv WHERE codigo_mesa LIKE ?""",
                (like,),
            ).fetchone()

            cov_2v = conn.execute(
                """SELECT COUNT(*) AS mesas,
                          COALESCE(SUM(electores_habiles),0) AS electores,
                          COALESCE(SUM(votos_emitidos),0) AS emitidos
                   FROM mesas_sv WHERE codigo_mesa LIKE ?""",
                (like,),
            ).fetchone()

        # Aplicar TRANSFER_MAP nacional partido por partido
        pred_keiko = 0.0
        pred_sanchez = 0.0
        pred_bn = 0.0
        pred_abs = 0.0
        breakdown: list[dict[str, Any]] = []
        total_1v_pool = 0
        for r in rows_1v:
            nombre = str(r["nombre"] or "").strip()
            votos = int(r["total"] or 0)
            total_1v_pool += votos
            pk, ps, pb, fuente = get_transfer(nombre or "DESCONOCIDO")
            pa = max(0.0, 1.0 - pk - ps - pb)
            vk = votos * pk
            vs = votos * ps
            vb = votos * pb
            va = votos * pa
            pred_keiko += vk
            pred_sanchez += vs
            pred_bn += vb
            pred_abs += va
            breakdown.append(
                {
                    "partido_id": str(r["partido_id"] or ""),
                    "nombre": nombre,
                    "votos_1v": votos,
                    "pct_keiko": round(pk * 100, 1),
                    "pct_sanchez": round(ps * 100, 1),
                    "pct_bn": round(pb * 100, 1),
                    "pct_abstencion": round(pa * 100, 1),
                    "pred_keiko": round(vk),
                    "pred_sanchez": round(vs),
                    "pred_bn": round(vb),
                    "pred_abstencion": round(va),
                    "fuente": fuente,
                }
            )

        keiko_obs = int(obs_2v["keiko"] or 0)
        sanchez_obs = int(obs_2v["sanchez"] or 0)
        blancos_obs = int(obs_2v["blancos"] or 0)
        nulos_obs = int(obs_2v["nulos"] or 0)
        pred_k_int = round(pred_keiko)
        pred_s_int = round(pred_sanchez)

        def _err_pct(pred: int, obs: int) -> float | None:
            if obs == 0:
                return None
            return round((pred - obs) / obs * 100, 2)

        breakdown.sort(key=lambda x: -int(x.get("votos_1v") or 0))

        return {
            "mesa_prefix": prefix,
            "primera_vuelta": {
                "mesas": int(cov_1v["mesas"] or 0),
                "electores_habiles": int(cov_1v["electores"] or 0),
                "votos_emitidos": int(cov_1v["emitidos"] or 0),
                "votos_validos": int(cov_1v["validos"] or 0),
                "pool_total_1v": total_1v_pool,
            },
            "segunda_vuelta_observada": {
                "mesas": int(cov_2v["mesas"] or 0),
                "electores_habiles": int(cov_2v["electores"] or 0),
                "votos_emitidos": int(cov_2v["emitidos"] or 0),
                "keiko": keiko_obs,
                "sanchez": sanchez_obs,
                "blancos": blancos_obs,
                "nulos": nulos_obs,
            },
            "proyeccion_nnls_nacional": {
                "keiko": pred_k_int,
                "sanchez": pred_s_int,
                "bn": round(pred_bn),
                "abstencion": round(pred_abs),
                "modelo": "NNLS calibrado 86,124 mesas (TRANSFER_MAP nacional)",
            },
            "error_modelo": {
                "keiko_abs": pred_k_int - keiko_obs,
                "sanchez_abs": pred_s_int - sanchez_obs,
                "keiko_pct": _err_pct(pred_k_int, keiko_obs),
                "sanchez_pct": _err_pct(pred_s_int, sanchez_obs),
                "interpretacion": (
                    "Error positivo: el modelo nacional sobreestima en este bloque. "
                    "Error negativo: el modelo nacional subestima."
                ),
            },
            "breakdown_partidos_top": breakdown[: max(0, int(top_partidos))],
            "notas": [
                "Las tasas de transferencia provienen del modelo NNLS nacional, no recalibradas para este bloque específico.",
                "Para bloques con perfil rural/urbano distinto al promedio nacional (ej: mesas 900K), el modelo puede tener sesgo sistemático.",
                "Compara `proyeccion_nnls_nacional` vs `segunda_vuelta_observada` para cuantificar el sesgo.",
            ],
        }

    def bootstrap_segunda_vuelta(
        self,
        sv_output_dir: Path,
        sv_resumen_dir: Path,
        *,
        force: bool = False,
    ) -> dict[str, Any]:
        """Master bootstrap: loads all SV data. Idempotent (UPSERT). Returns stats."""
        existing = self.total_mesas_sv_local()
        if existing > 0 and not force:
            return {"skipped": True, "mesas_sv": existing}

        result: dict[str, Any] = {"skipped": False}
        result["mesas_sv"] = self.bootstrap_sv_mesas(sv_output_dir)
        result["votos_sv"] = self.bootstrap_sv_votos(sv_output_dir)
        result["agrupaciones_sv"] = self.bootstrap_sv_agrupaciones(sv_output_dir)
        result["ubicaciones_sv"] = self.bootstrap_sv_ubicaciones(sv_output_dir)
        result["reasignados"] = self.bootstrap_sv_reasignados(sv_output_dir)
        result["resumen"] = self.bootstrap_resumen_sv(sv_resumen_dir)
        # rebuild_sv_ctas_levels y seed_transfer_map son opcionales (legacy / cycle tests)
        if hasattr(self, "rebuild_sv_ctas_levels"):
            result["ctas"] = self.rebuild_sv_ctas_levels()
        if hasattr(self, "seed_transfer_map"):
            result["transfer_map_seeded"] = self.seed_transfer_map()
        return result

    def onpe_sv_refresh_from_scraper(
        self,
        sv_output_dir: Path,
        sv_resumen_dir: Path,
    ) -> dict[str, Any]:
        """Incremental refresh: re-import all SV files (UPSERT). Returns change stats."""
        result: dict[str, Any] = {}
        result["mesas_sv"] = self.bootstrap_sv_mesas(sv_output_dir)
        result["votos_sv"] = self.bootstrap_sv_votos(sv_output_dir)
        result["agrupaciones_sv"] = self.bootstrap_sv_agrupaciones(sv_output_dir)
        result["ubicaciones_sv"] = self.bootstrap_sv_ubicaciones(sv_output_dir)
        result["reasignados"] = self.bootstrap_sv_reasignados(sv_output_dir)
        result["resumen"] = self.bootstrap_resumen_sv(sv_resumen_dir)
        if hasattr(self, "rebuild_sv_ctas_levels"):
            result["ctas"] = self.rebuild_sv_ctas_levels()
        return result

    def get_sv_conteo_actual(self) -> dict[str, Any]:
        """Retorna el conteo actual de segunda vuelta 2026 desde el cache hidratado.

        Combina la cifra oficial ONPE (sv_resumen_nacional, solo actas Contabilizadas)
        con el desglose mesa-a-mesa de votos_sv (que incluye actas E con votos crudos
        capturados pero aún no certificadas por ONPE).

        Retorna estructura con tres bloques:
          - oficial: lo que ONPE certificó (snapshot del último refresh).
          - desglose_por_estado: votos K/S por código_estado_acta (C/E/P).
          - proyectado_con_crudo: total combinando C+E (incluye crudo no certificado).

        Si las tablas SV no existen aún (DB sin hidratar), retorna estructura vacía
        con un flag `sv_hidratada=False` para que el llamador pueda dar fallback.
        """
        with self._connect() as conn:
            # Verificar si las tablas SV existen en esta DB
            sv_tables = {
                r[0] for r in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' "
                    "AND name IN ('sv_resumen_nacional','mesas_sv','votos_sv')"
                )
            }
            if not sv_tables:
                return {
                    "sv_hidratada": False,
                    "oficial": {"candidatos": [], "actas_contabilizadas": 0,
                                "total_actas": 0, "pct_contabilizadas": 0.0,
                                "participacion": 0.0, "fecha_actualizacion": None,
                                "loaded_at": None},
                    "desglose_por_estado": [],
                    "proyectado_con_crudo": {"mesas_con_votos": 0, "keiko": 0,
                                              "sanchez": 0, "blancos": 0, "nulos": 0,
                                              "margen_keiko_sanchez": 0,
                                              "pct_keiko": 0.0, "pct_sanchez": 0.0},
                    "cache_hidratado_al": None,
                    "nota_metodologia": (
                        "Esta DB no tiene tablas SV hidratadas. "
                        "Verifica el path ONPE_DATA_DIR o ejecuta el bootstrap SV."
                    ),
                }

            # Bloque 1: oficial certificado
            oficial_rows = conn.execute(
                """SELECT partido_id, nombre_candidato, nombre_agrupacion,
                          votos_validos, pct_votos_validos,
                          contabilizadas, total_actas, participacion_ciudadana,
                          fecha_actualizacion, loaded_at
                   FROM sv_resumen_nacional
                   ORDER BY votos_validos DESC"""
            ).fetchall()

            oficial = {
                "candidatos": [],
                "actas_contabilizadas": 0,
                "total_actas": 0,
                "pct_contabilizadas": 0.0,
                "participacion": 0.0,
                "fecha_actualizacion": None,
                "loaded_at": None,
            }
            for row in oficial_rows:
                nombre = (row["nombre_candidato"] or row["nombre_agrupacion"] or "").strip()
                oficial["candidatos"].append({
                    "partido_id": str(row["partido_id"] or ""),
                    "nombre": nombre,
                    "votos_validos": int(row["votos_validos"] or 0),
                    "pct_votos_validos": float(row["pct_votos_validos"] or 0),
                })
                if not oficial["fecha_actualizacion"]:
                    contab = int(row["contabilizadas"] or 0)
                    tot = int(row["total_actas"] or 0)
                    oficial["actas_contabilizadas"] = contab
                    oficial["total_actas"] = tot
                    oficial["pct_contabilizadas"] = round(contab / tot * 100, 4) if tot else 0.0
                    oficial["participacion"] = float(row["participacion_ciudadana"] or 0)
                    oficial["fecha_actualizacion"] = row["fecha_actualizacion"]
                    oficial["loaded_at"] = row["loaded_at"]

            # Bloque 2: desglose por estado de acta
            # Mesas y electores: agregados directos sobre mesas_sv (sin JOIN para evitar duplicación)
            mesas_por_estado = {
                str(r["codigo_estado_acta"] or ""): {
                    "mesas": int(r["mesas"] or 0),
                    "electores": int(r["electores"] or 0),
                }
                for r in conn.execute(
                    """SELECT codigo_estado_acta,
                              COUNT(*) AS mesas,
                              COALESCE(SUM(electores_habiles), 0) AS electores
                       FROM mesas_sv
                       GROUP BY codigo_estado_acta"""
                )
            }
            # Votos K/S/B/N por estado (con JOIN — agregaciones de votos sí necesitan el join)
            votos_por_estado = {
                str(r["estado"] or ""): r
                for r in conn.execute(
                    """SELECT m.codigo_estado_acta AS estado,
                              COALESCE(SUM(CASE WHEN v.partido_id='8' THEN v.votos END), 0) AS k,
                              COALESCE(SUM(CASE WHEN v.partido_id='10' THEN v.votos END), 0) AS s,
                              COALESCE(SUM(CASE WHEN v.partido_id='80' THEN v.votos END), 0) AS blancos,
                              COALESCE(SUM(CASE WHEN v.partido_id='81' THEN v.votos END), 0) AS nulos
                       FROM mesas_sv m
                       JOIN votos_sv v ON v.codigo_mesa = m.codigo_mesa
                       GROUP BY m.codigo_estado_acta"""
                )
            }

            estado_nombre = {"C": "Contabilizada", "E": "En proceso", "P": "Pendiente"}
            desglose = []
            for est, info in sorted(mesas_por_estado.items()):
                vr = votos_por_estado.get(est)
                k = int(vr["k"] or 0) if vr else 0
                s = int(vr["s"] or 0) if vr else 0
                desglose.append({
                    "codigo_estado": est,
                    "descripcion": estado_nombre.get(est, f"Otro ({est})"),
                    "mesas": info["mesas"],
                    "electores_habiles": info["electores"],
                    "keiko": k,
                    "sanchez": s,
                    "blancos": int(vr["blancos"] or 0) if vr else 0,
                    "nulos": int(vr["nulos"] or 0) if vr else 0,
                    "margen_keiko_sanchez": k - s,
                })

            # Bloque 3: total mesa-a-mesa (incluye C + E)
            total_row = conn.execute(
                """SELECT COUNT(DISTINCT codigo_mesa) AS n_mesas,
                          COALESCE(SUM(CASE WHEN partido_id='8' THEN votos END), 0) AS k,
                          COALESCE(SUM(CASE WHEN partido_id='10' THEN votos END), 0) AS s,
                          COALESCE(SUM(CASE WHEN partido_id='80' THEN votos END), 0) AS blancos,
                          COALESCE(SUM(CASE WHEN partido_id='81' THEN votos END), 0) AS nulos
                   FROM votos_sv"""
            ).fetchone()

            tot_k = int(total_row["k"] or 0) if total_row else 0
            tot_s = int(total_row["s"] or 0) if total_row else 0
            tot_val = tot_k + tot_s
            proyectado = {
                "mesas_con_votos": int(total_row["n_mesas"] or 0) if total_row else 0,
                "keiko": tot_k,
                "sanchez": tot_s,
                "blancos": int(total_row["blancos"] or 0) if total_row else 0,
                "nulos": int(total_row["nulos"] or 0) if total_row else 0,
                "margen_keiko_sanchez": tot_k - tot_s,
                "pct_keiko": round(tot_k / tot_val * 100, 4) if tot_val else 0.0,
                "pct_sanchez": round(tot_s / tot_val * 100, 4) if tot_val else 0.0,
            }

            # Timestamps de las tablas SV
            ts_row = conn.execute(
                """SELECT MAX(fetched_at) AS last_mesas FROM mesas_sv"""
            ).fetchone()
            cache_last = ts_row["last_mesas"] if ts_row else None

        return {
            "oficial": oficial,
            "desglose_por_estado": desglose,
            "proyectado_con_crudo": proyectado,
            "cache_hidratado_al": cache_last,
            "nota_metodologia": (
                "El bloque 'oficial' refleja solo actas Contabilizadas certificadas por ONPE. "
                "El bloque 'proyectado_con_crudo' suma C + E (votos crudos capturados pero "
                "aún no certificados oficialmente). La diferencia entre ambos representa el "
                "voto ya escaneado pero pendiente de cierre formal de acta."
            ),
            "sv_hidratada": True,
        }
